"""Inserted rows inherit particulars merge format from anchor journal block."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

import excel_engine as ee


def test_insert_standardizes_credit_particulars_merge_to_match_debit() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Particulars", "", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Entry", "Cash", None, 50.0, 0.0])
        ws.append([None, "", "Bank", None, 0.0, 50.0])
        ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=4)
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 0,
                "details": 1,
                "particulars": 2,
                "debit": 4,
                "credit": 5,
            },
        }
        plan = ee.GlEditPlan(
            updates=[],
            delete_rows=[],
            insert_rows=[
                {
                    "insert_after": 3,
                    "insert_below_exact": True,
                    "gl_date": "2025-06-02",
                    "account": "NewDr",
                    "description": "new",
                    "debit": 10.0,
                    "credit": 0.0,
                },
                {
                    "insert_after": 3,
                    "insert_below_exact": True,
                    "account": "NewCr",
                    "description": "",
                    "debit": 0.0,
                    "credit": 10.0,
                },
            ],
            swap_rows=[],
        )
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        wb2 = load_workbook(path)
        ws2 = wb2.active
        assert "C5:D5" in {str(m) for m in ws2.merged_cells.ranges}
        assert ws2.cell(5, 3).value == "NewCr"
        wb2.close()
