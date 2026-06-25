"""Particulars merged across C–D: read, write, and post-insert blank checks."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook

import excel_engine as ee


def _aab_layout() -> dict:
    """Ali Al Baker–style: particulars in col C (0-based 2), merged with D on each leg."""
    return {
        "mode": "manual",
        "header_first_row": 1,
        "data_start_row": 2,
        "columns": {
            "date": 0,
            "details": 1,
            "particulars": 2,
            "debit": 4,
            "credit": 5,
        },
    }


def test_particulars_columns_detects_horizontal_merge_cd() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Particulars", "", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Pay", "Cash", None, 100.0, 0.0])
        ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=4)
        wb.save(path)
        wb.close()

        from openpyxl import load_workbook

        wb2 = load_workbook(path)
        ws2 = wb2.active
        cols = ee.particulars_columns_1b(ws2, 2, 3)
        assert cols == [3, 4]
        colmap = ee.gl_column_map_from_layout_dict(_aab_layout())
        label = ee.particulars_label_on_worksheet_row(ws2, colmap, 2)
        assert label == "Cash"
        wb2.close()


def test_insert_rejects_blank_particulars_on_merged_credit_leg() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Particulars", "", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Entry", "Cash", None, 50.0, 0.0])
        ws.append([None, "", "Bank", None, 0.0, 50.0])
        ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=3)
        ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=4)
        ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=4)
        wb.save(path)
        wb.close()

        layout = _aab_layout()
        plan = ee.GlEditPlan(
            updates=[],
            delete_rows=[],
            insert_rows=[
                {
                    "insert_after": 2,
                    "insert_below_exact": True,
                    "gl_date": "2025-06-02",
                    "account": "NewDr",
                    "description": "new",
                    "debit": 10.0,
                    "credit": 0.0,
                },
                {
                    "insert_after": 2,
                    "insert_below_exact": True,
                    "account": "",
                    "description": "",
                    "debit": 0.0,
                    "credit": 10.0,
                },
            ],
            swap_rows=[],
        )
        try:
            ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)
            raised = False
        except ValueError as e:
            raised = True
            assert "particulars" in str(e).lower()
        assert raised


def test_insert_below_exact_skips_workbook_re_resolve() -> None:
    """Compound inserts coerce to block bottom even when insert_below_exact is set."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Account", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Entry A", "Cash", 50.0, 0.0])
        ws.append([None, "", "Bank", 0.0, 50.0])
        ws.append([date(2025, 6, 3), "Entry B", "Rent", 10.0, 0.0])
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 0,
                "details": 1,
                "particulars": 2,
                "debit": 3,
                "credit": 4,
            },
        }
        plan = ee.GlEditPlan(
            updates=[],
            delete_rows=[],
            insert_rows=[
                {
                    "insert_after": 2,
                    "insert_below_exact": True,
                    "gl_date": "2025-06-02",
                    "account": "NewDr",
                    "description": "new",
                    "debit": 5.0,
                    "credit": 0.0,
                },
                {
                    "insert_after": 2,
                    "insert_below_exact": True,
                    "account": "NewCr",
                    "description": "",
                    "debit": 0.0,
                    "credit": 5.0,
                },
            ],
            swap_rows=[],
        )
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)
        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert by_er[4]["account"] == "NewDr"
        assert by_er[5]["account"] == "NewCr"
        assert by_er[6]["account"] == "Rent"


def test_insert_writes_account_into_merged_particulars_band() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Particulars", "", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Entry", "Cash", None, 50.0, 0.0])
        ws.append([None, "", "Bank", None, 0.0, 50.0])
        ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=3)
        ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=4)
        ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=4)
        wb.save(path)
        wb.close()

        layout = _aab_layout()
        plan = ee.GlEditPlan(
            updates=[],
            delete_rows=[],
            insert_rows=[
                {
                    "insert_after": 2,
                    "insert_below_exact": True,
                    "gl_date": "2025-06-02",
                    "account": "NewDr",
                    "description": "new",
                    "debit": 10.0,
                    "credit": 0.0,
                },
                {
                    "insert_after": 2,
                    "insert_below_exact": True,
                    "account": "NewCr",
                    "description": "",
                    "debit": 0.0,
                    "credit": 10.0,
                },
            ],
            swap_rows=[],
        )
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)
        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_acct = {str(r["account"]): r for r in rows}
        assert "NewDr" in by_acct
        assert "NewCr" in by_acct
