"""Faux GL row delete: clear cells, exclude empty rows on read."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

import excel_engine as ee


def _build_merged_journal_sheet(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GL"
    ws.append(["Tr", "Date", "Details", "Account", "Debit", "Credit"])
    ws.append([1, date(2025, 6, 25), "Capital", "Cash", 100_000, 0])
    ws.append(["", None, "", "Owners Equity ( Ali )", 0, 50_000])
    ws.append(["", None, "", "Owners Equity ( Yousef)", 0, 50_000])
    ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
    ws.merge_cells(start_row=2, start_column=3, end_row=4, end_column=3)
    wb.save(path)
    wb.close()


def test_clear_middle_leg_preserves_sibling_values() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        _build_merged_journal_sheet(path)
        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 1,
                "details": 2,
                "particulars": 3,
                "debit": 4,
                "credit": 5,
                "tr_number": 0,
            },
        }
        plan = ee.GlEditPlan(updates=[], delete_rows=[3], insert_rows=[], swap_rows=[])
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        wb = load_workbook(path)
        ws = wb["GL"]
        assert ws.max_row == 4
        assert ws.cell(2, 4).value == "Cash"
        assert float(ws.cell(2, 5).value or 0) == 100_000.0
        assert ws.cell(3, 4).value in (None, "")
        assert ws.cell(4, 4).value == "Owners Equity ( Yousef)"
        assert float(ws.cell(4, 6).value or 0) == 50_000.0
        wb.close()


def test_clear_top_leg_preserves_credit_siblings_when_debit_merged() -> None:
    """Vertical debit merge must not zero credit legs when the anchor row is cleared."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Tr", "Date", "Details", "Account", "Debit", "Credit"])
        ws.append([1, date(2025, 6, 25), "Capital", "Cash", 100_000, 0])
        ws.append(["", None, "", "Owners Equity ( Ali )", 0, 50_000])
        ws.append(["", None, "", "Owners Equity ( Yousef)", 0, 50_000])
        ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
        ws.merge_cells(start_row=2, start_column=3, end_row=4, end_column=3)
        ws.merge_cells(start_row=2, start_column=5, end_row=4, end_column=5)
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 1,
                "details": 2,
                "particulars": 3,
                "debit": 4,
                "credit": 5,
                "tr_number": 0,
            },
        }
        plan = ee.GlEditPlan(updates=[], delete_rows=[2], insert_rows=[], swap_rows=[])
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        wb = load_workbook(path)
        ws = wb["GL"]
        assert ws.cell(2, 4).value in (None, "")
        assert float(ws.cell(3, 6).value or 0) == 50_000.0
        assert float(ws.cell(4, 6).value or 0) == 50_000.0
        wb.close()

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert 2 not in by_er
        assert float(by_er[3]["credit"]) == 50_000.0
        assert float(by_er[4]["credit"]) == 50_000.0


def test_delete_credit_leg_keeps_merged_debit_anchor_in_workbook() -> None:
    """Clearing a credit leg must not blank the shared vertical debit merge anchor."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Tr", "Date", "Details", "Account", "Debit", "Credit"])
        ws.append([1, date(2025, 6, 25), "Capital", "Cash", 100_000, 0])
        ws.append(["", None, "", "Owners Equity ( Ali )", 0, 50_000])
        ws.append(["", None, "", "Owners Equity ( Yousef)", 0, 50_000])
        ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
        ws.merge_cells(start_row=2, start_column=3, end_row=4, end_column=3)
        ws.merge_cells(start_row=2, start_column=5, end_row=4, end_column=5)
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 1,
                "details": 2,
                "particulars": 3,
                "debit": 4,
                "credit": 5,
                "tr_number": 0,
            },
        }
        plan = ee.GlEditPlan(updates=[], delete_rows=[3], insert_rows=[], swap_rows=[])
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        wb = load_workbook(path)
        ws = wb["GL"]
        assert float(ws.cell(2, 5).value or 0) == 100_000.0
        assert float(ws.cell(4, 6).value or 0) == 50_000.0
        wb.close()

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert float(by_er[2]["debit"]) == 100_000.0
        assert float(by_er[4]["credit"]) == 50_000.0


def test_delete_middle_leg_without_date_merge_restores_siblings() -> None:
    """Tr-blank continuation rows must be restored even when date cells are not merged."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Tr", "Date", "Details", "Account", "Debit", "Credit"])
        ws.append([29, date(2025, 8, 10), "Payables", "Cheques Payable", 3458, 0])
        ws.append([None, date(2025, 8, 10), "", "Equipment", 2859, 0])
        ws.append([None, date(2025, 8, 10), "", "Cash", 0, 10119])
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 1,
                "details": 2,
                "particulars": 3,
                "debit": 4,
                "credit": 5,
                "tr_number": 0,
            },
        }
        plan = ee.GlEditPlan(updates=[], delete_rows=[3], insert_rows=[], swap_rows=[])
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert 3 not in by_er
        assert float(by_er[2]["debit"]) == 3458.0
        assert float(by_er[4]["credit"]) == 10119.0


def test_clear_middle_leg_no_date_only_orphan_in_read() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        _build_merged_journal_sheet(path)
        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 1,
                "details": 2,
                "particulars": 3,
                "debit": 4,
                "credit": 5,
                "tr_number": 0,
            },
        }
        plan = ee.GlEditPlan(updates=[], delete_rows=[3], insert_rows=[], swap_rows=[])
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        assert len(rows) == 2
        for r in rows:
            assert str(r.get("account") or "").strip()
            assert float(r.get("debit") or 0) + float(r.get("credit") or 0) > 0
            assert not ee.is_non_posting_gl_row(r)
