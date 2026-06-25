"""Delete targets must match workbook row identity."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

import excel_engine as ee


def test_verify_rejects_wrong_account() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Account", "Debit", "Credit"])
        ws.append([date(2025, 1, 1), "a", "Cash", 1, 0])
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {"date": 0, "details": 1, "particulars": 2, "debit": 3, "credit": 4},
        }
        wb = ee.load_workbook(path)
        ws = wb["GL"]
        _, ws_o, colmap, _, credit_style, _, _ = ee._open_gl_sheet_for_write(
            path, sheet_name="GL", layout=layout
        )
        del ws_o
        err = ee.verify_gl_delete_targets(
            ws,
            colmap,
            [{"excel_row": 2, "account": "Rent", "debit": 0, "credit": 2}],
            credit_style=credit_style,
        )
        wb.close()
        assert err is not None
        assert "Cash" in err


def test_apply_with_checks_deletes_only_matching_rows() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Account", "Debit", "Credit"])
        ws.append([date(2025, 1, 1), "a", "Cash", 1, 0])
        ws.append([date(2025, 1, 2), "b", "Rent", 0, 2])
        ws.append([date(2025, 1, 3), "c", "Fees", 3, 0])
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {"date": 0, "details": 1, "particulars": 2, "debit": 3, "credit": 4},
        }
        plan = ee.GlEditPlan(
            updates=[],
            delete_rows=[2, 3],
            insert_rows=[],
            swap_rows=[],
            delete_row_checks=[
                {"excel_row": 2, "account": "Cash", "debit": 1, "credit": 0},
                {"excel_row": 3, "account": "Rent", "debit": 0, "credit": 2},
            ],
        )
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        from openpyxl import load_workbook

        wb2 = load_workbook(path)
        ws2 = wb2["GL"]
        assert ws2.max_row == 4
        assert ws2.cell(2, 3).value in (None, "")
        assert ws2.cell(3, 3).value in (None, "")
        assert ws2.cell(4, 3).value == "Fees"
        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0)
        assert len(rows) == 1
        assert rows[0].get("account") == "Fees"
        wb2.close()


def test_verify_finds_account_in_spillover_column() -> None:
    """Account one column right of mapped particulars (indented credit lines)."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Particulars", "Account", "Debit", "Credit"])
        ws.append([date(2025, 3, 15), "Payment", None, "Utilities", 0, 125.50])
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 0,
                "details": 1,
                "particulars": 2,
                "debit": 4,
                "credit": 5,
            },
        }
        wb = ee.load_workbook(path)
        ws = wb["GL"]
        _, _ws_o, colmap, _, credit_style, _, _ = ee._open_gl_sheet_for_write(
            path, sheet_name="GL", layout=layout
        )
        err = ee.verify_gl_delete_targets(
            ws,
            colmap,
            [{"excel_row": 2, "account": "Utilities", "debit": 0, "credit": 125.50}],
            credit_style=credit_style,
        )
        wb.close()
        assert err is None
