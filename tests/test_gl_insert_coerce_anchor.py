"""Insert anchor must land below full journal block even when plan says debit leg."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook

import excel_engine as ee


def _layout() -> dict:
    return {
        "mode": "manual",
        "header_first_row": 1,
        "data_start_row": 2,
        "columns": {
            "date": 0,
            "details": 1,
            "particulars": 2,
            "debit": 3,
            "credit": 4,
            "tr_number": None,
        },
    }


def test_coerce_anchor_when_exact_after_debit_leg() -> None:
    """insert_below_exact after row 10 must still insert at 12-13, not 11-12."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Account", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Entry A", "Cash", 50.0, 0.0])
        ws.append([None, "", "", None, None])
        ws.append([None, "", "Bank", 0.0, 50.0])
        ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
        ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
        wb.save(path)
        wb.close()

        layout = _layout()
        plan = ee.GlEditPlan(
            updates=[],
            delete_rows=[],
            insert_rows=[
                {
                    "insert_after": 2,
                    "insert_below_exact": True,
                    "gl_date": "2025-06-02",
                    "account": "NewDr",
                    "description": "new",
                    "debit": 10.0,
                    "credit": 0.0,
                },
                {
                    "insert_after": 2,
                    "insert_below_exact": True,
                    "account": "NewCr",
                    "description": "",
                    "debit": 0.0,
                    "credit": 10.0,
                },
            ],
            swap_rows=[],
        )
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)
        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert by_er[2]["account"] == "Cash"
        assert by_er[4]["account"] == "Bank"
        assert by_er[5]["account"] == "NewDr"
        assert by_er[6]["account"] == "NewCr"


def test_journal_block_bounds_with_spacer_row() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Account", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Entry", "Cash", 50.0, 0.0])
        ws.append([None, "", "", None, None])
        ws.append([None, "", "Bank", 0.0, 50.0])
        ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
        wb.save(path)
        wb.close()

        from openpyxl import load_workbook

        wb2 = load_workbook(path)
        ws2 = wb2.active
        layout = _layout()
        scan = ee._sheet_raw_cell_matrix(ws2, 10, 10)
        colmap = ee.resolve_gl_column_map([tuple(r) for r in scan], layout)
        date_c = ee._openpyxl_col(colmap.date)
        desc_c = ee._openpyxl_col(colmap.details)
        lo, hi = ee._journal_block_bounds_workbook(ws2, colmap, 2, date_c, desc_c)
        assert lo == 2
        assert hi == 4
        assert ee._coerce_insert_anchor_to_block_bottom(ws2, colmap, 2, line_count=2) == 4
        assert ee._coerce_insert_anchor_to_block_bottom(ws2, colmap, 2, line_count=1) == 2
        wb2.close()
