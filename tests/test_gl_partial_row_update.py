"""Partial GL row updates must not zero untouched amount columns."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

import excel_engine as ee


def test_update_account_only_preserves_amounts() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Account", "Debit", "Credit"])
        ws.append([date(2025, 4, 1), "rent", "Rent", 0, 75.5])
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {"date": 0, "details": 1, "particulars": 2, "debit": 3, "credit": 4},
        }
        plan = ee.GlEditPlan(
            updates=[{"excel_row": 2, "account": "Rent (edited)"}],
            delete_rows=[],
            insert_rows=[],
            swap_rows=[],
        )
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        wb2 = load_workbook(path)
        ws2 = wb2["GL"]
        assert ws2.cell(2, 3).value == "Rent (edited)"
        assert float(ws2.cell(2, 4).value or 0) == 0.0
        assert float(ws2.cell(2, 5).value or 0) == 75.5
        wb2.close()