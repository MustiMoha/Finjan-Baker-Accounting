"""Batch faux GL row clear in one apply."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

import excel_engine as ee


def _four_row_gl(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GL"
    ws.append(["Date", "Details", "Account", "Debit", "Credit"])
    ws.append([date(2025, 1, 1), "a", "Cash", 1, 0])
    ws.append([date(2025, 1, 2), "b", "Rent", 0, 2])
    ws.append([date(2025, 1, 3), "c", "Fees", 3, 0])
    ws.append([date(2025, 1, 4), "d", "Tax", 0, 4])
    wb.save(path)
    wb.close()


def test_apply_clears_multiple_rows_in_one_pass() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        _four_row_gl(path)
        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {"date": 0, "details": 1, "particulars": 2, "debit": 3, "credit": 4},
        }
        plan = ee.GlEditPlan(updates=[], delete_rows=[3, 5], insert_rows=[], swap_rows=[])
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        wb = load_workbook(path)
        ws = wb["GL"]
        assert ws.max_row == 5
        assert ws.cell(2, 3).value == "Cash"
        assert ws.cell(3, 3).value in (None, "")
        assert ws.cell(4, 3).value == "Fees"
        assert ws.cell(5, 3).value in (None, "")
        wb.close()

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0)
        assert len(rows) == 2
        accounts = {str(r.get("account")) for r in rows}
        assert accounts == {"Cash", "Fees"}
