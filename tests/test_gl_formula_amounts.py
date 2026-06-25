"""Debit/credit cells with Excel formulas must read cached numeric results, not 0."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

import excel_engine as ee


def test_read_posting_snapshot_uses_formula_cached_value() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Tr", "Date", "Details", "Account", "Debit", "Credit"])
        ws.append([1, date(2025, 6, 25), "Capital", "Cash", 100_000, 0])
        ws.append([None, None, "", "Equity", 0, 50_000])
        wb.save(path)
        wb.close()

        wb_f = load_workbook(path, data_only=False)
        wb_f.active.cell(3, 6).value = "=E2/2"
        wb_f.save(path)
        wb_f.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 1,
                "details": 2,
                "particulars": 3,
                "debit": 4,
                "credit": 5,
                "tr_number": 0,
            },
        }
        wb_f2, ws_f, colmap, _, credit_style, wb_v2, ws_v2 = ee._open_gl_sheet_for_write(
            path, sheet_name="GL", layout=layout
        )
        try:
            snap = ee._read_gl_row_posting_snapshot(
                ws_f, colmap, 3, credit_style=credit_style, ws_values=ws_v2
            )
            assert snap.get("_credit_formula") is True
            assert float(snap["credit"]) == 50_000.0
            assert float(snap["debit"]) == 0.0
        finally:
            ee._close_gl_workbooks(wb_f2, wb_v2)

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert float(by_er[3]["credit"]) == 50_000.0


def test_faux_delete_preserves_sibling_formula_cell() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Tr", "Date", "Details", "Account", "Debit", "Credit"])
        ws.append([1, date(2025, 6, 25), "Capital", "Cash", 100_000, 0])
        ws.append([None, None, "", "Equity (remove)", 0, 50_000])
        ws.append([None, None, "", "Equity (keep)", 0, 25_000])
        wb.save(path)
        wb.close()

        wb_f = load_workbook(path, data_only=False)
        wb_f.active.cell(4, 6).value = "=E2/2"
        wb_f.save(path)
        wb_f.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 1,
                "details": 2,
                "particulars": 3,
                "debit": 4,
                "credit": 5,
                "tr_number": 0,
            },
        }
        plan = ee.GlEditPlan(updates=[], delete_rows=[3], insert_rows=[], swap_rows=[])
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        wb2 = load_workbook(path, data_only=False)
        ws2 = wb2["GL"]
        # Formula is materialized to the cached numeric result so faux delete cannot zero it later.
        assert float(ws2.cell(4, 6).value or 0) == 50_000.0
        assert not str(ws2.cell(4, 6).value or "").startswith("=")
        wb2.close()

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert 3 not in by_er
        assert float(by_er[2]["debit"]) == 100_000.0
        assert float(by_er[4]["credit"]) == 50_000.0


def test_faux_delete_materializes_formula_referencing_deleted_row() -> None:
    """Formulas pointing at the deleted row are frozen before the row is cleared."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Tr", "Date", "Details", "Account", "Debit", "Credit"])
        ws.append([1, date(2025, 1, 5), "Entry A", "Cash", 80_000, 0])
        ws.append([2, date(2025, 1, 6), "Entry B", "Remove me", 20_000, 0])
        ws.append([3, date(2025, 2, 1), "Entry C", "Mirror", 0, 0])
        wb.save(path)
        wb.close()

        wb_f = load_workbook(path, data_only=False)
        ws_f = wb_f["GL"]
        ws_f.cell(4, 6).value = "=E3"
        wb_f.save(path)
        wb_f.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 1,
                "details": 2,
                "particulars": 3,
                "debit": 4,
                "credit": 5,
                "tr_number": 0,
            },
        }
        plan = ee.GlEditPlan(updates=[], delete_rows=[3], insert_rows=[], swap_rows=[])
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert 3 not in by_er
        assert float(by_er[4]["credit"]) == 20_000.0


def test_insert_row_preserves_formula_amounts_on_shifted_lines() -> None:
    """Inserting a GL line must not break debit/credit cells that used Excel formulas."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Tr", "Date", "Details", "Account", "Debit", "Credit"])
        ws.append([1, date(2025, 6, 25), "Capital", "Cash", 100_000, 0])
        ws.append([None, None, "", "Equity", 0, 50_000])
        wb.save(path)
        wb.close()

        wb_f = load_workbook(path, data_only=False)
        wb_f.active.cell(3, 6).value = "=E2/2"
        wb_f.save(path)
        wb_f.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 1,
                "details": 2,
                "particulars": 3,
                "debit": 4,
                "credit": 5,
                "tr_number": 0,
            },
        }
        plan = ee.GlEditPlan(
            updates=[],
            delete_rows=[],
            insert_rows=[
                {
                    "insert_after": 2,
                    "insert_below_exact": True,
                    "gl_date": "2025-06-26",
                    "account": "New line",
                    "description": "inserted",
                    "debit": 0.0,
                    "credit": 0.0,
                }
            ],
            swap_rows=[],
        )
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert float(by_er[2]["debit"]) == 100_000.0
        assert by_er[3]["account"] == "New line"
        assert float(by_er[4]["credit"]) == 50_000.0

        wb_chk = load_workbook(path, data_only=False)
        ws_chk = wb_chk["GL"]
        assert not str(ws_chk.cell(4, 6).value or "").startswith("=")
        assert float(ws_chk.cell(4, 6).value or 0) == 50_000.0
        wb_chk.close()


def test_eval_simple_amount_formula_addition() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Account", "Debit", "Credit"])
        ws.append([date(2025, 1, 1), "x", "A", 10, 0])
        ws.append([date(2025, 1, 1), "y", "B", 25, 0])
        ws.append([date(2025, 1, 1), "z", "C", 0, 0])
        wb.save(path)
        wb.close()

        wb_f = load_workbook(path, data_only=False)
        wb_f.active.cell(4, 4).value = "=D2+D3"
        wb_f.save(path)
        wb_f.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 0,
                "details": 1,
                "particulars": 2,
                "debit": 3,
                "credit": 4,
            },
        }
        wb_f2, ws_f, colmap, _, credit_style, wb_v2, ws_v2 = ee._open_gl_sheet_for_write(
            path, sheet_name="GL", layout=layout
        )
        try:
            snap = ee._read_gl_row_posting_snapshot(
                ws_f, colmap, 4, credit_style=credit_style, ws_values=ws_v2
            )
            assert float(snap["debit"]) == 35.0
        finally:
            ee._close_gl_workbooks(wb_f2, wb_v2)
