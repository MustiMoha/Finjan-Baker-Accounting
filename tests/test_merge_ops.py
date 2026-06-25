"""openpyxl row delete with merged-cell shrink (workbook_editor.merge_ops)."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from workbook_editor.merge_ops import delete_sheet_row, merged_ranges_on_row


def _three_leg_journal(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GL"
    ws.append(["Tr", "Date", "Details", "Account", "Debit", "Credit"])
    ws.append([1, date(2025, 6, 25), "Capital", "Cash", 100_000, 0])
    ws.append(["", None, "", "Owners Equity ( Ali )", 0, 50_000])
    ws.append(["", None, "", "Owners Equity ( Yousef)", 0, 50_000])
    ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
    ws.merge_cells(start_row=2, start_column=3, end_row=4, end_column=3)
    wb.save(path)
    wb.close()


def test_delete_middle_leg_only_one_row_removed() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        _three_leg_journal(path)
        wb = load_workbook(path)
        ws = wb["GL"]
        assert len(merged_ranges_on_row(ws, 3)) >= 1
        delete_sheet_row(ws, 3)
        wb.save(path)
        wb.close()

        wb2 = load_workbook(path)
        ws2 = wb2["GL"]
        assert ws2.max_row == 3
        assert ws2.cell(2, 4).value == "Cash"
        assert float(ws2.cell(2, 5).value or 0) == 100_000.0
        assert ws2.cell(3, 4).value == "Owners Equity ( Yousef)"
        assert float(ws2.cell(3, 6).value or 0) == 50_000.0
        merged = {str(m) for m in ws2.merged_cells.ranges}
        assert "B2:B3" in merged
        assert "C2:C3" in merged
        wb2.close()


def test_delete_top_leg_keeps_other_accounts() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        _three_leg_journal(path)
        wb = load_workbook(path)
        ws = wb["GL"]
        delete_sheet_row(ws, 2)
        wb.save(path)
        wb.close()

        wb2 = load_workbook(path)
        ws2 = wb2["GL"]
        assert ws2.max_row == 3
        accounts = [str(ws2.cell(r, 4).value or "") for r in range(2, 4)]
        assert "Owners Equity ( Ali )" in accounts[0]
        assert "Owners Equity ( Yousef)" in accounts[1]
        wb2.close()
