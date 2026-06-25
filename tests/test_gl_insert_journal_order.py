"""Multi-line GL insert must preserve debit-then-credit row order and account columns."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook

import excel_engine as ee


def _layout_indented_credit() -> dict:
    return {
        "mode": "manual",
        "header_first_row": 1,
        "data_start_row": 2,
        "columns": {
            "date": 0,
            "details": 1,
            "particulars": 2,
            "debit": 4,
            "credit": 5,
            "tr_number": None,
        },
    }


def test_insert_two_line_journal_keeps_debit_above_credit() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Account", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Old", "Cash", 50.0, 0.0])
        ws.append([None, "", "Bank", 0.0, 50.0])
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 0,
                "details": 1,
                "particulars": 2,
                "debit": 3,
                "credit": 4,
                "tr_number": None,
            },
        }
        plan = ee.GlEditPlan(
            updates=[],
            delete_rows=[],
            insert_rows=[
                {
                    "insert_after": 2,
                    "gl_date": "2025-06-02",
                    "account": "New debit acct",
                    "description": "new entry",
                    "debit": 10.0,
                    "credit": 0.0,
                },
                {
                    "insert_after": 2,
                    "account": "New credit acct",
                    "description": "",
                    "debit": 0.0,
                    "credit": 10.0,
                },
            ],
            swap_rows=[],
        )
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        # Insert below row 2 resolves to row 3 (credit leg), so new lines land at 4–5.
        assert float(by_er[4]["debit"]) == 10.0
        assert by_er[4]["account"] == "New debit acct"
        assert float(by_er[5]["credit"]) == 10.0
        assert by_er[5]["account"] == "New credit acct"


def test_update_preserves_indented_credit_account() -> None:
    layout = _layout_indented_credit()
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Particulars", "", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Opening", "Cash", "", 100.0, 0.0])
        ws.append([None, "", "", "Equity", 0.0, 100.0])
        wb.save(path)
        wb.close()

        plan = ee.GlEditPlan(
            updates=[
                {
                    "excel_row": 3,
                    "account": "Equity renamed",
                }
            ],
            delete_rows=[],
            insert_rows=[],
            swap_rows=[],
        )
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)
        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        equity = next(r for r in rows if int(r["_excel_row"]) == 3)
        assert equity["account"] == "Equity renamed"


def test_insert_specs_credit_first_still_debit_above_credit() -> None:
    """Plan rows listed credit-before-debit must still post debit leg on the upper line."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Account", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Old", "Cash", 50.0, 0.0])
        ws.append([None, "", "Bank", 0.0, 50.0])
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 0,
                "details": 1,
                "particulars": 2,
                "debit": 3,
                "credit": 4,
                "tr_number": None,
            },
        }
        plan = ee.GlEditPlan(
            updates=[],
            delete_rows=[],
            insert_rows=[
                {
                    "insert_after": 2,
                    "account": "Credit acct",
                    "description": "",
                    "debit": 0.0,
                    "credit": 10.0,
                },
                {
                    "insert_after": 2,
                    "gl_date": "2025-06-02",
                    "account": "Debit acct",
                    "description": "memo here",
                    "debit": 10.0,
                    "credit": 0.0,
                },
            ],
            swap_rows=[],
        )
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert float(by_er[4]["debit"]) == 10.0
        assert by_er[4]["account"] == "Debit acct"
        assert float(by_er[5]["credit"]) == 10.0
        assert by_er[5]["account"] == "Credit acct"


def test_sort_insert_specs_debit_first() -> None:
    specs = [
        {"credit": 5.0, "debit": 0.0, "account": "C"},
        {"credit": 0.0, "debit": 5.0, "account": "D"},
    ]
    out = ee._sort_insert_specs_debit_first(specs)
    assert [s["account"] for s in out] == ["D", "C"]
