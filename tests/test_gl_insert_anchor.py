"""Insert below a debit leg must not split that journal's credit leg."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook

import excel_engine as ee


def test_resolve_insert_after_debit_leg_uses_block_bottom() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Account", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Entry A", "Cash", 50.0, 0.0])
        ws.append([None, "", "Bank", 0.0, 50.0])
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
        wb.save(path)
        wb.close()

        from openpyxl import load_workbook

        wb2 = load_workbook(path)
        ws2 = wb2.active
        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 0,
                "details": 1,
                "particulars": 2,
                "debit": 3,
                "credit": 4,
                "tr_number": None,
            },
        }
        scan = ee._sheet_raw_cell_matrix(ws2, 10, 10)
        colmap = ee.resolve_gl_column_map([tuple(r) for r in scan], layout)
        assert ee.resolve_gl_insert_after_row(ws2, colmap, 2) == 3
        wb2.close()


def test_insert_below_debit_keeps_existing_credit_with_debit() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Account", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Entry A", "Cash", 50.0, 0.0])
        ws.append([None, "", "Bank", 0.0, 50.0])
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 0,
                "details": 1,
                "particulars": 2,
                "debit": 3,
                "credit": 4,
                "tr_number": None,
            },
        }
        plan = ee.GlEditPlan(
            updates=[],
            delete_rows=[],
            insert_rows=[
                {
                    "insert_after": 2,
                    "gl_date": "2025-06-02",
                    "account": "NewDr",
                    "description": "new",
                    "debit": 10.0,
                    "credit": 0.0,
                },
                {
                    "insert_after": 2,
                    "account": "NewCr",
                    "description": "",
                    "debit": 0.0,
                    "credit": 10.0,
                },
            ],
            swap_rows=[],
        )
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert by_er[2]["account"] == "Cash"
        assert float(by_er[2]["debit"]) == 50.0
        assert by_er[3]["account"] == "Bank"
        assert float(by_er[3]["credit"]) == 50.0
        assert by_er[4]["account"] == "NewDr"
        assert by_er[5]["account"] == "NewCr"


def test_resolve_insert_capped_by_next_visible_editor_row() -> None:
    """Full workbook must not push insert anchor to the last row of a long month slice."""
    import gl_editor as gled

    editor = [
        {"_excel_row": 10, "Date": "2025-06-01", "Account": "Cash", "Debit": 1.0, "Credit": 0.0},
        {"_excel_row": 11, "Date": "", "Account": "Bank", "Debit": 0.0, "Credit": 1.0},
        {"_excel_row": 12, "Date": "2025-06-02", "Account": "Rent", "Debit": 2.0, "Credit": 0.0},
        {"_excel_row": 200, "Date": "2025-06-30", "Account": "End", "Debit": 3.0, "Credit": 0.0},
    ]
    workbook = editor + [
        {"_excel_row": i, "Date": "2025-06-01", "Account": "X", "Debit": 0.0, "Credit": 0.0}
        for i in range(13, 200)
    ]
    assert (
        gled.resolve_insert_after_excel_row(
            editor, 10, include_tr=False, workbook_records=workbook
        )
        == 11
    )


def test_editor_resolve_insert_after_debit() -> None:
    from views.gl_sheet_editor import _insert_compound_below_anchor, _resolve_insert_after_in_editor_records

    records = [
        {"_excel_row": 10, "Date": "2025-01-01", "Account": "Cash", "Debit": 1.0, "Credit": 0.0, "Details": "a"},
        {"_excel_row": 11, "Date": "", "Account": "Bank", "Debit": 0.0, "Credit": 1.0, "Details": ""},
        {"_excel_row": 12, "Date": "2025-01-03", "Account": "Rent", "Debit": 2.0, "Credit": 0.0, "Details": "b"},
    ]
    assert _resolve_insert_after_in_editor_records(records, 10) == 11


def test_editor_resolve_insert_across_excel_row_gap() -> None:
    """Credit leg on row 7 when row 6 is blank on the sheet (not in the grid)."""
    from views.gl_sheet_editor import _resolve_insert_after_in_editor_records

    records = [
        {"_excel_row": 5, "Date": "2025-01-01", "Account": "Cash", "Debit": 1.0, "Credit": 0.0},
        {"_excel_row": 7, "Date": "2025-01-01", "Account": "Bank", "Debit": 0.0, "Credit": 1.0},
    ]
    assert _resolve_insert_after_in_editor_records(records, 5) == 7
    assert _resolve_insert_after_in_editor_records(records, 7) == 7


def test_workbook_insert_below_debit_with_blank_spacer_row() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Account", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Entry A", "Cash", 50.0, 0.0])
        ws.append([None, "", "", None, None])
        ws.append([None, "", "Bank", 0.0, 50.0])
        ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
        ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 0,
                "details": 1,
                "particulars": 2,
                "debit": 3,
                "credit": 4,
                "tr_number": None,
            },
        }
        plan = ee.GlEditPlan(
            updates=[],
            delete_rows=[],
            insert_rows=[
                {
                    "insert_after": 2,
                    "gl_date": "2025-06-02",
                    "account": "NewDr",
                    "description": "new",
                    "debit": 10.0,
                    "credit": 0.0,
                },
                {
                    "insert_after": 2,
                    "account": "NewCr",
                    "description": "",
                    "debit": 0.0,
                    "credit": 10.0,
                },
            ],
            swap_rows=[],
        )
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)
        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert by_er[2]["account"] == "Cash"
        assert by_er[4]["account"] == "Bank"
        assert by_er[5]["account"] == "NewDr"
        assert by_er[6]["account"] == "NewCr"


def test_editor_resolve_insert_skips_pending_rows_and_merged_date_on_credit() -> None:
    from views.gl_sheet_editor import _insert_compound_below_anchor, _resolve_insert_after_in_editor_records

    records = [
        {"_excel_row": 10, "Date": "2025-01-01", "Tr": "1", "Account": "Cash", "Debit": 1.0, "Credit": 0.0},
        {"_excel_row": 0, "After row": 10, "Account": "Stale pending", "Debit": 0.0, "Credit": 0.0},
        {"_excel_row": 11, "Date": "2025-01-01", "Tr": "", "Account": "Bank", "Debit": 0.0, "Credit": 1.0},
        {"_excel_row": 12, "Date": "2025-01-03", "Tr": "2", "Account": "Rent", "Debit": 2.0, "Credit": 0.0},
    ]
    assert _resolve_insert_after_in_editor_records(records, 10) == 11
    new_rows = [
        {"_excel_row": 0, "After row": 10, "Account": "N1", "Debit": 5.0, "Credit": 0.0},
        {"_excel_row": 0, "After row": 10, "Account": "N2", "Debit": 0.0, "Credit": 5.0},
    ]
    out = _insert_compound_below_anchor(
        records, insert_after=10, new_rows=new_rows, resolve_block_end=False
    )
    from views.gl_sheet_editor import _excel_row_int

    assert [str(r.get("Account") or "") for r in out] == [
        "Cash",
        "Stale pending",
        "N1",
        "N2",
        "Bank",
        "Rent",
    ]
    assert [_excel_row_int(r["_excel_row"]) for r in out] == [10, 0, 0, 0, 11, 12]


def test_resolve_insert_after_repeated_tr_on_credit_leg() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Tr", "Date", "Details", "Account", "Debit", "Credit"])
        ws.append([29, date(2025, 6, 1), "A", "Cash", 50.0, 0.0])
        ws.append([29, None, "", "Bank", 0.0, 50.0])
        ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
        wb.save(path)
        wb.close()

        from openpyxl import load_workbook

        wb2 = load_workbook(path)
        ws2 = wb2.active
        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 1,
                "details": 2,
                "particulars": 3,
                "debit": 4,
                "credit": 5,
                "tr_number": 0,
            },
        }
        scan = ee._sheet_raw_cell_matrix(ws2, 10, 10)
        colmap = ee.resolve_gl_column_map([tuple(r) for r in scan], layout)
        assert ee.resolve_gl_insert_after_row(ws2, colmap, 2) == 3
        wb2.close()
