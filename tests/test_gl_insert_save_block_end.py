"""Save plan must insert below journal block end, not between debit/credit legs."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

import excel_engine as ee
import gl_editor as gled
from views.gl_sheet_editor import (
    _editor_rows_from_compound_entry,
    _insert_compound_below_anchor,
    _plan_from_merged,
)


def test_plan_resolves_insert_after_debit_leg_to_block_end() -> None:
    baseline = pd.DataFrame(
        [
            {
                "_excel_row": 10,
                "After row": 0,
                "Date": "2025-06-01",
                "Account": "Cash",
                "Debit": 50.0,
                "Credit": 0.0,
                "Details": "Entry A",
            },
            {
                "_excel_row": 11,
                "After row": 0,
                "Date": "",
                "Account": "Bank",
                "Debit": 0.0,
                "Credit": 50.0,
                "Details": "",
            },
        ]
    )
    new_rows = _editor_rows_from_compound_entry(
        scope_id="save_block",
        insert_after=10,
        posting_date=date(2025, 6, 2),
        details="new",
        journal_lines=[
            {"account": "NewDr", "debit": "10.00", "credit": "0.00"},
            {"account": "NewCr", "debit": "0.00", "credit": "10.00"},
        ],
        include_tr=False,
    )
    edited_recs = _insert_compound_below_anchor(
        baseline.to_dict(orient="records"),
        insert_after=10,
        new_rows=new_rows,
        include_tr=False,
    )
    plan = gled.plan_from_editor_diff(baseline, pd.DataFrame(edited_recs), include_tr=False)
    assert len(plan.insert_rows) == 2
    assert plan.insert_rows[0]["insert_after"] == 11
    assert plan.insert_rows[1]["insert_after"] == 11
    assert plan.insert_rows[0].get("insert_below_exact") is True


def test_save_below_debit_keeps_credit_with_debit_on_workbook() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Date", "Details", "Particulars", "", "Debit", "Credit"])
        ws.append([date(2025, 6, 1), "Entry A", "Cash", None, 50.0, 0.0])
        ws.append([None, "", "Bank", None, 0.0, 50.0])
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
        ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=4)
        ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=4)
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 0,
                "details": 1,
                "particulars": 2,
                "debit": 4,
                "credit": 5,
            },
        }
        baseline = pd.DataFrame(
            [
                {
                    "_excel_row": 2,
                    "After row": 0,
                    "Date": "2025-06-01",
                    "Account": "Cash",
                    "Debit": 50.0,
                    "Credit": 0.0,
                    "Details": "Entry A",
                },
                {
                    "_excel_row": 3,
                    "After row": 0,
                    "Date": "",
                    "Account": "Bank",
                    "Debit": 0.0,
                    "Credit": 50.0,
                    "Details": "",
                },
            ]
        )
        new_rows = _editor_rows_from_compound_entry(
            scope_id="wb_block",
            insert_after=2,
            posting_date=date(2025, 6, 2),
            details="new",
            journal_lines=[
                {"account": "NewDr", "debit": "10.00", "credit": "0.00"},
                {"account": "NewCr", "debit": "0.00", "credit": "10.00"},
            ],
            include_tr=False,
        )
        merged = pd.DataFrame(
            _insert_compound_below_anchor(
                baseline.to_dict(orient="records"),
                insert_after=2,
                new_rows=new_rows,
                include_tr=False,
            )
        )
        plan = _plan_from_merged(baseline, merged, include_tr=False, scope_id="wb_block")
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert by_er[2]["account"] == "Cash"
        assert by_er[3]["account"] == "Bank"
        assert by_er[4]["account"] == "NewDr"
        assert by_er[5]["account"] == "NewCr"
