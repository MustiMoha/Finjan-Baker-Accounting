"""Writes must survive orphan MergedCell slaves (missing merge metadata)."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

import excel_engine as ee


def test_merge_aware_write_orphan_merged_slave() -> None:
    wb = Workbook()
    ws = wb.active
    ws["D3"] = 100
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=6)
    ws.merged_cells.remove(next(iter(ws.merged_cells.ranges)))

    ee._set_sheet_cell_value_merge_aware(ws, 3, 5, 999)
    assert ws.cell(3, 4).value == 999


def test_insert_after_orphan_style_horizontal_amount_merge() -> None:
    """Simulate insert_rows leaving slave cells without merge registry entries."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Tr", "Date", "Details", "Account", "Debit", "Credit"])
        ws.append([1, date(2025, 1, 1), "Memo", "Cash", 100, 0])
        ws.append([None, None, "", "Rent Expense", 0, 100])
        ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=6)
        wb.save(path)
        wb.close()

        wb2 = load_workbook(path)
        ws2 = wb2.active
        ws2.insert_rows(4, amount=2)
        ws2.merged_cells.remove(next(iter(ws2.merged_cells.ranges)))
        wb2.save(path)
        wb2.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 1,
                "details": 2,
                "particulars": 3,
                "debit": 4,
                "credit": 5,
                "tr_number": 0,
            },
        }
        plan = ee.GlEditPlan(
            updates=[],
            delete_rows=[],
            insert_rows=[
                {
                    "insert_after": 3,
                    "insert_below_exact": True,
                    "gl_date": "2025-01-02",
                    "account": "NewDr",
                    "description": "new",
                    "debit": 10.0,
                    "credit": 0.0,
                },
                {
                    "insert_after": 3,
                    "insert_below_exact": True,
                    "account": "NewCr",
                    "description": "",
                    "debit": 0.0,
                    "credit": 10.0,
                },
            ],
            swap_rows=[],
        )
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)
