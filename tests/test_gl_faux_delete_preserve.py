"""Faux delete must not zero unrelated postings on re-read."""

from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

import excel_engine as ee


def test_stabilize_restores_wiped_amounts_for_same_excel_row() -> None:
    prior = [
        {
            "_excel_row": 10,
            "account": "Cash",
            "debit": 100.0,
            "credit": 0.0,
            "gl_date": "2025-01-01",
            "description": "x",
        },
        {
            "_excel_row": 11,
            "account": "Rent",
            "debit": 0.0,
            "credit": 50.0,
            "gl_date": "2025-01-01",
            "description": "y",
        },
    ]
    parsed = [
        {
            "_excel_row": 10,
            "account": "Cash",
            "debit": 0.0,
            "credit": 0.0,
            "gl_date": "2025-01-01",
            "description": "x",
        },
        {
            "_excel_row": 11,
            "account": "Rent",
            "debit": 0.0,
            "credit": 0.0,
            "gl_date": "2025-01-01",
            "description": "y",
        },
    ]
    out = ee.stabilize_gl_records_after_faux_delete(prior, parsed, {12})
    by_er = {int(r["_excel_row"]): r for r in out}
    assert float(by_er[10]["debit"]) == 100.0
    assert float(by_er[11]["credit"]) == 50.0


def test_faux_delete_row_with_horizontal_account_merge() -> None:
    """Account cells merged horizontally must clear without MergedCell write errors."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Tr", "Date", "Details", "Account", "Debit", "Credit"])
        ws.append([1, date(2025, 1, 1), "Memo", "Cash", 100, 0])
        ws.append([None, None, "", "Rent Expense", 0, 100])
        ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=6)
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 1,
                "details": 2,
                "particulars": 3,
                "debit": 4,
                "credit": 5,
                "tr_number": 0,
            },
        }
        plan = ee.GlEditPlan(updates=[], delete_rows=[3], insert_rows=[], swap_rows=[])
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)

        rows = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        by_er = {int(r["_excel_row"]): r for r in rows}
        assert 3 not in by_er
        assert float(by_er[2]["debit"]) == 100.0


def test_column_scoped_faux_clear_keeps_sibling_credits_and_date_merge() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "gl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "GL"
        ws.append(["Tr", "Date", "Details", "Account", "Debit", "Credit"])
        ws.append([1, date(2025, 6, 25), "Capital", "Cash", 100_000, 0])
        ws.append(["", None, "", "Owners Equity ( Ali )", 0, 50_000])
        ws.append(["", None, "", "Owners Equity ( Yousef)", 0, 50_000])
        ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
        ws.merge_cells(start_row=2, start_column=3, end_row=4, end_column=3)
        wb.save(path)
        wb.close()

        layout = {
            "mode": "manual",
            "header_first_row": 1,
            "data_start_row": 2,
            "columns": {
                "date": 1,
                "details": 2,
                "particulars": 3,
                "debit": 4,
                "credit": 5,
                "tr_number": 0,
            },
        }
        before = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        plan = ee.GlEditPlan(updates=[], delete_rows=[3], insert_rows=[], swap_rows=[])
        ee.apply_gl_edit_plan(path, plan, sheet_name="GL", layout=layout)
        after = ee.read_gl_sheet_rows(path, sheet_name="GL", layout=layout, tail=0, keep_excel_row=True)
        stabilized = ee.stabilize_gl_records_after_faux_delete(before, after, {3})

        wb2 = load_workbook(path)
        ws2 = wb2["GL"]
        assert ws2.cell(2, 2).value is not None
        assert float(ws2.cell(2, 5).value or 0) == 100_000.0
        assert float(ws2.cell(4, 6).value or 0) == 50_000.0
        wb2.close()

        by_er = {int(r["_excel_row"]): r for r in stabilized}
        assert 3 not in by_er
        assert float(by_er[2]["debit"]) == 100_000.0
        assert float(by_er[4]["credit"]) == 50_000.0
