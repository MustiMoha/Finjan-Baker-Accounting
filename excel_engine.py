"""Append double-entry GL lines to the 'GL' sheet with openpyxl (macros preserved)."""

from __future__ import annotations

from collections import deque
from copy import copy as _copy_style
from dataclasses import dataclass, replace
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import io
import math
import re
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from workbook_editor.merge_ops import insert_sheet_rows

import gl_analytics as gla


# Column layout on 'GL' sheet (customize to match your workbook template)
# A: Date | B: Description | C: Account | D: Debit | E: Credit
GL_SHEET_NAME_DEFAULT = "GL"
COL_DATE = 1
COL_DESCRIPTION = 2
COL_ACCOUNT = 3
COL_DEBIT = 4
COL_CREDIT = 5

# Approvals post only to OpenXML workbooks (openpyxl).
APPENDABLE_EXTENSIONS = frozenset({".xlsx", ".xlsm"})

DEFAULT_CURRENCY_ISO = "USD"

# Largest physical worksheet height we scan in one shot (balances complete statements vs workbook size).
_MAX_GL_PHYSICAL_ROWS_SOFT_CAP = 200_000

_ISO_TOKEN_RE = re.compile(r"\b(?P<iso>[A-Z]{3})\b")

# Merged rows between headers and GL lines (e.g. full-width "July").
_GL_MONTH_NAMES = (
    "january|february|march|april|may|june|july|august|september|october|november|december"
)
_GL_MONTH_ABBR = r"jan\.?|feb\.?|mar\.?|apr\.?|may|jun\.?|jul\.?|aug\.?|sep\.?|sept\.?|oct\.?|nov\.?|dec\.?"
_GL_MONTH_UNIT = rf"(?:{_GL_MONTH_NAMES}|{_GL_MONTH_ABBR})"
_GL_MONTH_BANNER_ROW_FULL = re.compile(
    rf"^{_GL_MONTH_UNIT}(?:\s+\d{{4}})?(?:\s+{_GL_MONTH_UNIT}(?:\s+\d{{4}})?)*$",
    re.I,
)


def currency_iso_from_excel_number_format(fmt: Optional[Any]) -> str:
    """
    Map Excel ``cell.number_format`` to ISO 4217 (lightweight heuristics).
    Unknown patterns default to ``USD``.
    """
    if fmt is None or fmt == "":
        return DEFAULT_CURRENCY_ISO
    s = str(fmt)
    su = s.upper()
    # Qatar — Excel often uses locale 634 or "ر.ق" / QAR code; must run before generic '$' → USD.
    if "QAR" in su:
        return "QAR"
    if "ر.ق" in s or ("\u0631" in s and "\u0642" in s):
        return "QAR"
    if re.search(r"\[[^\]]*-634\]", s):
        return "QAR"
    if "€" in s:
        return "EUR"
    if "£" in s:
        return "GBP"
    if "¥" in s or "[$¥" in s:
        return "JPY"
    if "₹" in s or "\u20b9" in s:
        return "INR"
    if "CAD" in su or "C$" in s:
        return "CAD"
    if "AUD" in su or "A$" in s:
        return "AUD"
    if "CHF" in su:
        return "CHF"
    if "CNY" in su or "RMB" in su or "元" in s:
        return "CNY"
    if "$" in s or "[$" in s:
        return "USD"
    return DEFAULT_CURRENCY_ISO


def _pick_amount_number_format(fmt_debit: str, fmt_credit: str, deb: float, cred: float) -> str:
    """Prefer the format on the non-zero amount column (debit vs credit)."""
    if deb >= cred and deb > 1e-9:
        return fmt_debit or fmt_credit
    if cred > 1e-9:
        return fmt_credit or fmt_debit
    return fmt_debit or fmt_credit


def resolve_row_currency_from_formats(
    fmt_debit: str,
    fmt_credit: str,
    deb: float,
    cred: float,
    currency_column_hint: Optional[str],
) -> str:
    """Combine number formats with optional currency-column text."""
    fp = _pick_amount_number_format(fmt_debit or "", fmt_credit or "", deb, cred)
    from_fmt = currency_iso_from_excel_number_format(fp)

    col_iso: Optional[str] = None
    if currency_column_hint:
        col_iso = normalize_currency_hint_cell(currency_column_hint)
        if not col_iso:
            raw = str(currency_column_hint).strip().upper()
            if len(raw) >= 3 and raw[:3].isalpha():
                col_iso = raw[:3]
        if col_iso:
            col_iso = col_iso[:3]

    if col_iso and len(col_iso) == 3:
        # Explicit column ISO wins over plain "$" / default USD formatting (common for multi-currency books).
        if from_fmt == DEFAULT_CURRENCY_ISO and col_iso != DEFAULT_CURRENCY_ISO:
            return col_iso

    if from_fmt != DEFAULT_CURRENCY_ISO:
        return from_fmt
    if col_iso and len(col_iso) == 3:
        return col_iso
    return from_fmt


def excel_money_number_format_for_iso(iso: str) -> str:
    """Excel custom number format so reopened files keep symbols for sensing."""
    u = (iso or DEFAULT_CURRENCY_ISO).strip().upper()[:3]
    if u == "EUR":
        return '#,##0.00" €"'
    if u == "GBP":
        return '"£"#,##0.00'
    if u == "JPY":
        return '"¥"#,##0'
    if u == "INR":
        return "[$₹]#,##0.00"
    if u == "CAD":
        return '"C$"#,##0.00'
    if u == "AUD":
        return '"A$"#,##0.00'
    if u == "QAR":
        return "[$ر.ق-634]#,##0.00"
    return '"$"#,##0.00'


def _amount_number_format_for_posting(template_nf: Optional[str], posting_iso: str) -> str:
    """
    Prefer the template's amount style when it already matches the dashboard posting currency;
    otherwise apply ``excel_money_number_format_for_iso(posting_iso)``.

    Many masters use USD on the credit column only; debits then follow posting currency via the
    fallback path, so we infer currency from ``template_nf`` and replace when it disagrees.
    """
    fb = excel_money_number_format_for_iso(posting_iso)
    t = (template_nf or "").strip()
    if not t:
        return fb
    inferred = currency_iso_from_excel_number_format(t)
    post = (posting_iso or DEFAULT_CURRENCY_ISO).strip().upper()[:3]
    if len(post) != 3:
        post = DEFAULT_CURRENCY_ISO
    if inferred != post:
        return fb
    return t


def _enrich_record_currency_from_workbook_formats(
    rec: dict[str, Any],
    ws,
    excel_row_1based: int,
    colmap: GlColumnMap,
) -> None:
    deb_c = _openpyxl_col(colmap.debit)
    cred_c = _openpyxl_col(colmap.credit)
    fd = str(ws.cell(row=excel_row_1based, column=deb_c).number_format or "")
    fc = str(ws.cell(row=excel_row_1based, column=cred_c).number_format or "")
    d_amt = row_amount(rec.get("debit"))
    c_amt = row_amount(rec.get("credit"))
    col_hint: Optional[str] = None
    if colmap.currency is not None:
        cv = ws.cell(row=excel_row_1based, column=_openpyxl_col(colmap.currency)).value
        if cv is not None:
            col_hint = str(cv).strip()
    oc = resolve_row_currency_from_formats(fd, fc, d_amt, c_amt, col_hint)
    pre = str(rec.get("currency_iso") or "").strip().upper()[:3]
    if oc == DEFAULT_CURRENCY_ISO and len(pre) == 3 and pre != DEFAULT_CURRENCY_ISO:
        oc = pre
    rec["original_currency"] = oc
    rec["currency_iso"] = oc
    rec["original_amount"] = float(d_amt + c_amt)


def _ensure_original_currency_fallback(rec: dict[str, Any]) -> None:
    """Populate original_* from parsed amounts when Excel formats are unavailable."""
    d_amt = row_amount(rec.get("debit"))
    c_amt = row_amount(rec.get("credit"))
    oc = str(rec.get("currency_iso") or DEFAULT_CURRENCY_ISO).strip().upper()[:3]
    if len(oc) != 3:
        oc = DEFAULT_CURRENCY_ISO
    rec.setdefault("original_currency", oc)
    rec["currency_iso"] = rec.get("original_currency", oc)
    rec["original_amount"] = float(d_amt + c_amt)


@dataclass
class GlColumnMap:
    """0-based column indices for reading a journal / GL sheet."""

    date: int = 0
    particulars: int = 2  # account / ledger line (was col C in legacy template)
    debit: int = 3
    credit: int = 4
    details: int = 1  # narrative; legacy col B (description)
    currency: Optional[int] = None  # optional currency / ISO column
    tr_number: Optional[int] = None  # transaction / batch number (Tr. No.); blank rows continue prior entry
    data_start_row: int = 1  # 1-based first data row in Excel


def _norm_header_cell(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _classify_header_text(s: str) -> set[str]:
    """Map header label text to logical roles (a cell can hint multiple)."""
    roles: set[str] = set()
    if not s:
        return roles
    # Date
    if any(
        kw in s
        for kw in (
            "posting date",
            "txn date",
            "trans date",
            "gl date",
            "journal date",
            "transaction date",
            "value date",
            "doc date",
        )
    ):
        roles.add("date")
    elif s in ("date", "dt") or (s.endswith(" date") and "update" not in s):
        roles.add("date")
    # Particulars / account (ledger line)
    if "particular" in s and "detail" not in s:
        roles.add("particulars")
    if any(k in s for k in ("account", "ledger", "a/c", "acct", "gl code", "account code")):
        roles.add("particulars")
    # Description / narrative (separate from particulars when possible)
    if any(
        k in s
        for k in (
            "description",
            "narrative",
            "memo",
            "remark",
            "reference",
            "ref ",
            " ref",
            "comment",
            "notes",
            "details",
            "explanation",
        )
    ):
        roles.add("details")
    # Debit / credit
    if "debit" in s and "credit" not in s:
        roles.add("debit")
    if "credit" in s and "debit" not in s:
        roles.add("credit")
    if s in ("dr", "d/r", "d r"):
        roles.add("debit")
    if s in ("cr", "c/r", "c r"):
        roles.add("credit")
    if re.match(r"^dr[\s./-]", s) or s.startswith("debit "):
        roles.add("debit")
    if re.match(r"^cr[\s./-]", s) or s.startswith("credit "):
        roles.add("credit")
    for phrase in ("amount dr", "amt dr", "dr amount", "debit amount", "deb amt", "debits"):
        if phrase in s:
            roles.add("debit")
    for phrase in ("amount cr", "amt cr", "cr amount", "credit amount", "cr amt", "credits"):
        if phrase in s:
            roles.add("credit")
    if any(
        tok in s
        for tok in (
            "tr. no",
            "tr no",
            "tr#",
            "transaction no",
            "transaction number",
            "txn no",
            "txn number",
            "trans no",
            "jv no",
            "entry no",
            "doc no",
            "voucher no",
        )
    ):
        roles.add("tr_number")
    if re.match(r"^tr\.?\s*no\.?$", s) or s in ("tr no.", "trans. no."):
        roles.add("tr_number")
    if any(
        k in s
        for k in (
            "currency",
            "curr.",
            " curr",
            "ccy",
            "fx",
            "fx rate",
            "txn cur",
            "trans cur",
        )
    ):
        roles.add("currency")
    return roles


def detect_gl_column_map(header_matrix: list[tuple[Any, ...]], *, max_header_rows: int = 250) -> GlColumnMap:
    """
    Scan header rows for journal keywords and infer column roles.

    ``header_matrix`` is only the configured header segment (titles above the first data row),
    not the whole sheet.
    """
    best: dict[str, tuple[int, int]] = {}  # role -> (score, col_index)
    last_header_row_0based = -1

    nrows_cap = max(15, min(int(max_header_rows), 500))
    nrows = min(len(header_matrix), nrows_cap)
    ncols = 0
    for r in range(nrows):
        row = header_matrix[r]
        if _tuple_row_is_month_section_banner(tuple(row)):
            continue
        ncols = max(ncols, len(row))
        score_base = max(10, (nrows - r) * 10)
        for c, cell in enumerate(row):
            s = _norm_header_cell(cell)
            if not s:
                continue
            if _try_float(cell) is not None and not any(ch.isalpha() for ch in s):
                continue
            for role in _classify_header_text(s):
                sc = score_base + len(s)
                prev = best.get(role)
                if prev is None or sc > prev[0]:
                    best[role] = (sc, c)
                    last_header_row_0based = max(last_header_row_0based, r)

    m = GlColumnMap()
    if "date" in best:
        m.date = best["date"][1]
    if "particulars" in best:
        m.particulars = best["particulars"][1]
    if "debit" in best:
        m.debit = best["debit"][1]
    if "credit" in best:
        m.credit = best["credit"][1]
    if "details" in best:
        m.details = best["details"][1]
    if "currency" in best:
        m.currency = best["currency"][1]
    if "tr_number" in best:
        m.tr_number = best["tr_number"][1]

    # If particulars and details landed on same column, try to find a separate details column
    used = {m.date, m.particulars, m.debit, m.credit}
    if m.currency is not None:
        used.add(m.currency)
    if m.tr_number is not None:
        used.add(m.tr_number)
    if m.details == m.particulars:
        for c in range(ncols):
            if c not in used:
                m.details = c
                break

    if "particulars" not in best:
        used2 = {m.date, m.debit, m.credit, m.details}
        if m.currency is not None:
            used2.add(m.currency)
        if m.tr_number is not None:
            used2.add(m.tr_number)
        for c in range(ncols):
            if c not in used2:
                m.particulars = c
                break

    if m.debit == m.credit and ncols > 1:
        m.credit = min(m.debit + 1, ncols - 1)

    m.data_start_row = max(1, last_header_row_0based + 2)  # 1-based Excel row after header block
    if not best:
        m.data_start_row = 1
    return m


def _matrix_width(matrix: list[tuple[Any, ...]]) -> int:
    if not matrix:
        return 0
    return min(45, max(len(r) for r in matrix))


def gl_column_map_from_layout_dict(layout: dict[str, Any]) -> GlColumnMap:
    """Build GlColumnMap from saved Configuration JSON (manual mode)."""
    cols = layout.get("columns") if isinstance(layout.get("columns"), dict) else {}

    def _ix(key: str, default: int) -> int:
        v = cols.get(key)
        if v is None or str(v).strip() == "":
            return default
        return int(v)

    m = GlColumnMap()
    m.date = _ix("date", m.date)
    m.details = _ix("details", m.details)
    m.particulars = _ix("particulars", m.particulars)
    m.debit = _ix("debit", m.debit)
    m.credit = _ix("credit", m.credit)
    cur = cols.get("currency")
    if cur is not None and str(cur).strip() != "":
        m.currency = int(cur)
    else:
        m.currency = None
    trn = cols.get("tr_number")
    m.tr_number = None
    if trn is not None and str(trn).strip() != "":
        try:
            tri = int(trn)
            if tri >= 0:
                m.tr_number = tri
        except (TypeError, ValueError):
            pass
    h_first_excel = max(1, int(layout.get("header_first_row") or 1))
    dsr = layout.get("data_start_row")
    if dsr is not None:
        try:
            m.data_start_row = max(h_first_excel + 1, int(dsr))
        except (TypeError, ValueError):
            pass
    elif h_first_excel > 1:
        m.data_start_row = max(m.data_start_row, h_first_excel + 1)
    return m


def _score_dc_columns(rows: list[tuple[Any, ...]], c_dr: int, c_cr: int) -> float:
    if c_dr == c_cr:
        return -1e18
    both = dr_only = cr_only = 0
    for row in rows:
        d = abs(_try_float(_cell(row, c_dr)) or 0.0)
        c = abs(_try_float(_cell(row, c_cr)) or 0.0)
        if d > 1e-9 and c > 1e-9:
            both += 1
        elif d > 1e-9:
            dr_only += 1
        elif c > 1e-9:
            cr_only += 1
    denom = dr_only + cr_only + 1
    if both > max(3, int(denom * 0.45)):
        return -1e12 + dr_only + cr_only
    return float(dr_only + cr_only - 4.0 * both)


def refine_debit_credit_columns_from_data(
    data_rows: list[tuple[Any, ...]],
    baseline: GlColumnMap,
    ncols: int,
) -> GlColumnMap:
    """
    Disambiguate debit vs credit columns using journal-shaped data (mostly one-sided amounts).
    Tries baseline order, swapped order, and small column shifts.
    """
    if not data_rows or ncols < 2:
        return baseline

    candidates: list[tuple[int, int]] = []
    bd, bc = baseline.debit, baseline.credit
    candidates.append((bd, bc))
    candidates.append((bc, bd))
    for delta in (-1, 1):
        tbd, tbc = bd + delta, bc + delta
        if 0 <= tbd < ncols and 0 <= tbc < ncols and tbd != tbc:
            candidates.append((tbd, tbc))
    seen: set[tuple[int, int]] = set()
    uniq: list[tuple[int, int]] = []
    for p in candidates:
        if p not in seen and p[0] != p[1]:
            seen.add(p)
            uniq.append(p)

    best_pair = (bd, bc)
    best_sc = -1e18
    for dc, cc in uniq:
        sc = _score_dc_columns(data_rows, dc, cc)
        if sc > best_sc:
            best_sc = sc
            best_pair = (dc, cc)

    if best_pair == (bd, bc) or best_sc <= -1e11:
        return baseline
    return replace(baseline, debit=best_pair[0], credit=best_pair[1])


def resolve_gl_column_map(
    matrix: list[tuple[Any, ...]],
    layout: Optional[dict[str, Any]] = None,
    *,
    header_scan_rows: int = 40,
) -> GlColumnMap:
    """
    Combine optional saved layout (manual) with header detection and data refinement (auto).

    When ``data_start_row`` (and optionally ``header_first_row``) are saved, auto-detection scans
    only that header band — not the arbitrary top rows of the sheet.
    """
    layout = layout or {}
    mode = str(layout.get("mode") or "auto").strip().lower()
    h_first_excel = max(1, int(layout.get("header_first_row") or 1))

    ds_excel_opt: Optional[int]
    ds_raw = layout.get("data_start_row")
    if ds_raw is not None:
        try:
            ds_excel_opt = max(h_first_excel + 1, int(ds_raw))
        except (TypeError, ValueError):
            ds_excel_opt = None
    else:
        ds_excel_opt = None

    hdr_fallback = max(25, int(header_scan_rows))

    if mode == "manual" and isinstance(layout.get("columns"), dict) and layout["columns"]:
        cmap = gl_column_map_from_layout_dict(layout)
        if ds_excel_opt is not None:
            cmap = replace(cmap, data_start_row=int(ds_excel_opt))
        elif cmap.data_start_row < h_first_excel + 1:
            cmap = replace(cmap, data_start_row=h_first_excel + 1)
        return cmap

    # --- Auto: derive header_segment from configured band; fall back to sheet top -------------
    header_part_rows: list[tuple[Any, ...]]
    row_count = len(matrix)
    if ds_excel_opt is not None:
        ds_excel = int(ds_excel_opt)
        h0 = h_first_excel - 1
        end_exclusive = min(ds_excel - 1, row_count)
        if 0 <= h0 < row_count and h0 < end_exclusive:
            header_part_rows = list(matrix[h0:end_exclusive])
        else:
            header_part_rows = []
        if not header_part_rows:
            span = max(1, min(hdr_fallback, max(0, row_count - h0)))
            header_part_rows = list(matrix[h0 : min(row_count, h0 + span)])
            if not header_part_rows:
                header_part_rows = list(matrix[: min(hdr_fallback, row_count)])
        auto = detect_gl_column_map(header_part_rows)
        auto = replace(auto, data_start_row=max(h_first_excel + 1, ds_excel))
    else:
        h0 = max(0, h_first_excel - 1)
        header_part_rows = list(matrix[h0 : min(row_count, h0 + hdr_fallback)])
        auto = detect_gl_column_map(header_part_rows)
        if auto.data_start_row < h_first_excel + 1:
            auto = replace(auto, data_start_row=h_first_excel + 1)

    ncols = _matrix_width(matrix)
    ds_skip = max(0, auto.data_start_row - 1)
    sample = matrix[ds_skip : min(len(matrix), ds_skip + 150)]
    refined = refine_debit_credit_columns_from_data(sample, auto, ncols)

    cols_side = layout.get("columns") if isinstance(layout.get("columns"), dict) else {}
    tr_o = cols_side.get("tr_number")
    if tr_o is not None and str(tr_o).strip() != "":
        try:
            triv = int(tr_o)
            if triv >= 0:
                refined = replace(refined, tr_number=triv)
        except (TypeError, ValueError):
            pass
    return refined


def excel_column_letter(col_index_0based: int) -> str:
    """0-based column index → Excel column label (A, B, …, AA)."""
    return get_column_letter(col_index_0based + 1)


def _cell_value_resolving_merge(ws, row: int, col: int) -> Any:
    """Return display value for a coordinate, using merged range top-left when applicable."""
    for mrange in ws.merged_cells.ranges:
        if mrange.min_row <= row <= mrange.max_row and mrange.min_col <= col <= mrange.max_col:
            return ws.cell(mrange.min_row, mrange.min_col).value
    return ws.cell(row, col).value


def gl_header_column_options_from_worksheet(
    ws,
    *,
    header_first_row: int = 1,
    data_start_row: int = 2,
    max_columns: int = 45,
) -> list[tuple[int, str]]:
    """
    Build (0-based column index, display label) options for mapping the GL / journal.

    Heading cells are scanned on Excel rows ``header_first_row .. data_start_row - 1`` (inclusive).
    Horizontal merges that touch that band are surfaced as one label (left column index).
    """
    hf_excel = max(1, int(header_first_row))
    ds_excel = max(hf_excel + 1, int(data_start_row))
    header_last_excel = ds_excel - 1

    max_c = min(int(ws.max_column or 1), int(max_columns))
    if max_c < 1:
        return []

    assigned = [False] * (max_c + 2)  # 1-based column indexing into this list
    options: list[tuple[int, str]] = []

    ranges_sorted = sorted(ws.merged_cells.ranges, key=lambda r: (r.min_row, r.min_col, r.max_row, r.max_col))

    for m in ranges_sorted:
        if m.max_row < hf_excel or m.min_row > header_last_excel or m.max_col < 1:
            continue
        r_lo = max(hf_excel, m.min_row)
        r_hi = min(header_last_excel, m.max_row)
        if r_lo > r_hi:
            continue
        lo, hi = m.min_col, min(m.max_col, max_c)
        if lo > max_c or lo < 1:
            continue
        label_raw = ws.cell(m.min_row, m.min_col).value
        label_s = str(label_raw).strip() if label_raw is not None else ""
        if not label_s:
            label_s = f"(empty)"
        letter_lo = get_column_letter(lo)
        letter_hi = get_column_letter(hi)
        letters = letter_lo if hi == lo else f"{letter_lo}–{letter_hi}"
        display = f"{label_s} [{letters}]"
        options.append((lo - 1, display))
        for cc in range(lo, hi + 1):
            if 1 <= cc <= max_c:
                assigned[cc] = True

    for col in range(1, max_c + 1):
        if assigned[col]:
            continue
        label_s = ""
        for row in range(header_last_excel, hf_excel - 1, -1):
            v = _cell_value_resolving_merge(ws, row, col)
            if v is not None and str(v).strip():
                label_s = str(v).strip()
                break
        if not label_s:
            label_s = "(empty)"
        letter = get_column_letter(col)
        display = f"{label_s} [{letter}]"
        options.append((col - 1, display))

    options.sort(key=lambda x: x[0])
    return options


def gl_header_column_options_from_bytes(
    data: bytes,
    original_filename: str,
    sheet_name: str,
    *,
    header_first_row: int = 1,
    data_start_row: int = 2,
    max_columns: int = 45,
) -> list[tuple[int, str]]:
    """
    Load column heading choices from workbook bytes (.xlsx / .xlsm with merges), .xls,
    or .csv (bottom row inside the configured header span).
    """
    ext = Path(original_filename.strip()).suffix.lower()
    hf = max(1, int(header_first_row))
    ds_excel = max(hf + 1, int(data_start_row))
    buf = io.BytesIO(data)

    if ext in (".xlsx", ".xlsm"):
        wb = load_workbook(buf, read_only=False, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return []
            ws = wb[sheet_name]
            return gl_header_column_options_from_worksheet(
                ws,
                header_first_row=hf,
                data_start_row=ds_excel,
                max_columns=max_columns,
            )
        finally:
            wb.close()

    if ext == ".csv":
        import csv

        text = data.decode("utf-8-sig", errors="replace")
        try:
            dialect = csv.Sniffer().sniff(text[:4096])
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(io.StringIO(text), dialect)
        csv_lines_take = ds_excel - 1  # CSV lines 1.. = Excel-style rows above first data
        rows_accum: list[list[str]] = []
        for i, row in enumerate(reader):
            if i >= csv_lines_take:
                break
            rows_accum.append(list(row))
        if not rows_accum:
            return []
        h0 = hf - 1
        slice_end_exclusive = ds_excel - 1
        segment = rows_accum[h0:slice_end_exclusive] if slice_end_exclusive > h0 else []
        hdr_row_vals = segment[-1] if segment else rows_accum[-1]
        width = min(int(max_columns), max(len(hdr_row_vals), 1))
        out: list[tuple[int, str]] = []
        for i in range(width):
            cell = hdr_row_vals[i].strip() if i < len(hdr_row_vals) else ""
            lab = cell or "(empty)"
            letter = get_column_letter(i + 1)
            out.append((i, f"{lab} [{letter}]"))
        return out

    if ext == ".xls":
        import pandas as pd

        hdr_lines_excel_above_data = ds_excel - 1
        hdr_lines_excel_above_data = max(1, hdr_lines_excel_above_data)
        try:
            df = pd.read_excel(
                buf,
                sheet_name=sheet_name,
                header=None,
                nrows=min(500, hdr_lines_excel_above_data),
                engine="xlrd",
            )
        except Exception:
            return []
        if df.empty:
            return []
        h_idx = hf - 1
        end_idx_exclusive = ds_excel - 1
        slice_df = df.iloc[h_idx:end_idx_exclusive] if end_idx_exclusive > h_idx else df.iloc[h_idx:h_idx]
        last = slice_df.iloc[-1] if not slice_df.empty else df.iloc[min(len(df) - 1, end_idx_exclusive - 1)]
        n = min(int(max_columns), max(int(df.shape[1]), 1))
        out = []
        for i in range(n):
            v = last.iloc[i] if i < len(last) else None
            if v is None or (isinstance(v, float) and pd.isna(v)):
                lab = "(empty)"
            else:
                lab = str(v).strip() or "(empty)"
            letter = get_column_letter(i + 1)
            out.append((i, f"{lab} [{letter}]"))
        return out

    return []


def gl_header_column_options_from_path(
    path: str,
    sheet_name: str,
    *,
    header_first_row: int = 1,
    data_start_row: int = 2,
    max_columns: int = 45,
) -> list[tuple[int, str]]:
    """Read heading options from a local ``.xlsx`` / ``.xlsm`` path (uses openpyxl merge metadata)."""
    ext = Path(path).suffix.lower()
    if ext not in (".xlsx", ".xlsm"):
        return []
    hf = max(1, int(header_first_row))
    ds_ex = max(hf + 1, int(data_start_row))
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        return gl_header_column_options_from_worksheet(
            wb[sheet_name],
            header_first_row=hf,
            data_start_row=ds_ex,
            max_columns=max_columns,
        )
    finally:
        wb.close()


def fallback_gl_header_options(max_cols: int = 26) -> list[tuple[int, str]]:
    """Placeholder headings when the workbook is unavailable."""
    return [(i, f"Column {excel_column_letter(i)} [{excel_column_letter(i)}]") for i in range(max_cols)]


def read_numeric_cell_openpyxl(path: str | Path, *, sheet_name: str, cell_a1: str) -> tuple[Optional[float], Optional[str]]:
    """
    Load a workbook cell numeric value using cached formula results (``data_only=True``).

    Returns ``(amount, None)`` or ``(None, short_error_message)``. Only OpenXML formats
    (``.xlsx`` / ``.xlsm``); legacy ``.xls`` is not supported here.
    """
    fp = Path(path)
    suf = fp.suffix.lower()
    if suf not in APPENDABLE_EXTENSIONS:
        return None, f"Anchors require .xlsx or .xlsm (got {suf or 'unknown'})."
    sh = (sheet_name or "").strip()
    coord = str(cell_a1 or "").strip().upper().replace("$", "")
    if not sh:
        return None, "Sheet name is empty."
    if not coord:
        return None, "Cell reference is empty."

    wb = None
    try:
        wb = load_workbook(filename=str(fp), read_only=True, data_only=True, keep_links=False)
        if sh not in wb.sheetnames:
            avail = ", ".join(wb.sheetnames[:12])
            more = "" if len(wb.sheetnames) <= 12 else ", …"
            return None, f"Sheet {sh!r} not found (available: {avail}{more})."
        ws = wb[sh]
        val = ws[coord].value
        wb.close()
        wb = None
        if val is None:
            return None, f"Cell {coord} on {sh!r} is empty."
        if isinstance(val, bool):
            return None, f"Cell {coord} is boolean, not an amount."
        if isinstance(val, Decimal):
            return float(val), None
        fv = _try_float(val)
        if fv is None:
            return None, f"Cell {coord} is not numeric ({type(val).__name__})."
        return fv, None
    except Exception as e:
        return None, str(e).strip() or repr(e)
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _is_excel_formula_value(val: Any) -> bool:
    """True when a cell stores an Excel formula (``data_only=False`` load)."""
    return isinstance(val, str) and str(val).strip().startswith("=")


def _try_float(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    if _is_excel_formula_value(val):
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    s = str(val).strip().replace(",", "")
    for sym in ("$", "€", "£", "₹", "\u20b9"):
        s = s.replace(sym, "")
    if not s or s in ("-", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _gl_strings_form_month_banner(parts: list[str]) -> bool:
    """True when cells contain only month tokens / optional years (merged month heading row)."""
    cells: list[str] = []
    for raw in parts:
        s = str(raw).strip()
        if not s:
            continue
        tf = _try_float(s.replace(",", ""))
        if tf is not None and abs(tf) > 1e-9:
            return False
        cells.append(s)
    if not cells:
        return False
    blob = re.sub(r"\s+", " ", " ".join(cells).lower()).strip()
    return bool(_GL_MONTH_BANNER_ROW_FULL.match(blob))


def _tuple_row_is_month_section_banner(row: tuple[Any, ...], *, max_scan_cols: int = 40) -> bool:
    """Whole-row heuristic before GlColumnMap exists (header auto-detect path)."""
    parts: list[str] = []
    for c in row[:max_scan_cols]:
        if c is None:
            continue
        if isinstance(c, float) and math.isnan(c):
            continue
        s = str(c).strip()
        if not s:
            continue
        tf = _try_float(s.replace(",", ""))
        if tf is not None and abs(tf) > 1e-9:
            return False
        parts.append(s)
    return _gl_strings_form_month_banner(parts)


BEGINNING_BALANCE_LABEL = gla.BEGINNING_BALANCE_LABEL
_BROUGHT_FORWARD_ACCOUNT_RE = re.compile(r"^(.*?)\s+brought\s+(?:forward|fwd)\b", re.I)

_JOURNAL_HEADER_ACCOUNT_LABELS = frozenset(
    {
        "particulars",
        "account",
        "accounts",
        "debit",
        "credit",
        "date",
        "balance",
        "description",
        "journal entries template",
        "journal entries",
    }
)


def _normalize_headerish_label(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip().lower())


def is_non_posting_gl_row(rec: dict[str, Any], *, amt_tol: float = 0.015) -> bool:
    """
    Drop repeated journal template / column header rows (e.g. ``Date,#,Accounts,,Debit,Credit``).

    These appear between month sections in CSV/Excel journals and are not real postings.
    """
    deb = row_amount(rec.get("debit"))
    cre = row_amount(rec.get("credit"))
    if deb > amt_tol or cre > amt_tol:
        return False

    acct = _normalize_headerish_label(str(rec.get("account") or ""))
    desc = _normalize_headerish_label(str(rec.get("description") or ""))
    if acct in _JOURNAL_HEADER_ACCOUNT_LABELS or desc in _JOURNAL_HEADER_ACCOUNT_LABELS:
        return True

    tr_raw = str(rec.get("transaction_number") or "").strip()
    tr = tr_raw.lower()
    gd = _normalize_headerish_label(str(rec.get("gl_date") or ""))

    if acct in ("account", "accounts") and tr in ("#", ""):
        if gd in ("", "date", ":") or not any(ch.isdigit() for c in gd):
            return True

    if gd == "date" and acct in ("account", "accounts"):
        return True

    if tr == "#" and acct in ("account", "accounts") and desc in ("", "account", "accounts"):
        return True

    # Orphan merge stripe: date/description propagated but no account and no amounts
    # (can appear after a leg row is removed from a merged journal band).
    acct_raw = str(rec.get("account") or "").strip()
    if deb <= amt_tol and cre <= amt_tol and not acct_raw:
        if parse_gl_cell_to_date(rec.get("gl_date")) is not None:
            return True
        desc_raw = str(rec.get("description") or "").strip()
        if desc_raw and not acct_raw and parse_gl_cell_to_date(desc_raw) is not None:
            return True

    return False


def filter_non_posting_gl_rows(
    records: list[dict[str, Any]], *, amt_tol: float = 0.015
) -> list[dict[str, Any]]:
    return [r for r in records if not is_non_posting_gl_row(r, amt_tol=amt_tol)]


def _gl_row_is_month_banner_with_colmap(
    row: tuple[Any, ...],
    colmap: GlColumnMap,
    deb_amt: Optional[float],
    cre_amt: Optional[float],
) -> bool:
    """Skip merged month headings between header row(s) and journal lines."""
    if deb_amt is not None and abs(deb_amt) > 1e-12:
        return False
    if cre_amt is not None and abs(cre_amt) > 1e-12:
        return False
    idxs = sorted(
        {i for i in (colmap.date, colmap.details, colmap.particulars, colmap.tr_number) if i is not None}
    )
    structural_vals: list[str] = []
    for i in idxs:
        v = _cell(row, i)
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        tf = _try_float(s.replace(",", ""))
        if tf is not None and abs(tf) > 1e-9:
            return False
        structural_vals.append(s)
    if structural_vals and _gl_strings_form_month_banner(structural_vals):
        return True
    return _tuple_row_is_month_section_banner(row)


def sniff_currency_iso_from_primitive(val: Any) -> Optional[str]:
    """Infer ISO 4217 from symbols or (CODE) in formatted amount cells."""
    if val is None:
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return None
    raw = str(val)
    mpar = re.search(r"\(([A-Za-z]{3})\)", raw)
    if mpar:
        return mpar.group(1).upper()
    for ch, iso in (
        ("€", "EUR"),
        ("£", "GBP"),
        ("₹", "INR"),
        ("\u20b9", "INR"),
    ):
        if ch in raw:
            return iso
    if "¥" in raw:
        return "JPY"
    if "$" in raw:
        return "USD"
    mx = _ISO_TOKEN_RE.search(raw.strip().upper())
    if mx:
        return mx.group("iso")[:3]
    return None


def normalize_currency_hint_cell(val: Any) -> Optional[str]:
    """Standalone column cells with currency labels (USD, CAD, …)."""
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return None
    s_up = str(val).strip().upper()
    mx = _ISO_TOKEN_RE.search(s_up.replace("'", ""))
    if mx:
        return mx.group("iso")[:3]
    if len(s_up) == 3 and s_up.isalpha():
        return s_up
    lo = str(val).strip().lower().split()
    for needle, iso in (
        ("usd", "USD"),
        ("eur", "EUR"),
        ("gbp", "GBP"),
        ("cad", "CAD"),
        ("aud", "AUD"),
        ("inr", "INR"),
        ("jpy", "JPY"),
        ("chf", "CHF"),
        ("qar", "QAR"),
    ):
        if needle in lo:
            return iso
    return None


def parse_money_cell(val: Any) -> tuple[Optional[float], Optional[str]]:
    """Parse amount and currency hint from a debit/credit cell (e.g. ``QAR 10,119``)."""
    if _is_excel_formula_value(val):
        return None, sniff_currency_iso_from_primitive(val)
    iso = sniff_currency_iso_from_primitive(val)
    amt = _try_float(val)
    if amt is not None:
        return amt, iso
    if iso is None or val is None or str(val).strip() == "":
        return None, iso
    s = str(val).strip().replace(",", "")
    for sym in ("$", "€", "£", "₹", "\u20b9"):
        s = s.replace(sym, "")
    s = re.sub(rf"^{re.escape(iso)}\s*", "", s, flags=re.IGNORECASE).strip()
    if not s:
        return None, iso
    try:
        return float(s), iso
    except ValueError:
        return None, iso


def _sheet_raw_cell_matrix(ws, max_r: int, max_c: int) -> list[list[Any]]:
    """Read cell values into a grid (0-based row/col); merged slaves may be None until fill pass."""
    grid: list[list[Any]] = []
    for r in range(max_r):
        row_vals: list[Any] = []
        for c in range(max_c):
            row_vals.append(ws.cell(row=r + 1, column=c + 1).value)
        grid.append(row_vals)
    return grid


def _apply_merged_cell_fill_to_grid(
    grid: list[list[Any]],
    ws,
    *,
    debit_col_0based: Optional[int] = None,
    credit_col_0based: Optional[int] = None,
) -> None:
    """
    Propagate merged-range anchor values into covered cells (date/memo stripes, etc.).

    Skips **multi-row** rectangles that span **both** debit and credit columns: filling there would
    duplicate one amount into the other column and break GL reads (empty or wrong credits).
    """
    nrow = len(grid)
    ncol = len(grid[0]) if nrow else 0
    for mrange in ws.merged_cells.ranges:
        r1, r2 = mrange.min_row - 1, mrange.max_row - 1
        c1, c2 = mrange.min_col - 1, mrange.max_col - 1
        if nrow == 0 or ncol == 0 or r1 >= nrow or c1 >= ncol or r1 < 0 or c1 < 0:
            continue
        r2 = min(r2, nrow - 1)
        c2 = min(c2, ncol - 1)
        if r2 < r1 or c2 < c1:
            continue
        if (
            debit_col_0based is not None
            and credit_col_0based is not None
            and debit_col_0based != credit_col_0based
            and r2 > r1
            and c1 <= debit_col_0based <= c2
            and c1 <= credit_col_0based <= c2
        ):
            continue
        anchor = grid[r1][c1]
        if anchor is None:
            continue
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                if cc < len(grid[rr]):
                    grid[rr][cc] = anchor


def _sheet_to_matrix_expanded_merges(
    ws,
    max_r: int,
    max_c: int,
    *,
    debit_col_0based: Optional[int] = None,
    credit_col_0based: Optional[int] = None,
) -> list[list[Any]]:
    """Cell values with merged fills propagated (e.g. date spanning debit/credit rows)."""
    grid = _sheet_raw_cell_matrix(ws, max_r, max_c)
    _apply_merged_cell_fill_to_grid(
        grid,
        ws,
        debit_col_0based=debit_col_0based,
        credit_col_0based=credit_col_0based,
    )
    return grid


def _cell(row: tuple[Any, ...], idx: int) -> Any:
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _account_column_scan_upper_bound(colmap: GlColumnMap, particulars_0based: int, row_len: int) -> int:
    """Scan for account spillover stops before debit/credit/Tr/currency columns to the right of particulars."""
    p = int(particulars_0based)
    cand: list[int] = []
    for c in (colmap.debit, colmap.credit, colmap.tr_number, colmap.currency):
        if c is not None and int(c) > p:
            cand.append(int(c))
    return min(cand) if cand else row_len


def _worksheet_row_tuple_for_colmap(ws, row_1b: int, colmap: GlColumnMap) -> tuple[Any, ...]:
    """Build a 0-based column tuple for one sheet row (same shape GL read uses)."""
    p = int(colmap.particulars)
    scan_hi = _account_column_scan_upper_bound(colmap, p, int(ws.max_column or 1))
    row_len = max(
        scan_hi,
        int(colmap.debit or 0) + 1,
        int(colmap.credit or 0) + 1,
        int(colmap.date or 0) + 1,
        int(colmap.details or 0) + 1,
    )
    return tuple(_cell_value_resolving_merge(ws, row_1b, c + 1) for c in range(row_len))


def _plausible_account_label_from_cell(v: Any) -> str | None:
    """Return stripped account text, or None when the cell is empty or a money amount."""
    if v is None or not str(v).strip():
        return None
    s = str(v).strip()
    money, _iso = parse_money_cell(s)
    if money is not None:
        if abs(float(money)) > 1e-9:
            return None
        if re.fullmatch(r"[\d,.\s\-+]+", s):
            return None
    return s


def _effective_particulars_account_label(row: tuple[Any, ...], colmap: GlColumnMap) -> str:
    """
    Ledger account label: use the mapped **particulars** cell, or—when it is blank—the first
    plausible text cell nearby (indented credit lines, extra account columns, unmerged cells).
    Legacy sheets with every account in the primary column behave unchanged.
    """
    p = int(colmap.particulars)
    primary = _plausible_account_label_from_cell(_cell(row, p))
    if primary:
        return primary

    scan_hi = min(_account_column_scan_upper_bound(colmap, p, len(row)), len(row))
    if colmap.debit is not None:
        scan_hi = min(scan_hi, int(colmap.debit))
    scan_hi = min(scan_hi, p + 3)
    scan_lo = max(0, p - 1)

    def _scan_account_candidates(col_range: range) -> str | None:
        for c in col_range:
            if c == p or c < 0 or c >= len(row):
                continue
            if c == colmap.date:
                continue
            if c == colmap.details:
                if c != p - 1:
                    continue
                deb_amt = cre_amt = None
                if colmap.debit is not None:
                    deb_amt, _ = parse_money_cell(_cell(row, colmap.debit))
                if colmap.credit is not None:
                    cre_amt, _ = parse_money_cell(_cell(row, colmap.credit))
                deb_z = deb_amt is None or float(deb_amt) <= 1e-9
                cre_z = cre_amt is None or float(cre_amt) <= 1e-9
                if deb_z and cre_z:
                    continue
            if colmap.tr_number is not None and c == colmap.tr_number:
                continue
            if colmap.currency is not None and c == colmap.currency:
                continue
            hit = _plausible_account_label_from_cell(_cell(row, c))
            if hit:
                return hit
        return None

    right = _scan_account_candidates(range(p + 1, scan_hi))
    if right:
        return right
    left = _scan_account_candidates(range(scan_lo, p))
    return left or ""


def _row_tuple_to_record_dynamic(
    row: tuple[Any, ...],
    colmap: GlColumnMap,
    *,
    header_skipped_ref: list[bool],
    excel_row_1based: Optional[int] = None,
) -> dict[str, Any] | None:
    date_v = _cell(row, colmap.date)
    details_v = _cell(row, colmap.details)
    deb_v = _cell(row, colmap.debit)
    cre_v = _cell(row, colmap.credit)

    deb_amt, deb_iso_from_cell = parse_money_cell(deb_v)
    cre_amt, cre_iso_from_cell = parse_money_cell(cre_v)
    row_iso = None
    if colmap.currency is not None:
        row_iso = normalize_currency_hint_cell(_cell(row, colmap.currency))
    row_iso = row_iso or deb_iso_from_cell or cre_iso_from_cell or DEFAULT_CURRENCY_ISO

    deb_f = deb_amt
    cre_f = cre_amt

    if _gl_row_is_month_banner_with_colmap(row, colmap, deb_f, cre_f):
        return None

    particulars_eff = _effective_particulars_account_label(row, colmap)

    # Faux-deleted rows are cleared on the sheet but may still inherit date/description from
    # merges — treat as empty when there is no account and no amounts.
    row_empty = (
        not particulars_eff
        and (deb_f is None or deb_f == 0)
        and (cre_f is None or cre_f == 0)
    )
    if row_empty:
        return None

    # Skip repeated journal column-header rows (e.g. Date / # / Accounts with zero amounts).
    probe = {
        "account": particulars_eff,
        "description": "" if details_v is None else str(details_v),
        "debit": float(deb_f or 0),
        "credit": float(cre_f or 0),
        "gl_date": "",
        "transaction_number": _cell(row, colmap.tr_number) if colmap.tr_number is not None else None,
    }
    if is_non_posting_gl_row(probe):
        return None

    dv = ""
    if date_v is not None:
        parsed_dt = _coerce_journal_sheet_date(date_v)
        if parsed_dt is not None:
            dv = parsed_dt.isoformat()
        elif hasattr(date_v, "isoformat"):
            dv = str(date_v.isoformat())[:10]
        else:
            dv_str = str(date_v).strip()
            if dv_str and any(ch.isdigit() for ch in dv_str[:16]):
                dv = dv_str
            elif not header_skipped_ref[0]:
                combined = f"{particulars_eff} {details_v or ''}".lower()
                if ("description" in combined or "particular" in combined) and deb_f is None and cre_f is None:
                    header_skipped_ref[0] = True
                    return None

    particulars_s = particulars_eff
    details_s = "" if details_v is None else str(details_v)
    if colmap.details == colmap.particulars and particulars_s == details_s:
        details_s = particulars_s

    deb_out = float(deb_f) if deb_f is not None else 0.0
    cre_out = float(cre_f) if cre_f is not None else 0.0

    out: dict[str, Any] = {
        "gl_date": dv,
        "description": details_s,
        "account": particulars_s,
        "debit": deb_out,
        "credit": cre_out,
        "currency_iso": row_iso,
    }
    if colmap.tr_number is not None:
        tr_v = _cell(row, colmap.tr_number)
        if tr_v is not None and str(tr_v).strip() != "":
            out["transaction_number"] = tr_v
    if excel_row_1based is not None:
        out["_excel_row"] = int(excel_row_1based)
    return out


def path_supports_gl_append(path: str) -> bool:
    return Path(path).suffix.lower() in APPENDABLE_EXTENSIONS


def list_sheet_names_from_bytes(original_filename: str, data: bytes) -> list[str]:
    """
    Inspect workbook bytes and return worksheet names (.xlsx / .xlsm / .xls).
    Returns an empty list for formats without sheets (e.g. .csv) or if the workbook cannot be read.
    """
    ext = Path(original_filename.strip()).suffix.lower()
    try:
        if ext in (".xlsx", ".xlsm"):
            wb = load_workbook(io.BytesIO(data), read_only=True)
            try:
                return list(wb.sheetnames)
            finally:
                wb.close()
        if ext == ".xls":
            import xlrd

            book = xlrd.open_workbook(file_contents=data)
            return list(book.sheet_names())
    except Exception:
        return []
    return []


def _openpyxl_col(i0: int) -> int:
    """0-based column index → openpyxl 1-based column number."""
    return i0 + 1


def _copy_row_style(ws, src_row: int, dest_row: int, *, max_col: int) -> None:
    for c in range(1, max_col + 1):
        if _is_merged_slave_cell(ws, int(dest_row), c):
            continue
        src_ar, src_ac = _merged_top_left(ws, int(src_row), c)
        src = ws.cell(row=src_ar, column=src_ac)
        dest_ar, dest_ac = _merged_top_left(ws, int(dest_row), c)
        dest = ws.cell(row=dest_ar, column=dest_ac)
        if src.font:
            dest.font = _copy_style(src.font)
        if src.border:
            dest.border = _copy_style(src.border)
        if src.fill:
            dest.fill = _copy_style(src.fill)
        dest.number_format = src.number_format
        if src.protection:
            dest.protection = _copy_style(src.protection)
        if src.alignment:
            dest.alignment = _copy_style(src.alignment)


def _merge_range_covering_cell(ws, row_1b: int, col_1b: int):
    """Return openpyxl merge range if (row, col) sits inside a merged rectangle, else None."""
    for m in ws.merged_cells.ranges:
        if m.min_row <= row_1b <= m.max_row and m.min_col <= col_1b <= m.max_col:
            return m
    return None


def infer_last_entry_template_rows(ws, last_data_row_1b: int, date_col_1b: int, desc_col_1b: int) -> tuple[int, int]:
    """
    Contiguous rows for the workbook's **last journal** footprint.

    Unions merged **date** and **description** column spans touching the last posted row — they often
    differ by one row; using only the tallest would miss merges that anchor the memo block.

    Fallback: previous two sheet rows when no sane vertical merges are found (2–40 rows).
    """
    if last_data_row_1b < 1:
        return 1, 1
    lo = last_data_row_1b
    hi = last_data_row_1b
    found = False
    for col in (date_col_1b, desc_col_1b):
        m = _merge_range_covering_cell(ws, last_data_row_1b, col)
        if m is None or m.max_row <= m.min_row:
            continue
        h = m.max_row - m.min_row + 1
        if not (2 <= h <= 40):
            continue
        lo = min(lo, int(m.min_row))
        hi = max(hi, int(m.max_row))
        found = True
    if found and hi > lo:
        return lo, hi
    if last_data_row_1b >= 2:
        return last_data_row_1b - 1, last_data_row_1b
    return last_data_row_1b, last_data_row_1b


def _merged_top_left(ws, row_1b: int, col_1b: int) -> tuple[int, int]:
    m = _merge_range_covering_cell(ws, row_1b, col_1b)
    if m:
        return int(m.min_row), int(m.min_col)
    return row_1b, col_1b


def particulars_columns_1b(ws, row_1b: int, particulars_col_1b: int) -> list[int]:
    """
    1-based columns that form the **Particulars** band on ``row_1b`` (e.g. merged C–D).

    When the mapped particulars cell sits in a single-row horizontal merge, every column in
    that merge is included so reads/writes and blank checks cover the full label stripe.
    """
    row_1b = int(row_1b)
    p = int(particulars_col_1b)
    m = _merge_range_covering_cell(ws, row_1b, p)
    if (
        m is not None
        and int(m.min_row) == int(m.max_row) == row_1b
        and int(m.max_col) > int(m.min_col)
        and int(m.min_col) <= p <= int(m.max_col)
    ):
        return list(range(int(m.min_col), int(m.max_col) + 1))
    return [p]


def particulars_label_on_worksheet_row(
    ws,
    colmap: GlColumnMap,
    row_1b: int,
    *,
    credit_style: str = "standard",
) -> str:
    """Account label for one sheet row, scanning the full merged particulars band when present."""
    row_1b = int(row_1b)
    acct_c = _openpyxl_col(colmap.particulars)
    deb_f, cre_f, _, _ = _logical_debit_credit_from_row(
        ws,
        None,
        colmap,
        row_1b,
        credit_style=credit_style,
        physical=True,
    )
    if cre_f > 1e-9 and deb_f <= 1e-9:
        date_c = _openpyxl_col(colmap.date)
        desc_c = _openpyxl_col(colmap.details)
        block_lo, _block_hi = _journal_block_bounds_workbook(ws, colmap, row_1b, date_c, desc_c)
        tpl_top, tpl_bot = infer_last_entry_template_rows(ws, block_lo, date_c, desc_c)
        read_col = _gl_account_write_col_1b(
            ws,
            row_1b,
            colmap,
            tpl_top=tpl_top,
            tpl_bot=tpl_bot,
            credit_style=credit_style,
        )
        scan_cols = [read_col] + [
            c for c in particulars_columns_1b(ws, row_1b, acct_c) if c != read_col
        ]
        for col in scan_cols:
            anchor_r, _anchor_c = _merged_top_left(ws, row_1b, col)
            if anchor_r < row_1b:
                continue
            hit = _plausible_account_label_from_cell(_cell_value_resolving_merge(ws, row_1b, col))
            if hit:
                return hit
    else:
        for col in particulars_columns_1b(ws, row_1b, acct_c):
            hit = _plausible_account_label_from_cell(_cell_value_resolving_merge(ws, row_1b, col))
            if hit:
                return hit
    row_t = _worksheet_row_tuple_for_colmap(ws, row_1b, colmap)
    return _effective_particulars_account_label(row_t, colmap)


def validate_gl_insert_specs_have_accounts(rows: list[dict[str, Any]]) -> str | None:
    """Return an error when an insert line has amounts but no account name in the plan."""
    bad: list[int] = []
    for i, spec in enumerate(rows):
        deb = float(spec.get("debit") or 0)
        cred = float(spec.get("credit") or 0)
        if deb <= 1e-9 and cred <= 1e-9:
            continue
        if not _norm_account_label(spec.get("account")):
            bad.append(i + 1)
    if not bad:
        return None
    legs = ", ".join(str(x) for x in bad[:6])
    if len(bad) > 6:
        legs += f", … (+{len(bad) - 6} more)"
    return (
        f"Journal line(s) {legs} have debit/credit but no account (particulars). "
        "Enter an account on each leg before saving."
    )


def verify_gl_posting_rows_have_particulars(
    ws,
    colmap: GlColumnMap,
    rows_1b: list[int],
    *,
    credit_style: str = "standard",
) -> str | None:
    """
    Return an error when a posting row has amounts but no account across its particulars merge.

    Used after GL inserts (e.g. Ali Al Baker sheets with particulars merged across C–D).
    """
    data_floor = int(colmap.data_start_row)
    missing: list[int] = []
    for row_1b in sorted({int(r) for r in rows_1b if int(r) >= data_floor}):
        deb_f, cre_f, _, _ = _logical_debit_credit_from_row(
            ws,
            None,
            colmap,
            row_1b,
            credit_style=credit_style,
            physical=False,
        )
        if deb_f <= 1e-9 and cre_f <= 1e-9:
            continue
        if not _norm_account_label(
            particulars_label_on_worksheet_row(
                ws, colmap, row_1b, credit_style=credit_style
            )
        ):
            missing.append(row_1b)
    if not missing:
        return None
    acct_c = _openpyxl_col(colmap.particulars)
    cols = particulars_columns_1b(ws, missing[0], acct_c)
    from openpyxl.utils import get_column_letter

    letters = (
        get_column_letter(cols[0])
        if len(cols) == 1
        else f"{get_column_letter(cols[0])}–{get_column_letter(cols[-1])}"
    )
    shown = ", ".join(str(r) for r in missing[:8])
    if len(missing) > 8:
        shown += f", … (+{len(missing) - 8} more)"
    return (
        f"Row(s) {shown} have debit/credit but blank particulars ({letters} merged band). "
        "Fill the account on each new line before saving."
    )


def _set_gl_account_on_row(
    ws,
    row_1b: int,
    colmap: GlColumnMap,
    account: Any,
    *,
    tpl_top: int | None = None,
    tpl_bot: int | None = None,
    credit_style: str = "standard",
) -> None:
    """Write an account label onto the particulars merge band (e.g. merged C–D)."""
    acct_s = _norm_account_label(account)
    if not acct_s:
        return
    acct_c = _openpyxl_col(colmap.particulars)
    row_1b = int(row_1b)
    write_col = particulars_columns_1b(ws, row_1b, acct_c)[0]
    if tpl_top is not None and tpl_bot is not None:
        write_col = _gl_account_write_col_1b(
            ws,
            row_1b,
            colmap,
            tpl_top=tpl_top,
            tpl_bot=tpl_bot,
            credit_style=credit_style,
        )
        band = particulars_columns_1b(ws, row_1b, acct_c)
        if write_col not in band:
            write_col = band[0]
    _set_sheet_cell_value_merge_aware(ws, row_1b, write_col, acct_s)


def _gl_account_write_col_1b(
    ws,
    row_1b: int,
    colmap: GlColumnMap,
    *,
    tpl_top: int | None = None,
    tpl_bot: int | None = None,
    credit_style: str = "standard",
) -> int:
    """
    Column for writing an account label on ``row_1b`` (particulars or indented credit column).

    Uses horizontal merges on the journal template leg when available; otherwise the first
    text column right of particulars (same rule as :func:`_effective_particulars_account_label`).
    """
    acct_c = _openpyxl_col(colmap.particulars)
    row_1b = int(row_1b)
    band = particulars_columns_1b(ws, row_1b, acct_c)
    if len(band) > 1:
        return band[0]
    template_r = row_1b
    if tpl_top is not None and tpl_bot is not None:
        block_h = max(1, int(tpl_bot) - int(tpl_top) + 1)
        date_c = _openpyxl_col(colmap.date)
        desc_c = _openpyxl_col(colmap.details)
        block_lo, block_hi = _journal_block_bounds_workbook(ws, colmap, row_1b, date_c, desc_c)
        template_r = int(tpl_top) + ((row_1b - int(block_lo)) % block_h)
        _, horizontal_by_tpl_row = _collect_journal_template_merges(
            ws, int(tpl_top), int(tpl_bot)
        )
        write_col = acct_c
        for _mr1, mc1a, _mr2, mc2a in horizontal_by_tpl_row.get(template_r, ()):
            if mc1a <= acct_c <= mc2a:
                write_col = min(write_col, mc1a)
                return write_col

    for m in ws.merged_cells.ranges:
        if (
            int(m.min_row) <= row_1b <= int(m.max_row)
            and int(m.min_col) <= acct_c <= int(m.max_col)
            and int(m.max_col) > int(m.min_col)
        ):
            return min(acct_c, int(m.min_col))

    deb_f, cre_f, _, _ = _logical_debit_credit_from_row(
        ws,
        None,
        colmap,
        row_1b,
        credit_style=credit_style,
        physical=True,
    )
    if cre_f > 1e-9 and deb_f <= 1e-9:
        ind = int(colmap.particulars) + 1
        bound = _account_column_scan_upper_bound(colmap, int(colmap.particulars), 64)
        if ind < bound and ind != colmap.date and ind != colmap.details:
            if colmap.tr_number is None or ind != colmap.tr_number:
                if colmap.currency is None or ind != colmap.currency:
                    return ind + 1

    row_t = _worksheet_row_tuple_for_colmap(ws, row_1b, colmap)
    if not str(_cell(row_t, colmap.particulars) or "").strip():
        bound = _account_column_scan_upper_bound(colmap, int(colmap.particulars), len(row_t))
        for c in range(int(colmap.particulars) + 1, bound):
            if c == colmap.date or c == colmap.details:
                continue
            if colmap.tr_number is not None and c == colmap.tr_number:
                continue
            if colmap.currency is not None and c == colmap.currency:
                continue
            v = _cell(row_t, c)
            if v is None or not str(v).strip():
                continue
            s = str(v).strip()
            money, _iso = parse_money_cell(s)
            if money is not None and abs(float(money)) > 1e-9:
                continue
            return c + 1
    return acct_c


def _column_in_template_vertical_block(
    block_merges: list[tuple[int, int, int, int]],
    col_1based: int,
    tpl_top: int,
    tpl_bot: int,
) -> tuple[bool, int, int]:
    """
    Whether ``col`` lies in a rectangle that merges **every** sheet row tpl_top–tpl_bot.
    Returns (True, merged_left_col, merged_right_col) or (False, col, col) if discrete.
    """
    for mr1, mc1, mr2, mc2 in block_merges:
        if mr1 == tpl_top and mr2 == tpl_bot and mc1 <= col_1based <= mc2:
            return True, mc1, mc2
    return False, col_1based, col_1based


def _collect_journal_template_merges(
    ws, tpl_top: int, tpl_bot: int
) -> tuple[list[tuple[int, int, int, int]], dict[int, list[tuple[int, int, int, int]]]]:
    """
    * block_vertical*: merges spanning **exactly** ``tpl_top``–``tpl_bot`` (whole journal stripe).
    * horizontal_by_tpl_row*: single-row merges (e.g. account across F–H) keyed by Excel row inside block.
      Excludes rects already covered by a multi-row vertical merge in the journal block so we do not duplicate.
    """
    block_vertical: list[tuple[int, int, int, int]] = []
    cand_h: list[tuple[int, int, int, int]] = []
    for m in ws.merged_cells.ranges:
        mr1, mr2, mc1, mc2 = int(m.min_row), int(m.max_row), int(m.min_col), int(m.max_col)
        if mr1 == tpl_top and mr2 == tpl_bot:
            block_vertical.append((mr1, mc1, mr2, mc2))
        elif mr1 == mr2 and tpl_top <= mr1 <= tpl_bot:
            cand_h.append((mr1, mc1, mr2, mc2))

    def _overlap_single_row_horiz_multiline_vertical(
        r: int, hc1: int, hc2: int, vert: tuple[int, int, int, int]
    ) -> bool:
        vmr1, vmc1, vmr2, vmc2 = vert
        if vmr2 <= vmr1:
            return False
        if vmr1 > r or r > vmr2:
            return False
        return max(hc1, vmc1) <= min(hc2, vmc2)

    horiz_group: dict[int, list[tuple[int, int, int, int]]] = {}
    for mr1, mc1, mr2, mc2 in cand_h:
        r = mr1
        hc1, hc2 = mc1, mc2
        skip = False
        for bv in block_vertical:
            if _overlap_single_row_horiz_multiline_vertical(r, hc1, hc2, bv):
                skip = True
                break
        if not skip:
            horiz_group.setdefault(r, []).append((mr1, mc1, mr2, mc2))
    return block_vertical, horiz_group


def _infer_next_tr_number(previous: Any) -> Any | None:
    """Best-effort next transaction / journal number — follows integer tail when possible."""
    if previous is None:
        return 1
    s = str(previous).strip()
    if not s:
        return 1
    if s.isdigit():
        try:
            return int(s) + 1
        except ValueError:
            return None
    try:
        d = Decimal(s)
        if d == int(d):
            return int(d) + 1
        return float(d + 1)
    except (InvalidOperation, ValueError):
        pass
    m = re.search(r"(\d+)$", s)
    if not m:
        return None
    head, tail = s[: m.start()], m.group(1)
    padded = tail.lstrip("0") or tail
    try:
        nxt = str(int(padded) + 1)
    except ValueError:
        return None
    pad_w = len(tail)
    if pad_w and tail.isdigit() and tail.startswith("0") and pad_w >= len(str(int(padded) + 1)):
        return head + str(int(padded) + 1).zfill(pad_w)
    return head + nxt


def coerce_gl_transaction_number(value: Any) -> Any | None:
    """Normalize a UI/database override for the GL transaction-number cell."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if abs(value - round(value)) < 1e-9:
            return int(round(value))
        return value
    s = str(value).strip()
    if not s:
        return None
    if s.isdigit():
        return int(s)
    try:
        d = Decimal(s.replace(",", ""))
        if d == int(d):
            return int(d)
        return float(d)
    except (InvalidOperation, ValueError):
        return s


def peek_next_transaction_number_from_workbook(
    workbook_path: str,
    *,
    sheet_name: str = GL_SHEET_NAME_DEFAULT,
    layout: Optional[dict[str, Any]] = None,
) -> tuple[Any | None, bool, Optional[str]]:
    """
    Infer the transaction number that would follow the last posted row (same rules as posting).

    Returns ``(next_number, has_tr_column, error_message)``.
    When ``has_tr_column`` is False, ``next_number`` is ``None`` and this is not an error.
    """
    ext = Path(workbook_path).suffix.lower()
    if ext not in APPENDABLE_EXTENSIONS:
        return None, False, f"Requires .xlsx or .xlsm workbook; got {ext!r}"
    wb = load_workbook(workbook_path, keep_vba=(ext == ".xlsm"))
    try:
        if sheet_name not in wb.sheetnames:
            return None, False, f"Sheet {sheet_name!r} not found"
        ws = wb[sheet_name]
        mr = ws.max_row or 1
        max_c = min(ws.max_column or 1, 45)
        scan_r = min(ws.max_row or 1, max(60, mr + 2))
        raw_scan = _sheet_raw_cell_matrix(ws, scan_r, max_c)
        colmap = resolve_gl_column_map([tuple(r) for r in raw_scan], layout)
        grid = [list(r) for r in raw_scan]
        _apply_merged_cell_fill_to_grid(
            grid, ws, debit_col_0based=colmap.debit, credit_col_0based=colmap.credit
        )
        matrix: list[tuple[Any, ...]] = [tuple(row) for row in grid]
        colmap = resolve_gl_column_map(matrix, layout)
        tr_c = _openpyxl_col(colmap.tr_number) if colmap.tr_number is not None else None
        if tr_c is None:
            return None, False, None
        prev_tr_raw = _cell_value_resolving_merge(ws, mr, tr_c)
        nxt = _infer_next_tr_number(prev_tr_raw)
        if nxt is None:
            nxt = 1
        return nxt, True, None
    finally:
        wb.close()


def _value_for_new_gl_date(ws, tpl_top: int, date_col_1b: int, gl_dt: date) -> Any:
    """Match stored type of existing GL date cells where possible."""
    tpl_r, tpl_c = _merged_top_left(ws, tpl_top, date_col_1b)
    sample = _cell_value_resolving_merge(ws, tpl_top, date_col_1b)
    if isinstance(sample, datetime):
        return datetime(gl_dt.year, gl_dt.month, gl_dt.day, sample.hour or 0, sample.minute or 0, sample.second or 0)
    return gl_dt


def _mirror_anchor_number_format(ws, tpl_row: int, tpl_col: int, dst_row: int, dst_col: int) -> None:
    src_ar, src_ac = _merged_top_left(ws, int(tpl_row), int(tpl_col))
    src = ws.cell(row=src_ar, column=src_ac)
    if src.number_format:
        dst_ar, dst_ac = _merged_top_left(ws, int(dst_row), int(dst_col))
        ws.cell(row=dst_ar, column=dst_ac).number_format = src.number_format


def _set_cell_number_format_merge_aware(ws, row_1b: int, col_1b: int, number_format: str) -> None:
    wr, wc = _merged_top_left(ws, int(row_1b), int(col_1b))
    if _is_merged_slave_cell(ws, wr, wc):
        wr, wc = _resolve_writable_coordinate(ws, int(row_1b), int(col_1b))
    ws.cell(row=wr, column=wc).number_format = number_format


def _resolve_gl_date_number_format(
    ws,
    tpl_top: int,
    tpl_bot: int,
    date_col_1b: int,
    last_data_row_1b: int,
) -> str:
    """
    Excel date *display* follows ``number_format``. Prefer the format on the **last posted GL row**
    (what users already see), then any non-``General`` format in the inferred template rows.
    """

    def _nf_explicit(row_1b: int) -> str:
        ar, ac = _merged_top_left(ws, row_1b, date_col_1b)
        nf = ws.cell(row=ar, column=ac).number_format
        s = (nf or "").strip()
        if not s or s.lower() == "general":
            return ""
        return nf

    if last_data_row_1b >= 1:
        got = _nf_explicit(last_data_row_1b)
        if got:
            return got
    for r in range(tpl_top, tpl_bot + 1):
        got = _nf_explicit(r)
        if got:
            return got
    if last_data_row_1b >= 1:
        ar, ac = _merged_top_left(ws, last_data_row_1b, date_col_1b)
        return ws.cell(row=ar, column=ac).number_format or ""
    ar, ac = _merged_top_left(ws, tpl_top, date_col_1b)
    return ws.cell(row=ar, column=ac).number_format or ""


def _detect_credit_leg_amount_style(ws, colmap: GlColumnMap, r_credit: int) -> str:
    """``standard``: credit in credit column; ``both_in_debit``: credit row amount in debit column."""
    cdeb = _openpyxl_col(colmap.debit)
    ccred = _openpyxl_col(colmap.credit)
    deb_raw = _cell_value_resolving_merge(ws, int(r_credit), cdeb)
    cred_raw = _cell_value_resolving_merge(ws, int(r_credit), ccred)
    vd = float(_try_float(deb_raw) or 0.0)
    vc = float(_try_float(cred_raw) or 0.0)
    if vd > 1e-6 and abs(vc) <= 1e-6:
        return "both_in_debit"
    return "standard"


def _prepare_gl_structure_from_last_template(
    ws,
    *,
    row_start: int,
    row_end: int,
    tpl_top: int,
    block_h: int,
    block_merges: list[tuple[int, int, int, int]],
    horizontal_by_tpl_row: dict[int, list[tuple[int, int, int, int]]],
    max_col: int,
) -> None:
    """
    Phase 1 of posting (no amounts yet): replicate the **last journal** layout.

    Copies font / fill / borders / alignment / number_format from cycling template rows, then reapplies
    the same merged regions — whole-entry vertical rectangles first follow per-row horizontal merges in
    the template (same order as before value writes).

    Caller must pass ``row_end = row_start + k - 1`` for the posting's line count ``k``.
    """
    k = row_end - row_start + 1
    if k <= 0:
        return

    for i in range(k):
        dest_row = row_start + i
        template_r = tpl_top + (i % block_h)
        _copy_row_style(ws, template_r, dest_row, max_col=max_col)

    for i in range(k):
        new_r = row_start + i
        template_r = tpl_top + (i % block_h)
        for _mr1h, mc1h, _mr2h, mc2h in horizontal_by_tpl_row.get(template_r, ()):
            ws.merge_cells(start_row=new_r, start_column=mc1h, end_row=new_r, end_column=mc2h)

    for _mr1_v, mc1_v, _mr2_v, mc2_v in block_merges:
        ws.merge_cells(start_row=row_start, start_column=mc1_v, end_row=row_end, end_column=mc2_v)


def append_double_entry(
    workbook_path: str,
    *,
    gl_date: date,
    description: str,
    debit_account: str,
    credit_account: str,
    amount: Decimal,
    sheet_name: str = GL_SHEET_NAME_DEFAULT,
    layout: Optional[dict[str, Any]] = None,
    currency_iso: str = DEFAULT_CURRENCY_ISO,
    transaction_number: Optional[Any] = None,
) -> None:
    """Append one balanced two-line entry (one debit, one credit leg)."""
    append_journal_entry(
        workbook_path,
        gl_date=gl_date,
        description=description,
        lines=[
            (debit_account, amount, Decimal("0")),
            (credit_account, Decimal("0"), amount),
        ],
        sheet_name=sheet_name,
        layout=layout,
        currency_iso=currency_iso,
        transaction_number=transaction_number,
    )


def append_journal_entry(
    workbook_path: str,
    *,
    gl_date: date,
    description: str,
    lines: list[tuple[str, Decimal, Decimal]],
    sheet_name: str = GL_SHEET_NAME_DEFAULT,
    layout: Optional[dict[str, Any]] = None,
    currency_iso: str = DEFAULT_CURRENCY_ISO,
    transaction_number: Optional[Any] = None,
) -> None:
    """
    Append a balanced compound journal: each line is (account, debit, credit) with exactly one side > 0.

    **Two phases** (single admin post, one save):

    1. **Structure** — Infer the last journal row band, copy row-level styles (alignment, formats,
       borders, …) and reapply the same merged regions for exactly ``k`` new rows. No business values yet.
    2. **Write** — Fill accounts, debits/credits, date, description, Tr. No., then balance-check and save.

    If ``transaction_number`` is set (after coercion), it is written to the Tr. No. column instead of
    inferring the next number from the row above.
    """
    if len(lines) < 2:
        raise ValueError("A journal entry needs at least two lines")
    debit_sum = Decimal("0")
    credit_sum = Decimal("0")
    norm: list[tuple[str, float, float]] = []
    for acct, deb, cred in lines:
        acct_s = (acct or "").strip()
        if not acct_s:
            raise ValueError("Each line needs an account")
        d = Decimal(deb)
        c = Decimal(cred)
        if d < 0 or c < 0:
            raise ValueError("Debit and credit amounts must be non-negative")
        if (d > Decimal("0") and c > Decimal("0")) or (d <= Decimal("0") and c <= Decimal("0")):
            raise ValueError(f"Each line needs exactly one positive amount: {acct_s!r}")
        debit_sum += d
        credit_sum += c
        norm.append((acct_s, float(d), float(c)))
    if debit_sum <= 0 or credit_sum <= 0:
        raise ValueError("Journal must have both debits and credits")
    if round(float(debit_sum), 2) != round(float(credit_sum), 2):
        raise ValueError(f"Debits {debit_sum} must equal credits {credit_sum}")

    ext = Path(workbook_path).suffix.lower()
    if ext not in APPENDABLE_EXTENSIONS:
        raise ValueError(
            f"Posting requires a linked .xlsx or .xlsm workbook (openpyxl); this file is {ext!r}. "
            "CSV and legacy .xls can be viewed from Financials but cannot receive approval posts."
        )

    wb = load_workbook(workbook_path, keep_vba=(ext == ".xlsm"))
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Sheet {sheet_name!r} not found; available: {wb.sheetnames}")
    ws = wb[sheet_name]

    mr = ws.max_row or 1
    max_c = min(ws.max_column or 1, 45)
    # Include enough rows so column detection + merged template match the real sheet tail.
    scan_r = min(ws.max_row or 1, max(60, mr + 2))
    raw_scan = _sheet_raw_cell_matrix(ws, scan_r, max_c)
    colmap = resolve_gl_column_map([tuple(r) for r in raw_scan], layout)
    grid = [list(r) for r in raw_scan]
    _apply_merged_cell_fill_to_grid(
        grid, ws, debit_col_0based=colmap.debit, credit_col_0based=colmap.credit
    )
    matrix: list[tuple[Any, ...]] = [tuple(row) for row in grid]
    colmap = resolve_gl_column_map(matrix, layout)

    date_c = _openpyxl_col(colmap.date)
    desc_c = _openpyxl_col(colmap.details)
    acct_c = _openpyxl_col(colmap.particulars)
    deb_c = _openpyxl_col(colmap.debit)
    cred_c = _openpyxl_col(colmap.credit)
    tr_c = _openpyxl_col(colmap.tr_number) if colmap.tr_number is not None else None

    tpl_top, tpl_bot = infer_last_entry_template_rows(ws, mr, date_c, desc_c)
    block_merges, horizontal_by_tpl_row = _collect_journal_template_merges(ws, tpl_top, tpl_bot)
    credit_style = _detect_credit_leg_amount_style(ws, colmap, tpl_bot)
    block_h = max(1, tpl_bot - tpl_top + 1)

    date_vert, _, _ = _column_in_template_vertical_block(block_merges, date_c, tpl_top, tpl_bot)
    desc_vert, _, _ = _column_in_template_vertical_block(block_merges, desc_c, tpl_top, tpl_bot)
    tr_vert = False
    if tr_c is not None:
        tr_vert, _, _ = _column_in_template_vertical_block(block_merges, tr_c, tpl_top, tpl_bot)

    date_tpl_r, date_tpl_anchor_c = _merged_top_left(ws, tpl_top, date_c)
    desc_tpl_r, desc_tpl_anchor_c = _merged_top_left(ws, tpl_top, desc_c)

    excel_date_val = _value_for_new_gl_date(ws, tpl_top, date_c, gl_date)
    date_number_format = (
        (_resolve_gl_date_number_format(ws, tpl_top, tpl_bot, date_c, mr) or "").strip() or "yyyy-mm-dd"
    )

    override_tr = coerce_gl_transaction_number(transaction_number)
    next_tr_val: Any = None
    tpl_tr_r, tpl_tr_anchor_c = tpl_top, tpl_top
    if tr_c is not None:
        tpl_tr_r, tpl_tr_anchor_c = _merged_top_left(ws, tpl_top, tr_c)
        if override_tr is not None:
            next_tr_val = override_tr
        else:
            prev_tr_raw = _cell_value_resolving_merge(ws, mr, tr_c)
            next_tr_val = _infer_next_tr_number(prev_tr_raw)
            if next_tr_val is None:
                next_tr_val = 1

    k = len(norm)
    row_start = mr + 1
    row_end = row_start + k - 1

    _prepare_gl_structure_from_last_template(
        ws,
        row_start=row_start,
        row_end=row_end,
        tpl_top=tpl_top,
        block_h=block_h,
        block_merges=block_merges,
        horizontal_by_tpl_row=horizontal_by_tpl_row,
        max_col=max_c,
    )

    for i in range(k):
        r = row_start + i
        template_r = tpl_top + (i % block_h)

        ac_write_col = acct_c
        for _mr1, mc1a, _mr2, mc2a in horizontal_by_tpl_row.get(template_r, ()):
            if mc1a <= acct_c <= mc2a:
                ac_write_col = min(ac_write_col, mc1a)

        acct_s, deb_f, cred_f = norm[i]
        if acct_s:
            _set_gl_account_on_row(
                ws,
                r,
                colmap,
                acct_s,
                tpl_top=tpl_top,
                tpl_bot=tpl_bot,
                credit_style=credit_style,
            )

        _write_gl_amounts_on_row(
            ws,
            colmap,
            r,
            debit=deb_f,
            credit=cred_f,
            credit_style=credit_style,
        )

        if date_vert:
            if i == 0:
                _set_sheet_cell_value_merge_aware(ws, int(row_start), date_tpl_anchor_c, excel_date_val)
                nf = _resolve_gl_date_number_format(ws, tpl_top, tpl_bot, date_c, r)
                _set_cell_number_format_merge_aware(
                    ws, int(row_start), date_tpl_anchor_c, (nf or "").strip() or "yyyy-mm-dd"
                )
        else:
            date_ar, date_ac = _merged_top_left(ws, r, date_c)
            _set_sheet_cell_value_merge_aware(ws, date_ar, date_ac, excel_date_val)
            nf = _resolve_gl_date_number_format(ws, tpl_top, tpl_bot, date_c, r)
            _set_cell_number_format_merge_aware(
                ws, date_ar, date_ac, (nf or "").strip() or "yyyy-mm-dd"
            )

        if desc_vert:
            if i == 0:
                _set_sheet_cell_value_merge_aware(ws, int(row_start), desc_tpl_anchor_c, description)
        else:
            desc_ar, desc_ac = _merged_top_left(ws, r, desc_c)
            _set_sheet_cell_value_merge_aware(ws, desc_ar, desc_ac, description)

        if tr_c is not None and next_tr_val is not None:
            if tr_vert:
                if i == 0:
                    _set_sheet_cell_value_merge_aware(ws, int(row_start), tpl_tr_anchor_c, next_tr_val)
            else:
                tr_ar, tr_ac = _merged_top_left(ws, r, tr_c)
                _set_sheet_cell_value_merge_aware(ws, tr_ar, tr_ac, next_tr_val)

        tpl_deb_nf = ws.cell(row=template_r, column=deb_c).number_format or ""
        tpl_cred_nf = ws.cell(row=template_r, column=cred_c).number_format or ""
        _set_cell_number_format_merge_aware(
            ws, r, deb_c, _amount_number_format_for_posting(tpl_deb_nf, currency_iso)
        )
        _set_cell_number_format_merge_aware(
            ws, r, cred_c, _amount_number_format_for_posting(tpl_cred_nf, currency_iso)
        )

        ac_tpl_r_a, ac_tpl_c_a = _merged_top_left(ws, template_r, acct_c)
        ac_write_col = particulars_columns_1b(ws, r, acct_c)[0] if acct_s else acct_c
        _mirror_anchor_number_format(ws, ac_tpl_r_a, ac_tpl_c_a, r, ac_write_col)

    if desc_vert:
        _mirror_anchor_number_format(ws, desc_tpl_r, desc_tpl_anchor_c, row_start, desc_tpl_anchor_c)
    else:
        for ii in range(k):
            rr = row_start + ii
            tpl_rr = tpl_top + (ii % block_h)
            dr_r, dc_c = _merged_top_left(ws, tpl_rr, desc_c)
            _mirror_anchor_number_format(ws, dr_r, dc_c, rr, dc_c)

    if tr_c is not None and next_tr_val is not None:
        tpl_tr_nf = ws.cell(row=tpl_tr_r, column=tpl_tr_anchor_c).number_format or ""
        if tr_vert:
            _set_cell_number_format_merge_aware(
                ws, row_start, tpl_tr_anchor_c, tpl_tr_nf or "General"
            )
        else:
            for ii in range(k):
                rr = row_start + ii
                tpl_rr = tpl_top + (ii % block_h)
                tr_mr, tr_mc = _merged_top_left(ws, tpl_rr, tr_c)
                _mirror_anchor_number_format(ws, tr_mr, tr_mc, rr, tr_mc)

    _assert_period_balanced(ws, row_start, row_end, colmap)
    wb.save(workbook_path)
    wb.close()


def _assert_period_balanced(ws, start_row: int, end_row: int, colmap: GlColumnMap) -> None:
    dcol = _openpyxl_col(colmap.debit)
    ccol = _openpyxl_col(colmap.credit)
    debit_sum = 0.0
    credit_sum = 0.0
    for r in range(start_row, end_row + 1):
        d = ws.cell(row=r, column=dcol).value or 0
        c = ws.cell(row=r, column=ccol).value or 0
        try:
            debit_sum += float(d)
            credit_sum += float(c)
        except (TypeError, ValueError) as e:
            raise ValueError(f"Non-numeric GL amounts at row {r}") from e
    if round(debit_sum, 2) != round(credit_sum, 2):
        raise ValueError(f"Debits {debit_sum} != Credits {credit_sum} for rows {start_row}-{end_row}")


def _first_journal_gl_date_in_matrix(
    matrix: list[tuple[Any, ...]],
    colmap: GlColumnMap,
    data_start_0based: int,
) -> Optional[date]:
    """First parseable posting date on or after the journal data band."""
    if colmap.date is None:
        return None
    for row in matrix[data_start_0based:]:
        d = _coerce_journal_sheet_date(_cell(row, colmap.date))
        if d is not None:
            return d
    return None


def _is_beginning_balance_record(rec: dict[str, Any]) -> bool:
    return gla.is_opening_or_brought_forward_rec(rec)


def tag_opening_and_brought_forward_gl_records(
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Mark brought-forward rows when the sheet text explicitly says so (never on plain amounts)."""
    out: list[dict[str, Any]] = []
    for rec in records:
        nr = dict(rec)
        acct = str(nr.get("account") or "").strip()
        desc = str(nr.get("description") or "").strip()
        hay = f"{acct} {desc}".strip()
        explicit = (
            gla.is_explicit_opening_balance_text(hay)
            or gla.is_explicit_opening_balance_text(desc)
            or gla.is_explicit_opening_balance_text(acct)
        )
        if gla.is_brought_forward_text(acct):
            m = _BROUGHT_FORWARD_ACCOUNT_RE.match(acct)
            if m and m.group(1).strip():
                nr["account"] = m.group(1).strip()
            nr["brought_forward"] = True
            if not gla.is_brought_forward_text(desc):
                nr["description"] = gla.BROUGHT_FORWARD_LABEL
            explicit = True
        elif gla.is_brought_forward_text(desc) or gla.is_brought_forward_text(nr.get("account")):
            nr["brought_forward"] = True
            if not gla.is_brought_forward_text(desc):
                nr["description"] = gla.BROUGHT_FORWARD_LABEL
            explicit = True
        if nr.get("opening_balance") and explicit:
            if nr.get("brought_forward"):
                nr["description"] = gla.BROUGHT_FORWARD_LABEL
            elif not desc or desc == BEGINNING_BALANCE_LABEL:
                nr["description"] = gla.BEGINNING_BALANCE_LABEL
        elif nr.get("opening_balance") and not explicit:
            nr.pop("opening_balance", None)
            nr.pop("brought_forward", None)
        if nr.get("brought_forward") or (nr.get("opening_balance") and explicit):
            nr["opening_balance"] = True
        out.append(nr)
    return out


def extract_opening_balance_gl_rows(
    matrix: list[tuple[Any, ...]],
    colmap: GlColumnMap,
    *,
    amt_tol: float = 0.015,
) -> list[dict[str, Any]]:
    """
    Opening / brought-forward rows above the journal ``data_start_row`` (common Finjan layout).

    These carry account balances (e.g. cash ~6,167) that never appear as dated journal lines but
    must be included in trial-balance totals.
    """
    ds_excel = max(1, int(colmap.data_start_row))
    if ds_excel <= 1:
        return []
    first_journal_date = _first_journal_gl_date_in_matrix(
        matrix, colmap, max(0, ds_excel - 1)
    )
    header_skip = [False]
    out: list[dict[str, Any]] = []
    for excel_1b in range(1, ds_excel):
        row_idx = excel_1b - 1
        if row_idx >= len(matrix):
            break
        row = matrix[row_idx]
        rec = _row_tuple_to_record_dynamic(
            row, colmap, header_skipped_ref=header_skip, excel_row_1based=excel_1b
        )
        if not rec:
            continue
        deb = row_amount(rec.get("debit"))
        cre = row_amount(rec.get("credit"))
        hay = f"{rec.get('account') or ''} {rec.get('description') or ''}"
        cell_bits: list[str] = [hay]
        for c0 in (colmap.debit, colmap.credit, colmap.details, colmap.particulars):
            if c0 is None:
                continue
            cell_bits.append(str(_cell(row, c0) or ""))
        row_hay = " ".join(cell_bits)
        if not gla.is_explicit_opening_balance_text(row_hay):
            continue
        if deb <= amt_tol and cre <= amt_tol:
            from t_account_sheet import _parse_ta_amount

            if colmap.debit is not None:
                d_amt, d_iso, _ = _parse_ta_amount(_cell(row, colmap.debit))
                if d_amt is not None and abs(float(d_amt)) > amt_tol:
                    deb = float(d_amt)
                    rec = dict(rec)
                    rec["debit"] = deb
                    if d_iso:
                        rec["currency_iso"] = d_iso
            if cre <= amt_tol and colmap.credit is not None:
                c_amt, c_iso, _ = _parse_ta_amount(_cell(row, colmap.credit))
                if c_amt is not None and abs(float(c_amt)) > amt_tol:
                    cre = float(c_amt)
                    rec = dict(rec)
                    rec["credit"] = cre
                    if c_iso:
                        rec["currency_iso"] = c_iso
        if deb <= amt_tol and cre <= amt_tol:
            continue
        if is_non_posting_gl_row(rec, amt_tol=amt_tol):
            continue
        nr = dict(rec)
        nr["opening_balance"] = True
        if gla.is_brought_forward_text(row_hay):
            nr["brought_forward"] = True
            nr["description"] = gla.BROUGHT_FORWARD_LABEL
        else:
            nr["description"] = gla.BEGINNING_BALANCE_LABEL
        if first_journal_date is not None:
            nr["gl_date"] = first_journal_date.isoformat()
        out.append(nr)
    return out


def _supplement_matrix_formula_amounts(
    workbook_path: str,
    sheet_name: str,
    grid: list[list[Any]],
    colmap: GlColumnMap,
    *,
    data_start_0based: int,
) -> None:
    """Fill debit/credit matrix cells when formulas have no cached ``data_only`` value yet."""
    amt_cols = [c for c in (colmap.debit, colmap.credit) if c is not None and int(c) >= 0]
    if not amt_cols:
        return
    needs = False
    for ri in range(0, len(grid)):
        row = grid[ri]
        for c0 in amt_cols:
            if c0 >= len(row):
                continue
            amt, _ = parse_money_cell(row[c0])
            if amt is None or abs(float(amt)) <= 1e-9:
                needs = True
                break
        if needs:
            break
    if not needs:
        return

    wb_f = wb_v = None
    try:
        wb_f = load_workbook(workbook_path, data_only=False, read_only=False)
        wb_v = load_workbook(workbook_path, data_only=True, read_only=False)
        if sheet_name not in wb_f.sheetnames:
            return
        ws_f = wb_f[sheet_name]
        ws_v = wb_v[sheet_name]
        for ri in range(0, len(grid)):
            excel_1b = ri + 1
            row = grid[ri]
            for c0 in amt_cols:
                if c0 >= len(row):
                    continue
                cur_amt, _ = parse_money_cell(row[c0])
                if cur_amt is not None and abs(float(cur_amt)) > 1e-9:
                    continue
                if _is_excel_formula_value(row[c0]):
                    ev = _eval_simple_amount_formula(ws_f, ws_v, str(row[c0]))
                    if ev is not None:
                        row[c0] = ev
                        continue
                c1b = _openpyxl_col(int(c0))
                fr = _physical_cell_value(ws_f, excel_1b, c1b)
                if not _is_excel_formula_value(fr):
                    continue
                ev = _eval_simple_amount_formula(ws_f, ws_v, str(fr))
                if ev is not None:
                    row[c0] = ev
    finally:
        _close_gl_workbooks(wb_f, wb_v)


def enrich_gl_records_with_tac_openings(
    records: list[dict[str, Any]],
    workbook_path: str,
    *,
    gl_sheet_name: str,
    t_accounts_sheet_name: Optional[str] = None,
    default_year: Optional[int] = None,
    keep_excel_row: bool = False,
) -> list[dict[str, Any]]:
    """
    Analytics-only: merge explicit T-account Beg Balance rows into a copy for trial balance / charts.

    Does not modify the journal worksheet; synthetic rows are tagged ``source=t_account_sheet``.
  """
    from t_account_sheet import merge_t_account_opening_supplements, read_t_account_opening_from_workbook

    try:
        supp = read_t_account_opening_from_workbook(
            workbook_path,
            gl_sheet_name=gl_sheet_name,
            t_accounts_sheet_name=t_accounts_sheet_name,
            default_year=default_year,
        )
    except Exception:
        return list(records)
    if not supp:
        return list(records)
    merged = merge_t_account_opening_supplements(records, supp)
    return finalize_gl_records(merged, keep_excel_row=keep_excel_row)


def read_gl_sheet_rows(
    workbook_path: str,
    *,
    sheet_name: str = GL_SHEET_NAME_DEFAULT,
    tail: int = 500,
    layout: Optional[dict[str, Any]] = None,
    keep_excel_row: bool = False,
) -> list[dict[str, Any]]:
    """Read GL / journal rows using header keyword detection.

    ``tail``: max **logical** postings to retain after skipping headers. ``tail <= 0`` keeps **every**
    parsed row — required for dormant balance-sheet accounts whose last activity sits far above the
    sheet tail. Omitting historical lines made inactive payables look empty while frequently posted
    accounts (e.g. cheques payable) still had balances.
    """

    wb = load_workbook(workbook_path, read_only=False, data_only=True)
    if sheet_name not in wb.sheetnames:
        names = wb.sheetnames
        wb.close()
        raise KeyError(f"Sheet {sheet_name!r} not found; available: {names}")
    ws = wb[sheet_name]
    max_r = min(ws.max_row or 1, _MAX_GL_PHYSICAL_ROWS_SOFT_CAP)
    max_c = min(ws.max_column or 1, 45)
    raw_gl = _sheet_raw_cell_matrix(ws, max_r, max_c)
    colmap = resolve_gl_column_map([tuple(r) for r in raw_gl], layout)
    grid = [list(r) for r in raw_gl]
    _apply_merged_cell_fill_to_grid(
        grid, ws, debit_col_0based=colmap.debit, credit_col_0based=colmap.credit
    )
    matrix: list[tuple[Any, ...]] = [tuple(row) for row in grid]

    colmap = resolve_gl_column_map(matrix, layout)
    ds = max(0, colmap.data_start_row - 1)
    _supplement_matrix_formula_amounts(
        workbook_path,
        sheet_name,
        grid,
        colmap,
        data_start_0based=ds,
    )
    matrix = [tuple(row) for row in grid]

    unlimited = tail is None or int(tail) <= 0
    buf: deque[dict[str, Any]] | list[dict[str, Any]]
    if unlimited:
        buf = []
    else:
        buf = deque(maxlen=max(int(tail), 1))
    header_skip = [False]
    opening_rows = extract_opening_balance_gl_rows(matrix, colmap)
    for offset, row in enumerate(matrix[ds:]):
        excel_1b = ds + offset + 1
        rec = _row_tuple_to_record_dynamic(
            row, colmap, header_skipped_ref=header_skip, excel_row_1based=excel_1b
        )
        if rec:
            try:
                _enrich_record_currency_from_workbook_formats(rec, ws, excel_1b, colmap)
            except Exception:
                _ensure_original_currency_fallback(rec)
            buf.append(rec)  # type: ignore[union-attr]
    wb.close()
    merged_buf = opening_rows + list(buf)
    staged = apply_spatial_credit_from_matrix(matrix, merged_buf, colmap)
    return finalize_gl_records(staged, keep_excel_row=keep_excel_row)


def read_gl_csv_rows(
    path: str,
    *,
    tail: int = 500,
    layout: Optional[dict[str, Any]] = None,
    keep_excel_row: bool = False,
) -> list[dict[str, Any]]:
    """Read CSV with the same keyword-based header detection as Excel (``tail <= 0`` = all rows)."""
    import csv

    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        f.seek(0)
        reader = csv.reader(f, dialect)
        raw_list = list(reader)[:_MAX_GL_PHYSICAL_ROWS_SOFT_CAP]

    if not raw_list:
        return []

    width = min(45, max(len(r) for r in raw_list))
    matrix: list[tuple[Any, ...]] = [tuple((list(r) + [None] * width)[:width]) for r in raw_list]

    colmap = resolve_gl_column_map(matrix, layout)
    ds = max(0, colmap.data_start_row - 1)

    unlimited = tail is None or int(tail) <= 0
    buf: deque[dict[str, Any]] | list[dict[str, Any]]
    if unlimited:
        buf = []
    else:
        buf = deque(maxlen=max(int(tail), 1))
    header_skip = [False]
    opening_rows = extract_opening_balance_gl_rows(matrix, colmap)
    for offset, row in enumerate(matrix[ds:]):
        excel_1b = ds + offset + 1
        rec = _row_tuple_to_record_dynamic(
            row, colmap, header_skipped_ref=header_skip, excel_row_1based=excel_1b
        )
        if rec:
            _ensure_original_currency_fallback(rec)
            buf.append(rec)  # type: ignore[union-attr]
    merged_buf = opening_rows + list(buf)
    staged = apply_spatial_credit_from_matrix(matrix, merged_buf, colmap)
    return finalize_gl_records(staged, keep_excel_row=keep_excel_row)


def read_gl_xls_rows(
    path: str,
    *,
    sheet_name: str = GL_SHEET_NAME_DEFAULT,
    tail: int = 500,
    layout: Optional[dict[str, Any]] = None,
    keep_excel_row: bool = False,
) -> list[dict[str, Any]]:
    """Read legacy .xls GL sheet via xlrd (pandas) + header detection (``tail <= 0`` = all rows)."""
    import pandas as pd

    try:
        df = pd.read_excel(path, sheet_name=sheet_name, header=None, engine="xlrd")
    except ValueError:
        df = pd.read_excel(path, sheet_name=0, header=None, engine="xlrd")

    raw_list: list[list[Any]] = []
    _, ncols_df = df.shape
    ncol = min(45, int(ncols_df))
    for _, row in df.iterrows():
        lst: list[Any] = []
        for i in range(ncol):
            v = row.iloc[i] if i < len(row) else None
            lst.append(None if pd.isna(v) else v)
        raw_list.append(lst)

    if not raw_list:
        return []

    width = min(45, max(len(r) for r in raw_list))
    matrix = [tuple((list(r) + [None] * width)[:width]) for r in raw_list[:_MAX_GL_PHYSICAL_ROWS_SOFT_CAP]]

    colmap = resolve_gl_column_map(matrix, layout)
    ds = max(0, colmap.data_start_row - 1)

    unlimited = tail is None or int(tail) <= 0
    buf: deque[dict[str, Any]] | list[dict[str, Any]]
    if unlimited:
        buf = []
    else:
        buf = deque(maxlen=max(int(tail), 1))
    header_skip = [False]
    opening_rows = extract_opening_balance_gl_rows(matrix, colmap)
    for offset, row in enumerate(matrix[ds:]):
        excel_1b = ds + offset + 1
        rec = _row_tuple_to_record_dynamic(
            row, colmap, header_skipped_ref=header_skip, excel_row_1based=excel_1b
        )
        if rec:
            _ensure_original_currency_fallback(rec)
            buf.append(rec)  # type: ignore[union-attr]
    merged_buf = opening_rows + list(buf)
    staged = apply_spatial_credit_from_matrix(matrix, merged_buf, colmap)
    return finalize_gl_records(staged, keep_excel_row=keep_excel_row)


def partition_journal_blocks(records: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    """
    Split parsed GL rows into visual journal entries.

    A row with a non-blank ``transaction_number`` starts a new block; continuation rows
    (blank Tr. #) belong to the preceding block.
    """
    ordered = sorted(records, key=lambda r: int(r.get("_excel_row") or 0))
    blocks: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    for rec in ordered:
        tr = rec.get("transaction_number")
        has_tr = tr is not None and str(tr).strip() != ""
        if has_tr and current:
            blocks.append(current)
            current = []
        current.append(rec)
    if current:
        blocks.append(current)
    return blocks


def journal_block_balanced(block: list[dict[str, Any]], *, amt_tol: float = 0.015) -> bool:
    """True when total debits and credits in one journal block match within tolerance."""
    if not block:
        return True
    deb = sum(row_amount(r.get("debit")) for r in block)
    cre = sum(row_amount(r.get("credit")) for r in block)
    scale = max(amt_tol, 0.01 * max(abs(deb), abs(cre), 1.0))
    return abs(deb - cre) <= scale


def _balanced_journal_record_ids(records: list[dict[str, Any]], *, amt_tol: float = 0.015) -> set[int]:
    """``id(rec)`` for rows in compound entries that already tie debits to credits."""
    out: set[int] = set()
    for block in partition_journal_blocks(records):
        if journal_block_balanced(block, amt_tol=amt_tol):
            out.update(id(r) for r in block)
    return out


def apply_spatial_credit_from_matrix(
    matrix: list[tuple[Any, ...]],
    records: list[dict[str, Any]],
    colmap: GlColumnMap,
    *,
    amt_tol: float = 0.015,
) -> list[dict[str, Any]]:
    """
    For debit-only lines, recover credits entered one column right of the debit column,
    starting one row below, continuing downward while Tr. No. is blank until amounts balance.
    Each non-zero scanned row becomes its own synthetic credit leg (multiple split credits keep
    their row's account labels instead of collapsing onto the final fragment).

    Requires ``colmap.tr_number`` (auto-detected or set in Configuration).
    """
    if colmap.tr_number is None:
        return records
    width = _matrix_width(matrix)
    amt_col = colmap.debit + 1
    if amt_col >= width or colmap.debit < 0:
        return records
    trcol = colmap.tr_number
    nrows = len(matrix)

    def tr_blank(row: tuple[Any, ...]) -> bool:
        v = _cell(row, trcol)
        return v is None or str(v).strip() == ""

    consumed: set[int] = set()
    extras: list[dict[str, Any]] = []
    skip_spatial = _balanced_journal_record_ids(records, amt_tol=amt_tol)
    block_by_rec_id: dict[int, list[dict[str, Any]]] = {}
    for block in partition_journal_blocks(records):
        for rec in block:
            block_by_rec_id[id(rec)] = block

    def _block_has_explicit_credit_column(block: list[dict[str, Any]]) -> bool:
        for item in block:
            er = item.get("_excel_row")
            if not isinstance(er, int) or er < 1 or er > nrows:
                continue
            cre_v = _try_float(_cell(matrix[er - 1], colmap.credit))
            if cre_v is not None and abs(float(cre_v)) > amt_tol:
                return True
        return False

    sorted_for_scan = sorted(
        records,
        key=lambda r: int(r.get("_excel_row") or 0),
    )

    for rec in sorted_for_scan:
        if id(rec) in skip_spatial:
            continue
        block = block_by_rec_id.get(id(rec))
        if block and _block_has_explicit_credit_column(block):
            deb_amt = row_amount(rec.get("debit"))
            cre_amt = row_amount(rec.get("credit"))
            if deb_amt > amt_tol and cre_amt <= amt_tol:
                continue
        er = rec.get("_excel_row")
        if not isinstance(er, int):
            continue
        deb_amt = row_amount(rec.get("debit"))
        cre_amt = row_amount(rec.get("credit"))
        if deb_amt <= amt_tol or cre_amt > amt_tol:
            continue

        mi = er - 1
        if mi < 0 or mi >= nrows:
            continue

        cum = 0.0
        frag_credit_rows: list[tuple[int, float]] = []
        step = 0
        aborted = False
        while step < 80:
            ridx = mi + 1 + step
            if ridx >= nrows:
                break
            excel_here = ridx + 1
            if excel_here in consumed:
                aborted = True
                break
            row_t = matrix[ridx]
            if not tr_blank(row_t):
                break
            raw_amt = _try_float(_cell(row_t, amt_col))
            if raw_amt is None or abs(float(raw_amt)) <= amt_tol:
                if step == 0:
                    break
                step += 1
                continue
            chunk = float(raw_amt)
            cum += chunk
            frag_credit_rows.append((excel_here, chunk))
            rel = max(amt_tol, 0.01 * max(deb_amt, cum, 1.0))
            if abs(cum - deb_amt) <= rel:
                break
            step += 1

        if aborted:
            continue

        rel = max(amt_tol, 0.01 * max(deb_amt, cum, 1.0))
        if cum <= amt_tol or abs(cum - deb_amt) > rel or not frag_credit_rows:
            continue

        for fr, _amt in frag_credit_rows:
            consumed.add(fr)

        parent_desc_fallback = str(rec.get("description") or "")
        curr_iso = str(rec.get("currency_iso") or DEFAULT_CURRENCY_ISO)
        ocur = str(rec.get("original_currency") or curr_iso or DEFAULT_CURRENCY_ISO)
        for fr_excel, fr_amt in frag_credit_rows:
            row_ix = max(0, min(fr_excel - 1, nrows - 1))
            slice_t = matrix[row_ix]
            acct = str(_effective_particulars_account_label(slice_t, colmap) or "").strip()
            det = str(_cell(slice_t, colmap.details) or "").strip()
            extras.append(
                {
                    "gl_date": rec.get("gl_date", ""),
                    "description": det or parent_desc_fallback,
                    "account": acct,
                    "debit": 0.0,
                    "credit": float(fr_amt),
                    "currency_iso": curr_iso,
                    "original_currency": ocur,
                    "original_amount": float(fr_amt),
                    "_excel_row": fr_excel,
                }
            )

    if not consumed and not extras:
        return records

    out: list[dict[str, Any]] = []
    for r in records:
        er = r.get("_excel_row")
        if isinstance(er, int) and er in consumed:
            continue
        out.append(r)
    out.extend(extras)
    out.sort(key=lambda x: int(x.get("_excel_row") or 0))
    return out


def _is_narrative_memo_account_label(account: str) -> bool:
    """Parenthetical / free-text notes parked in the account column (not a real ledger account)."""
    s = (account or "").strip()
    if not s:
        return False
    if s.startswith("(") and s.endswith(")"):
        return True
    return len(s) > 64


def merge_backward_amountless_journal_notes(
    records: list[dict[str, Any]],
    *,
    amt_tol: float = 0.015,
) -> list[dict[str, Any]]:
    """
    Rows with no non‑trivial debit or credit amounts are memo lines — often the amount cells are
    blank or ``0.00`` — not separate GL legs. Fold their text onto the preceding **posting row**
    ``description`` (common when the memo sits under split credits and uses the particulars column).

    Rows that still carry a real **account** label (including split-credit continuation lines whose
    amounts live in an adjacent column) are kept so deleting one journal leg does not cascade into
    dropping the rest of the entry on reload.

    Rows are processed in workbook order when ``_excel_row`` is set; otherwise list order applies.
    """
    if not records:
        return records

    def _order_key(rr: dict[str, Any]) -> tuple[int, int]:
        er = rr.get("_excel_row")
        if isinstance(er, int):
            return (0, er)
        return (1, id(rr))

    def _header_like_account(s: str) -> bool:
        return _normalize_headerish_label(s) in _JOURNAL_HEADER_ACCOUNT_LABELS

    sorted_recs = sorted(records, key=_order_key)
    out: list[dict[str, Any]] = []
    for rec in sorted_recs:
        d_amt = row_amount(rec.get("debit"))
        c_amt = row_amount(rec.get("credit"))
        negligible = d_amt <= amt_tol and c_amt <= amt_tol
        if negligible:
            ac = str(rec.get("account") or "").strip()
            desc_raw = str(rec.get("description") or "").strip()
            if _header_like_account(ac) and not desc_raw:
                continue
            if ac and not _is_narrative_memo_account_label(ac):
                out.append(dict(rec))
                continue
            parts: list[str] = []
            if desc_raw:
                parts.append(desc_raw)
            if ac and ac != desc_raw:
                parts.append(ac)
            nar = "; ".join(parts).strip()
            if nar and out:
                prev = out[-1]
                base_desc = str(prev.get("description") or "").strip()
                if nar not in base_desc:
                    sep = "; " if base_desc else ""
                    prev["description"] = base_desc + sep + nar
                continue
            if not nar:
                continue
            # Orphan memo (no preceding row): drop from GL lines
            continue

        nr = dict(rec)
        out.append(nr)
    return out


def finalize_gl_records(
    records: list[dict[str, Any]], *, keep_excel_row: bool = False
) -> list[dict[str, Any]]:
    """Repair common workbook patterns: credits entered under Debit column; split narrative rows."""
    r = filter_non_posting_gl_rows(records)
    r = normalize_adjacent_debit_only_journal_legs(r)
    r = merge_backward_amountless_journal_notes(r)
    r = fold_consecutive_description_only_lines(r)
    r = annotate_journal_transaction_groups(r)
    r = filter_non_posting_gl_rows(r)
    r = tag_opening_and_brought_forward_gl_records(r)
    if not keep_excel_row:
        for x in r:
            x.pop("_excel_row", None)
    return r


def annotate_journal_transaction_groups(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Tag consecutive lines that belong to one **visual journal** in the sheet.

    After merged cells are expanded in the matrix, several rows often share the same propagated
    ``gl_date`` and ``description``; this groups them so dashboards can treat one logical entry.
    """
    if not records:
        return records

    def row_order(rec: dict[str, Any]) -> int:
        return int(rec.get("_excel_row") or 0)

    def norm_date(rec: dict[str, Any]) -> str:
        d = rec.get("gl_date")
        if d is not None and hasattr(d, "isoformat"):
            return str(d.isoformat())[:10]
        return str(d or "").strip()[:10]

    sorted_recs = sorted(records, key=row_order)
    out: list[dict[str, Any]] = []
    gid = 0
    i = 0
    n = len(sorted_recs)
    while i < n:
        gid += 1
        base = sorted_recs[i]
        dk = norm_date(base)
        desc = str(base.get("description") or "").strip()
        block: list[dict[str, Any]] = [base]
        j = i + 1
        while j < n:
            nxt = sorted_recs[j]
            if norm_date(nxt) == dk and str(nxt.get("description") or "").strip() == desc:
                block.append(nxt)
                j += 1
            else:
                break
        lines_in_block = len(block)
        rows_nums = [int(x.get("_excel_row") or 0) for x in block if int(x.get("_excel_row") or 0) > 0]
        span_s = ""
        if rows_nums:
            lo, hi = min(rows_nums), max(rows_nums)
            span_s = f"{lo}-{hi}" if lo != hi else str(lo)
        for li, rec in enumerate(block, 1):
            nr = dict(rec)
            nr["journal_entry_seq"] = gid
            nr["journal_line_in_entry"] = li
            nr["journal_lines_in_entry"] = lines_in_block
            if span_s:
                nr["journal_excel_row_span"] = span_s
            out.append(nr)
        i = j
    return out


def normalize_adjacent_debit_only_journal_legs(
    records: list[dict[str, Any]],
    *,
    amt_tol: float = 0.015,
) -> list[dict[str, Any]]:
    """
    Some templates put the credit leg's amount in the **Debit** column on the following row.
    When two consecutive lines both show as debit-only and amounts balance, move the second
    row's debit into ``credit`` so double-entry analytics and journal pairing work.

    Skip compound journals that already have credit-column legs (e.g. Cash + A/R debits with
    Unearned credit below, or paired salary debits with a Cash credit in the same entry).
    """
    if len(records) < 2:
        return records
    block_by_rec_id: dict[int, list[dict[str, Any]]] = {}
    for block in partition_journal_blocks(records):
        for rec in block:
            block_by_rec_id[id(rec)] = block

    def _block_allows_debit_pair_flip(r0: dict[str, Any], r1: dict[str, Any]) -> bool:
        block = block_by_rec_id.get(id(r0))
        if not block or block_by_rec_id.get(id(r1)) is not block:
            return False
        credit_sum = sum(row_amount(r.get("credit")) for r in block)
        if credit_sum > amt_tol:
            return False
        amount_lines = sum(
            1
            for r in block
            if row_amount(r.get("debit")) > amt_tol or row_amount(r.get("credit")) > amt_tol
        )
        return amount_lines == 2

    out: list[dict[str, Any]] = []
    i = 0
    n = len(records)
    while i < n:
        if i + 1 < n:
            r0 = records[i]
            r1 = records[i + 1]
            d0, c0 = row_amount(r0.get("debit")), row_amount(r0.get("credit"))
            d1, c1 = row_amount(r1.get("debit")), row_amount(r1.get("credit"))
            date0 = str(r0.get("gl_date") or "").strip()
            date1 = str(r1.get("gl_date") or "").strip()
            ac0 = str(r0.get("account") or "").strip()
            ac1 = str(r1.get("account") or "").strip()
            dust0 = max(amt_tol, 0.001 * max(abs(d0), abs(c0), 1.0))
            dust1 = max(amt_tol, 0.001 * max(abs(d1), abs(c1), 1.0))
            already_standard = (
                d0 > amt_tol
                and c0 <= dust0
                and c1 > amt_tol
                and d1 <= dust1
                and _journal_dates_compatible(date0, date1)
            )
            if already_standard:
                out.append(r0)
                out.append(r1)
                i += 2
                continue
            both_in_debit_column = (
                d0 > amt_tol
                and c0 <= dust0
                and d1 > amt_tol
                and c1 <= dust1
                and _journal_dates_compatible(date0, date1)
                and ac0 != ac1
                and _paired_amounts_match(d0, d1, amt_tol=amt_tol)
                and _block_allows_debit_pair_flip(r0, r1)
            )
            if both_in_debit_column:
                r1n = dict(r1)
                r1n["credit"] = float(d1)
                r1n["debit"] = 0.0
                out.append(r0)
                out.append(r1n)
                i += 2
                continue
        out.append(records[i])
        i += 1
    return out


def _is_description_only_row(rec: dict[str, Any], *, amt_tol: float) -> bool:
    d, c = row_amount(rec.get("debit")), row_amount(rec.get("credit"))
    ac = str(rec.get("account") or "").strip()
    desc = str(rec.get("description") or "").strip()
    return d <= amt_tol and c <= amt_tol and not ac and bool(desc)


def _looks_like_counterparty_account_line(name: str) -> bool:
    """Heuristic: short counterparty / bank line often placed below the expense description."""
    n = (name or "").strip().lower()
    if not n or len(n) > 48:
        return False
    return any(
        k in n
        for k in (
            "bank",
            "cash",
            "paypal",
            "wallet",
            "fawran",
            "western union",
            "union",
            "transfer",
            "clearing",
            "payable",
            "receivable",
        )
    )


def fold_consecutive_description_only_lines(
    records: list[dict[str, Any]],
    *,
    amt_tol: float = 0.015,
) -> list[dict[str, Any]]:
    """
    Merge runs of rows that have narrative in ``description`` but no account and no amounts
    (often caused by merged cells or split entry). If exactly two fragments match the pattern
    expense line + bank/counterparty, map the second to ``account``.
    """
    out: list[dict[str, Any]] = []
    i = 0
    n = len(records)
    while i < n:
        if _is_description_only_row(records[i], amt_tol=amt_tol):
            descs: list[str] = []
            j = i
            while j < n and _is_description_only_row(records[j], amt_tol=amt_tol):
                t = str(records[j].get("description") or "").strip()
                if t:
                    descs.append(t)
                j += 1
            if len(descs) >= 2:
                base = dict(records[i])
                d0, d1 = descs[0], descs[1]
                if len(descs) == 2 and _looks_like_counterparty_account_line(d1) and not _looks_like_counterparty_account_line(d0):
                    base["account"] = d1
                    base["description"] = d0
                else:
                    base["description"] = "; ".join(descs)
                    base["account"] = ""
                # Preserve earliest non-empty date from the run
                for k in range(i, j):
                    gv = str(records[k].get("gl_date") or "").strip()
                    if gv:
                        base["gl_date"] = records[k].get("gl_date")
                        break
                out.append(base)
                i = j
                continue
            if len(descs) == 1:
                out.append(records[i])
                i = j
                continue
        out.append(records[i])
        i += 1
    return out


def row_amount(val: Any) -> float:
    f = _try_float(val)
    return float(f) if f is not None else 0.0


def _coerce_journal_sheet_date(val: Any) -> Optional[date]:
    """Like :func:`parse_gl_cell_to_date`, but lifts two-digit / missing years to the current calendar year."""
    d = parse_gl_cell_to_date(val)
    if d is None:
        return None
    if d.year < 1900:
        try:
            return d.replace(year=date.today().year)
        except ValueError:
            return d
    return d


def parse_gl_cell_to_date(val: Any) -> Optional[date]:
    """Normalize Excel / CSV cell values to a calendar date."""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        x = float(val)
        # ``data_only`` workbooks often cache dates as Excel serial numbers (plain floats).
        if 300 < abs(x) < 1_200_000:
            try:
                from openpyxl.utils.datetime import from_excel

                dtv = from_excel(x)
                if isinstance(dtv, datetime):
                    return dtv.date()
                if isinstance(dtv, date):
                    return dtv
            except Exception:
                pass

    if isinstance(val, str):
        st = val.strip()
        if st and any(ch.isdigit() for ch in st[:14]):
            try:
                x = float(st.replace(",", ""))
                if 300 < abs(x) < 1_200_000:
                    from openpyxl.utils.datetime import from_excel

                    dtv = from_excel(x)
                    if isinstance(dtv, datetime):
                        return dtv.date()
                    if isinstance(dtv, date):
                        return dtv
            except Exception:
                pass

    import pandas as pd

    ts = pd.to_datetime(val, errors="coerce")
    if pd.isna(ts):
        ts = pd.to_datetime(str(val).strip(), errors="coerce")
    if pd.isna(ts):
        return None
    return ts.date()


def gl_flat_records_activity_dataframe(
    records: list[dict[str, Any]],
    fiscal_start_month: int,
    *,
    table_rates_foreign_to_usd: Optional[dict[str, float]] = None,
):
    """DataFrame with original_currency, original_amount, USD columns for charts."""
    import pandas as pd

    import database as db
    import fiscal

    rows_out: list[dict[str, Any]] = []
    last_d: Optional[date] = None
    for r in records:
        gd = _coerce_journal_sheet_date(r.get("gl_date"))
        if gd is None:
            gd = last_d
        else:
            last_d = gd
        if gd is None and _is_beginning_balance_record(r):
            gd = date.today()
        if gd is None:
            continue
        fp = fiscal.fiscal_period_for(gd, fiscal_start_month)
        d_amt = row_amount(r.get("debit"))
        c_amt = row_amount(r.get("credit"))
        oc_raw = str(r.get("original_currency") or r.get("currency_iso") or DEFAULT_CURRENCY_ISO).strip().upper()
        oc = oc_raw[:3] if len(oc_raw) >= 3 else DEFAULT_CURRENCY_ISO
        oa = r.get("original_amount")
        if oa is None:
            oa = d_amt + c_amt
        else:
            try:
                oa = float(oa)
            except (TypeError, ValueError):
                oa = d_amt + c_amt
        rate_usd = db.get_conversion_rate(
            oc, "USD", gd, table_rates_foreign_to_usd=table_rates_foreign_to_usd
        )
        debit_usd = d_amt * rate_usd
        credit_usd = c_amt * rate_usd
        base_currency_amount = float(oa) * rate_usd
        row_extra: dict[str, Any] = {}
        for key in ("journal_entry_seq", "journal_line_in_entry", "journal_lines_in_entry"):
            v = r.get(key)
            if v is None or str(v).strip() == "":
                continue
            try:
                row_extra[key] = int(v)
            except (TypeError, ValueError):
                continue
        rows_out.append(
            {
                "gl_date": gd.isoformat(),
                "description": str(r.get("description") or ""),
                "account": str(r.get("account") or ""),
                "debit": d_amt,
                "credit": c_amt,
                "fiscal_year": fp.fiscal_year,
                "fiscal_period": fp.fiscal_period,
                "activity": d_amt + c_amt,
                "currency_iso": oc,
                "original_currency": oc,
                "original_amount": float(oa),
                "debit_usd": debit_usd,
                "credit_usd": credit_usd,
                "base_currency_amount": base_currency_amount,
                "opening_balance": bool(r.get("opening_balance")),
                "brought_forward": bool(r.get("brought_forward")),
                **row_extra,
            }
        )
    return pd.DataFrame(rows_out)


_DISPLAY_SYM = {
    "USD": "$",
    "EUR": "€",
    "GBP": "£",
    "JPY": "¥",
    "INR": "₹",
    "CAD": "C$",
    "AUD": "A$",
}


def format_amount_display(amount: float, display_iso: str) -> str:
    try:
        a = float(amount)
    except (TypeError, ValueError):
        return ""
    if not math.isfinite(a) or abs(a) < 1e-11:
        return ""
    iso = (display_iso or DEFAULT_CURRENCY_ISO).upper()
    sym = _DISPLAY_SYM.get(iso, f"{iso} ")
    return f"{sym}{a:,.2f}"


def _credit_particulars_display(acct: str) -> str:
    a = (acct or "").strip()
    if not a:
        return ""
    if a.lower().startswith("to "):
        return a
    return f"To {a}"


def _row_currency_iso_for_format(r: dict[str, Any], fallback_iso: str) -> str:
    raw = str(r.get("currency_iso") or "").strip().upper()
    if len(raw) >= 3:
        return raw[:3]
    fb = (fallback_iso or DEFAULT_CURRENCY_ISO).strip().upper()
    return fb[:3] if len(fb) >= 3 else DEFAULT_CURRENCY_ISO


def _journal_dates_compatible(date_s: str, date_s2: str) -> bool:
    return (not date_s2.strip()) or (not date_s.strip()) or (date_s.strip() == date_s2.strip())


def _one_sided_debit(d: float, c: float, *, amt_tol: float) -> bool:
    """True if row is treated as a debit line (non-trivial debit, credit is negligible or formula dust)."""
    if d <= amt_tol:
        return False
    dust = max(amt_tol, 0.001 * max(abs(d), abs(c), 1.0))
    return c <= dust


def _one_sided_credit(d: float, c: float, *, amt_tol: float) -> bool:
    if c <= amt_tol:
        return False
    dust = max(amt_tol, 0.001 * max(abs(d), abs(c), 1.0))
    return d <= dust


def _paired_amounts_match(d_debit: float, c_credit: float, *, amt_tol: float) -> bool:
    rel = max(amt_tol, 0.01 * max(d_debit, c_credit, 1.0))
    return abs(d_debit - c_credit) <= rel


def _emit_journal_pair(
    *,
    debit_rec: dict[str, Any],
    credit_rec: dict[str, Any],
    d_amt: float,
    c_amt: float,
    display_currency_iso: str,
    credit_indent: str,
) -> list[dict[str, str]]:
    date_s = str(debit_rec.get("gl_date") or "").strip()
    dt2 = str(credit_rec.get("gl_date") or "").strip()
    narr = str(debit_rec.get("description") or "").strip()
    nar2 = str(credit_rec.get("description") or "").strip()
    acct_d = str(debit_rec.get("account") or "").strip()
    acct_c = str(credit_rec.get("account") or "").strip()
    desc_combined = narr or nar2
    iso_d = _row_currency_iso_for_format(debit_rec, display_currency_iso)
    iso_c = _row_currency_iso_for_format(credit_rec, display_currency_iso)
    return [
        {
            "Date": date_s or dt2,
            "Particulars": acct_d,
            "Debit": format_amount_display(d_amt, iso_d),
            "Credit": "",
            "Description / details": desc_combined,
        },
        {
            "Date": "",
            "Particulars": credit_indent + _credit_particulars_display(acct_c),
            "Debit": "",
            "Credit": format_amount_display(c_amt, iso_c),
            "Description / details": "",
        },
    ]


def gl_flat_records_to_journal_display_rows(
    records: list[dict[str, Any]],
    *,
    amt_tol: float = 0.015,
    display_currency_iso: str = DEFAULT_CURRENCY_ISO,
    credit_indent: str = "    ",
) -> list[dict[str, str]]:
    """
    Pair adjacent journal lines into a classic layout:
    Date | Particulars | Debit | Credit | Description / details

    Tries debit-then-credit (Excel post order) and credit-then-debit (reversed templates),
    with tolerant amount matching for rounding / display noise.
    """
    out: list[dict[str, str]] = []
    i = 0
    n = len(records)
    disp_fallback = display_currency_iso
    while i < n:
        r = records[i]
        d = row_amount(r.get("debit"))
        c = row_amount(r.get("credit"))
        date_s = str(r.get("gl_date") or "").strip()
        narr = str(r.get("description") or "").strip()
        acct = str(r.get("account") or "").strip()

        if i + 1 < n:
            r2 = records[i + 1]
            d2 = row_amount(r2.get("debit"))
            c2 = row_amount(r2.get("credit"))
            dt2 = str(r2.get("gl_date") or "").strip()

            fwd = (
                _one_sided_debit(d, c, amt_tol=amt_tol)
                and _one_sided_credit(d2, c2, amt_tol=amt_tol)
                and _journal_dates_compatible(date_s, dt2)
                and _paired_amounts_match(d, c2, amt_tol=amt_tol)
            )
            if fwd:
                out.extend(
                    _emit_journal_pair(
                        debit_rec=r,
                        credit_rec=r2,
                        d_amt=d,
                        c_amt=c2,
                        display_currency_iso=disp_fallback,
                        credit_indent=credit_indent,
                    )
                )
                i += 2
                continue

            rev = (
                _one_sided_credit(d, c, amt_tol=amt_tol)
                and _one_sided_debit(d2, c2, amt_tol=amt_tol)
                and _journal_dates_compatible(date_s, dt2)
                and _paired_amounts_match(d2, c, amt_tol=amt_tol)
            )
            if rev:
                out.extend(
                    _emit_journal_pair(
                        debit_rec=r2,
                        credit_rec=r,
                        d_amt=d2,
                        c_amt=c,
                        display_currency_iso=disp_fallback,
                        credit_indent=credit_indent,
                    )
                )
                i += 2
                continue

        iso_row = _row_currency_iso_for_format(r, disp_fallback)
        out.append(
            {
                "Date": date_s,
                "Particulars": acct,
                "Debit": format_amount_display(d, iso_row) if d > amt_tol else "",
                "Credit": format_amount_display(c, iso_row) if c > amt_tol else "",
                "Description / details": narr,
            }
        )
        i += 1
    return out


def read_gl_sheet_rows_from_path(
    path: str,
    *,
    sheet_name: str = GL_SHEET_NAME_DEFAULT,
    tail: int = 500,
    layout: Optional[dict[str, Any]] = None,
    keep_excel_row: bool = False,
) -> list[dict[str, Any]]:
    """Read GL preview from .xlsx/.xlsm (openpyxl), .xls (xlrd), or .csv."""
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return read_gl_csv_rows(path, tail=tail, layout=layout, keep_excel_row=keep_excel_row)
    if ext == ".xls":
        return read_gl_xls_rows(
            path, sheet_name=sheet_name, tail=tail, layout=layout, keep_excel_row=keep_excel_row
        )
    if ext in (".xlsx", ".xlsm"):
        return read_gl_sheet_rows(
            path,
            sheet_name=sheet_name,
            tail=tail,
            layout=layout,
            keep_excel_row=keep_excel_row,
        )
    raise ValueError(f"Unsupported file type {ext!r} for GL preview")


def apply_gl_transaction_number_edits(
    workbook_path: str,
    updates: list[tuple[int, Any]],
    *,
    sheet_name: str = GL_SHEET_NAME_DEFAULT,
    layout: Optional[dict[str, Any]] = None,
) -> None:
    """Write transaction numbers to the Tr. No. column for specific 1-based Excel rows."""
    if not updates:
        return
    ext = Path(workbook_path).suffix.lower()
    if ext not in APPENDABLE_EXTENSIONS:
        raise ValueError(
            f"Updating transaction numbers requires .xlsx or .xlsm; this file is {ext!r}. "
            "Convert or upload a modern Excel workbook in Settings."
        )
    wb = load_workbook(workbook_path, keep_vba=(ext == ".xlsm"))
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet {sheet_name!r} not found; available: {wb.sheetnames}")
        ws = wb[sheet_name]
        scan_r = min(ws.max_row or 1, max(60, (ws.max_row or 1) + 2))
        max_c = min(ws.max_column or 1, 45)
        raw_scan = _sheet_raw_cell_matrix(ws, scan_r, max_c)
        colmap = resolve_gl_column_map([tuple(r) for r in raw_scan], layout)
        grid = [list(r) for r in raw_scan]
        _apply_merged_cell_fill_to_grid(
            grid, ws, debit_col_0based=colmap.debit, credit_col_0based=colmap.credit
        )
        matrix: list[tuple[Any, ...]] = [tuple(row) for row in grid]
        colmap = resolve_gl_column_map(matrix, layout)
        if colmap.tr_number is None:
            raise ValueError("This workbook layout has no transaction-number column.")
        tr_c_1b = _openpyxl_col(colmap.tr_number)
        for row_1b, val in updates:
            try:
                ri = int(row_1b)
            except (TypeError, ValueError):
                continue
            if ri < 1:
                continue
            coerced = coerce_gl_transaction_number(val)
            tr_ar, tr_ac = _merged_top_left(ws, int(ri), tr_c_1b)
            _set_sheet_cell_value_merge_aware(ws, tr_ar, tr_ac, coerced)
        wb.save(workbook_path)
    finally:
        try:
            wb.close()
        except Exception:
            pass


@dataclass
class GlEditPlan:
    """Batch of GL sheet mutations applied in one workbook save."""

    updates: list[dict[str, Any]]
    delete_rows: list[int]
    insert_rows: list[dict[str, Any]]
    swap_rows: list[tuple[int, int]]
    # Optional identity checks: each dict has excel_row, account, debit, credit from the UI.
    delete_row_checks: list[dict[str, Any]] | None = None


def _open_gl_values_workbook(workbook_path: str, *, sheet_name: str) -> tuple[Any, Any]:
    """Companion ``data_only`` workbook for cached formula results on amount cells."""
    wb_v = load_workbook(workbook_path, data_only=True, keep_links=False)
    if sheet_name not in wb_v.sheetnames:
        wb_v.close()
        raise KeyError(f"Sheet {sheet_name!r} not found; available: {wb_v.sheetnames}")
    return wb_v, wb_v[sheet_name]


def _close_gl_workbooks(wb: Any, wb_values: Any | None = None) -> None:
    for book in (wb_values, wb):
        if book is None:
            continue
        try:
            book.close()
        except Exception:
            pass


def _open_gl_sheet_for_write(
    workbook_path: str,
    *,
    sheet_name: str = GL_SHEET_NAME_DEFAULT,
    layout: Optional[dict[str, Any]] = None,
) -> tuple[Any, Any, GlColumnMap, int, str, Any, Any]:
    """
    Return ``(workbook, worksheet, colmap, max_col, credit_style, wb_values, ws_values)``.

    ``ws`` keeps formulas for writes; ``ws_values`` supplies cached numeric results for
    debit/credit reads (``data_only=True``).
    """
    ext = Path(workbook_path).suffix.lower()
    if ext not in APPENDABLE_EXTENSIONS:
        raise ValueError(
            f"GL edits require .xlsx or .xlsm; this file is {ext!r}. "
            "Upload a modern Excel workbook in Settings."
        )
    wb = load_workbook(workbook_path, keep_vba=(ext == ".xlsm"))
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Sheet {sheet_name!r} not found; available: {wb.sheetnames}")
    ws = wb[sheet_name]
    wb_v, ws_v = _open_gl_values_workbook(workbook_path, sheet_name=sheet_name)
    mr = ws.max_row or 1
    max_c = min(ws.max_column or 1, 45)
    scan_r = min(mr, max(60, mr + 2))
    raw_scan = _sheet_raw_cell_matrix(ws, scan_r, max_c)
    colmap = resolve_gl_column_map([tuple(r) for r in raw_scan], layout)
    grid = [list(r) for r in raw_scan]
    _apply_merged_cell_fill_to_grid(
        grid, ws, debit_col_0based=colmap.debit, credit_col_0based=colmap.credit
    )
    matrix: list[tuple[Any, ...]] = [tuple(row) for row in grid]
    colmap = resolve_gl_column_map(matrix, layout)
    date_c = _openpyxl_col(colmap.date)
    desc_c = _openpyxl_col(colmap.details)
    tpl_top, tpl_bot = infer_last_entry_template_rows(ws, mr, date_c, desc_c)
    credit_style = _detect_credit_leg_amount_style(ws, colmap, tpl_bot)
    return wb, ws, colmap, max_c, credit_style, wb_v, ws_v


def _coerce_gl_date_write_value(
    ws,
    *,
    tpl_top: int,
    date_col_1b: int,
    raw: Any,
) -> Any:
    if raw is None or (isinstance(raw, str) and not str(raw).strip()):
        return None
    if isinstance(raw, date):
        return _value_for_new_gl_date(ws, tpl_top, date_col_1b, raw)
    if isinstance(raw, datetime):
        return _value_for_new_gl_date(ws, tpl_top, date_col_1b, raw.date())
    parsed = parse_gl_cell_to_date(raw)
    if parsed is not None:
        return _value_for_new_gl_date(ws, tpl_top, date_col_1b, parsed)
    return str(raw).strip()


def _write_gl_amounts_on_row(
    ws,
    colmap: GlColumnMap,
    row_1b: int,
    *,
    debit: float,
    credit: float,
    credit_style: str,
    preserve_formula_cols: frozenset[str] | None = None,
) -> None:
    deb_c = _openpyxl_col(colmap.debit)
    cred_c = _openpyxl_col(colmap.credit)
    d = max(0.0, float(debit))
    c = max(0.0, float(credit))
    if d > 1e-9 and c > 1e-9:
        raise ValueError(f"Row {row_1b}: enter debit or credit, not both.")
    keep = preserve_formula_cols or frozenset()
    if d > 1e-9:
        if "debit" not in keep:
            _set_sheet_cell_value_merge_aware(ws, row_1b, deb_c, d)
        if "credit" not in keep:
            _set_sheet_cell_value_merge_aware(ws, row_1b, cred_c, 0.0)
    elif c > 1e-9:
        if credit_style == "both_in_debit":
            if "debit" not in keep:
                _set_sheet_cell_value_merge_aware(ws, row_1b, deb_c, c)
            if "credit" not in keep:
                _set_sheet_cell_value_merge_aware(ws, row_1b, cred_c, 0.0)
        else:
            if "debit" not in keep:
                _set_sheet_cell_value_merge_aware(ws, row_1b, deb_c, 0.0)
            if "credit" not in keep:
                _set_sheet_cell_value_merge_aware(ws, row_1b, cred_c, c)
    else:
        if "debit" not in keep:
            _set_sheet_cell_value_merge_aware(ws, row_1b, deb_c, 0.0)
        if "credit" not in keep:
            _set_sheet_cell_value_merge_aware(ws, row_1b, cred_c, 0.0)


def _physical_cell_value(ws, row_1b: int, col_1b: int) -> Any:
    """Cell value on this coordinate only (``None`` for non-anchor merged slaves)."""
    row_1b = int(row_1b)
    col_1b = int(col_1b)
    m = _merge_range_covering_cell(ws, row_1b, col_1b)
    if m is not None and (row_1b != int(m.min_row) or col_1b != int(m.min_col)):
        return None
    return ws.cell(row=row_1b, column=col_1b).value


def _amount_cell_has_formula(ws_formula: Any, row_1b: int, col_1b: int) -> bool:
    return _is_excel_formula_value(_physical_cell_value(ws_formula, row_1b, col_1b))


def _eval_simple_amount_formula(
    ws_formula: Any,
    ws_values: Any | None,
    formula: str,
) -> Optional[float]:
    """
    Best-effort numeric result for simple reference formulas (e.g. ``=E2/2``, ``=$F$10``).

    Used when ``data_only`` loads have no cached ``<v>`` yet (common after openpyxl saves).
    """
    s = str(formula or "").strip()
    if not s.startswith("="):
        return None
    from openpyxl.utils import column_index_from_string

    def _ref_amount(col_letters: str, row_n: int) -> Optional[float]:
        try:
            c = int(column_index_from_string(col_letters.upper()))
            r = int(row_n)
        except (TypeError, ValueError):
            return None
        raw = _resolve_amount_cell_raw(
            ws_formula, ws_values, r, c, physical=False
        )
        amt, _ = parse_money_cell(raw)
        return float(amt) if amt is not None else None

    m = re.fullmatch(
        r"=\s*(?:'?[^']*'?!)?\$?([A-Za-z]{1,3})\$?(\d+)\s*/\s*(\d+(?:\.\d+)?)\s*",
        s,
        flags=re.IGNORECASE,
    )
    if m:
        base = _ref_amount(m.group(1), int(m.group(2)))
        if base is not None:
            try:
                return base / float(m.group(3))
            except (TypeError, ValueError, ZeroDivisionError):
                return None
    m = re.fullmatch(
        r"=\s*(?:'?[^']*'?!)?\$?([A-Za-z]{1,3})\$?(\d+)\s*",
        s,
        flags=re.IGNORECASE,
    )
    if m:
        return _ref_amount(m.group(1), int(m.group(2)))
    m = re.fullmatch(
        r"=\s*(?:'?[^']*'?!)?\$?([A-Za-z]{1,3})\$?(\d+)\s*([+\-])\s*(?:'?[^']*'?!)?\$?([A-Za-z]{1,3})\$?(\d+)\s*",
        s,
        flags=re.IGNORECASE,
    )
    if m:
        left = _ref_amount(m.group(1), int(m.group(2)))
        right = _ref_amount(m.group(4), int(m.group(5)))
        if left is None or right is None:
            return None
        if m.group(3) == "+":
            return left + right
        return left - right
    return None


def _resolve_amount_cell_raw(
    ws_formula: Any,
    ws_values: Any | None,
    row_1b: int,
    col_1b: int,
    *,
    physical: bool = False,
) -> Any:
    """
    Numeric amount source for one debit/credit coordinate.

    Uses cached formula results from ``ws_values`` when the formula sheet stores ``=...``.
    """
    row_1b = int(row_1b)
    col_1b = int(col_1b)
    formula_raw: Any = None
    if physical:
        formula_raw = _physical_cell_value(ws_formula, row_1b, col_1b)
    else:
        formula_raw = _cell_value_resolving_merge(ws_formula, row_1b, col_1b)
    if ws_values is not None:
        if physical:
            raw = _physical_cell_value(ws_values, row_1b, col_1b)
        else:
            raw = _cell_value_resolving_merge(ws_values, row_1b, col_1b)
        if raw is not None and not _is_excel_formula_value(raw):
            amt, _ = parse_money_cell(raw)
            if amt is not None and abs(float(amt)) > 1e-9:
                return raw
    if _is_excel_formula_value(formula_raw):
        evaluated = _eval_simple_amount_formula(ws_formula, ws_values, str(formula_raw))
        if evaluated is not None:
            return evaluated
        return None
    if physical:
        return _physical_cell_value(ws_formula, row_1b, col_1b)
    return formula_raw


def _logical_debit_credit_from_row(
    ws_formula: Any,
    ws_values: Any | None,
    colmap: GlColumnMap,
    row_1b: int,
    *,
    credit_style: str,
    physical: bool = False,
) -> tuple[float, float, bool, bool]:
    """Return ``(debit, credit, debit_is_formula, credit_is_formula)``."""
    deb_c = _openpyxl_col(colmap.debit)
    cred_c = _openpyxl_col(colmap.credit)
    deb_formula = _amount_cell_has_formula(ws_formula, row_1b, deb_c)
    cred_formula = _amount_cell_has_formula(ws_formula, row_1b, cred_c)
    deb_amt, _ = parse_money_cell(
        _resolve_amount_cell_raw(
            ws_formula, ws_values, row_1b, deb_c, physical=physical
        )
    )
    cre_amt, _ = parse_money_cell(
        _resolve_amount_cell_raw(
            ws_formula, ws_values, row_1b, cred_c, physical=physical
        )
    )
    deb_f = float(deb_amt) if deb_amt is not None else 0.0
    cre_f = float(cre_amt) if cre_amt is not None else 0.0
    if credit_style == "both_in_debit" and deb_f > 1e-9 and cre_f <= 1e-9:
        return 0.0, deb_f, deb_formula, cred_formula
    return deb_f, cre_f, deb_formula, cred_formula


def _read_gl_row_posting_snapshot(
    ws,
    colmap: GlColumnMap,
    row_1b: int,
    *,
    credit_style: str,
    ws_values: Any | None = None,
) -> dict[str, Any]:
    """
    Row snapshot for faux-delete restore: account/date use merge resolution; amounts use
    **physical** debit/credit cells so vertically merged debits are not copied onto credit legs.
    """
    snap = _read_gl_row_snapshot(
        ws, colmap, row_1b, credit_style=credit_style, ws_values=ws_values
    )
    deb_f, cre_f, deb_fx, cre_fx = _logical_debit_credit_from_row(
        ws,
        ws_values,
        colmap,
        row_1b,
        credit_style=credit_style,
        physical=True,
    )
    snap["debit"] = deb_f
    snap["credit"] = cre_f
    snap["_debit_formula"] = deb_fx
    snap["_credit_formula"] = cre_fx
    return snap


def _read_gl_row_snapshot(
    ws,
    colmap: GlColumnMap,
    row_1b: int,
    *,
    credit_style: str,
    ws_values: Any | None = None,
) -> dict[str, Any]:
    date_c = _openpyxl_col(colmap.date)
    desc_c = _openpyxl_col(colmap.details)
    acct_c = _openpyxl_col(colmap.particulars)
    tr_c = _openpyxl_col(colmap.tr_number) if colmap.tr_number is not None else None

    debit_out, credit_out, _, _ = _logical_debit_credit_from_row(
        ws,
        ws_values,
        colmap,
        row_1b,
        credit_style=credit_style,
        physical=False,
    )

    row_t = _worksheet_row_tuple_for_colmap(ws, row_1b, colmap)
    account_eff = particulars_label_on_worksheet_row(
        ws, colmap, row_1b, credit_style=credit_style
    ) or _effective_particulars_account_label(row_t, colmap)

    snap: dict[str, Any] = {
        "gl_date": _cell_value_resolving_merge(ws, row_1b, date_c),
        "description": _cell_value_resolving_merge(ws, row_1b, desc_c),
        "account": account_eff or _cell_value_resolving_merge(ws, row_1b, acct_c),
        "debit": debit_out,
        "credit": credit_out,
    }
    if tr_c is not None:
        snap["transaction_number"] = _cell_value_resolving_merge(ws, row_1b, tr_c)
    return snap


def _apply_gl_row_snapshot(
    ws,
    colmap: GlColumnMap,
    row_1b: int,
    snap: dict[str, Any],
    *,
    credit_style: str,
    tpl_top: int,
) -> None:
    date_c = _openpyxl_col(colmap.date)
    desc_c = _openpyxl_col(colmap.details)
    tr_c = _openpyxl_col(colmap.tr_number) if colmap.tr_number is not None else None

    if snap.get("account") is not None:
        acct_s = str(snap.get("account") or "").strip()
        if acct_s:
            _tt, _tb = infer_last_entry_template_rows(ws, row_1b, date_c, desc_c)
            _set_gl_account_on_row(
                ws,
                row_1b,
                colmap,
                acct_s,
                tpl_top=_tt,
                tpl_bot=_tb,
                credit_style=credit_style,
            )
    if snap.get("description") is not None:
        desc_s = str(snap.get("description") or "").strip()
        if desc_s:
            desc_ar, desc_ac = _merged_top_left(ws, row_1b, desc_c)
            _set_sheet_cell_value_merge_aware(ws, desc_ar, desc_ac, desc_s)
    if "gl_date" in snap:
        dv = _coerce_gl_date_write_value(
            ws, tpl_top=tpl_top, date_col_1b=date_c, raw=snap.get("gl_date")
        )
        if dv is not None:
            date_ar, date_ac = _merged_top_left(ws, row_1b, date_c)
            _set_sheet_cell_value_merge_aware(ws, date_ar, date_ac, dv)
            nf = _resolve_gl_date_number_format(ws, tpl_top, tpl_top, date_c, row_1b)
            _set_cell_number_format_merge_aware(
                ws, date_ar, date_ac, (nf or "").strip() or "yyyy-mm-dd"
            )
    if "debit" in snap or "credit" in snap:
        ws_values = snap.pop("_ws_values", None)
        preserve: set[str] = set()
        if snap.get("_debit_formula"):
            preserve.add("debit")
        if snap.get("_credit_formula"):
            preserve.add("credit")
        snap.pop("_debit_formula", None)
        snap.pop("_credit_formula", None)
        cur = _read_gl_row_snapshot(
            ws, colmap, row_1b, credit_style=credit_style, ws_values=ws_values
        )
        deb_out = float(snap["debit"]) if "debit" in snap else float(cur.get("debit") or 0)
        cre_out = float(snap["credit"]) if "credit" in snap else float(cur.get("credit") or 0)
        _write_gl_amounts_on_row(
            ws,
            colmap,
            row_1b,
            debit=deb_out,
            credit=cre_out,
            credit_style=credit_style,
            preserve_formula_cols=frozenset(preserve),
        )
    if tr_c is not None and "transaction_number" in snap:
        tr_coerced = coerce_gl_transaction_number(snap.get("transaction_number"))
        if tr_coerced is not None and str(tr_coerced).strip() != "":
            _set_sheet_cell_value_merge_aware(ws, row_1b, tr_c, tr_coerced)


def _set_gl_row_fields(
    ws,
    colmap: GlColumnMap,
    row_1b: int,
    fields: dict[str, Any],
    *,
    credit_style: str,
    tpl_top: int,
) -> None:
    snap: dict[str, Any] = {}
    if "gl_date" in fields:
        snap["gl_date"] = fields["gl_date"]
    if "account" in fields:
        snap["account"] = fields["account"]
    if "description" in fields and str(fields.get("description") or "").strip():
        snap["description"] = fields["description"]
    if "debit" in fields:
        snap["debit"] = float(fields["debit"])
    if "credit" in fields:
        snap["credit"] = float(fields["credit"])
    if "transaction_number" in fields and str(fields.get("transaction_number") or "").strip():
        snap["transaction_number"] = fields["transaction_number"]
    _apply_gl_row_snapshot(ws, colmap, row_1b, snap, credit_style=credit_style, tpl_top=tpl_top)


def _sort_insert_specs_debit_first(specs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Keep debit legs above credit legs for multi-line journal inserts."""
    debits: list[dict[str, Any]] = []
    credits: list[dict[str, Any]] = []
    other: list[dict[str, Any]] = []
    for spec in specs:
        row = dict(spec)
        if float(row.get("debit") or 0) > 1e-9:
            debits.append(row)
        elif float(row.get("credit") or 0) > 1e-9:
            credits.append(row)
        else:
            other.append(row)
    return debits + credits + other


def _write_journal_block_lines(
    ws,
    colmap: GlColumnMap,
    *,
    row_start: int,
    line_specs: list[dict[str, Any]],
    tpl_top: int,
    tpl_bot: int,
    block_merges: list[tuple[int, int, int, int]],
    horizontal_by_tpl_row: dict[int, list[tuple[int, int, int, int]]],
    credit_style: str,
    max_col: int,
) -> None:
    """Write business values onto a journal block that already has template structure."""
    if not line_specs:
        return
    block_h = max(1, int(tpl_bot) - int(tpl_top) + 1)
    date_c = _openpyxl_col(colmap.date)
    desc_c = _openpyxl_col(colmap.details)
    acct_c = _openpyxl_col(colmap.particulars)
    deb_c = _openpyxl_col(colmap.debit)
    cred_c = _openpyxl_col(colmap.credit)
    tr_c = _openpyxl_col(colmap.tr_number) if colmap.tr_number is not None else None

    date_vert, _, _ = _column_in_template_vertical_block(block_merges, date_c, tpl_top, tpl_bot)
    desc_vert, _, _ = _column_in_template_vertical_block(block_merges, desc_c, tpl_top, tpl_bot)
    tr_vert = False
    if tr_c is not None:
        tr_vert, _, _ = _column_in_template_vertical_block(block_merges, tr_c, tpl_top, tpl_bot)

    _, date_tpl_anchor_c = _merged_top_left(ws, tpl_top, date_c)
    _, desc_tpl_anchor_c = _merged_top_left(ws, tpl_top, desc_c)
    tpl_tr_anchor_c = tpl_top
    if tr_c is not None:
        _, tpl_tr_anchor_c = _merged_top_left(ws, tpl_top, tr_c)

    for i, spec in enumerate(line_specs):
        r = int(row_start) + i
        template_r = int(tpl_top) + (i % block_h)

        acct_s = str(spec.get("account") or "").strip()
        deb_f = float(spec.get("debit") or 0)
        cred_f = float(spec.get("credit") or 0)
        ac_mirror_col = acct_c
        for _mr1, mc1a, _mr2, mc2a in horizontal_by_tpl_row.get(template_r, ()):
            if mc1a <= acct_c <= mc2a:
                ac_mirror_col = min(ac_mirror_col, mc1a)
        if acct_s:
            _set_gl_account_on_row(
                ws,
                r,
                colmap,
                acct_s,
                tpl_top=tpl_top,
                tpl_bot=tpl_bot,
                credit_style=credit_style,
            )
            ac_mirror_col = particulars_columns_1b(ws, r, acct_c)[0]

        _write_gl_amounts_on_row(
            ws,
            colmap,
            r,
            debit=deb_f,
            credit=cred_f,
            credit_style=credit_style,
        )

        raw_date = spec.get("gl_date")
        if raw_date not in (None, ""):
            dv = _coerce_gl_date_write_value(
                ws, tpl_top=tpl_top, date_col_1b=date_c, raw=raw_date
            )
            if dv is not None:
                if date_vert:
                    if i == 0:
                        _set_sheet_cell_value_merge_aware(ws, int(row_start), date_tpl_anchor_c, dv)
                        nf = _resolve_gl_date_number_format(ws, tpl_top, tpl_bot, date_c, r)
                        _set_cell_number_format_merge_aware(
                            ws, int(row_start), date_tpl_anchor_c, (nf or "").strip() or "yyyy-mm-dd"
                        )
                else:
                    date_ar, date_ac = _merged_top_left(ws, r, date_c)
                    _set_sheet_cell_value_merge_aware(ws, date_ar, date_ac, dv)
                    nf = _resolve_gl_date_number_format(ws, tpl_top, tpl_bot, date_c, r)
                    _set_cell_number_format_merge_aware(
                        ws, date_ar, date_ac, (nf or "").strip() or "yyyy-mm-dd"
                    )

        desc_s = str(spec.get("description") or "").strip()
        if desc_s:
            if desc_vert:
                if i == 0:
                    _set_sheet_cell_value_merge_aware(ws, int(row_start), desc_tpl_anchor_c, desc_s)
            else:
                desc_ar, desc_ac = _merged_top_left(ws, r, desc_c)
                _set_sheet_cell_value_merge_aware(ws, desc_ar, desc_ac, desc_s)

        if tr_c is not None and "transaction_number" in spec:
            tr_coerced = coerce_gl_transaction_number(spec.get("transaction_number"))
            if tr_coerced is not None and str(tr_coerced).strip() != "":
                if tr_vert:
                    if i == 0:
                        _set_sheet_cell_value_merge_aware(ws, int(row_start), tpl_tr_anchor_c, tr_coerced)
                else:
                    tr_ar, tr_ac = _merged_top_left(ws, r, tr_c)
                    _set_sheet_cell_value_merge_aware(ws, tr_ar, tr_ac, tr_coerced)

        tpl_deb_nf = ws.cell(row=template_r, column=deb_c).number_format or ""
        tpl_cred_nf = ws.cell(row=template_r, column=cred_c).number_format or ""
        _set_cell_number_format_merge_aware(ws, r, deb_c, tpl_deb_nf or "General")
        _set_cell_number_format_merge_aware(ws, r, cred_c, tpl_cred_nf or "General")

        ac_tpl_r_a, ac_tpl_c_a = _merged_top_left(ws, template_r, acct_c)
        _mirror_anchor_number_format(ws, ac_tpl_r_a, ac_tpl_c_a, r, ac_mirror_col)


def _particulars_horizontal_merge_span(
    ws, row_1b: int, particulars_col_1b: int
) -> tuple[int, int] | None:
    """Return ``(min_col, max_col)`` for a single-row particulars horizontal merge, if any."""
    row_1b = int(row_1b)
    p = int(particulars_col_1b)
    band = particulars_columns_1b(ws, row_1b, p)
    if len(band) > 1:
        return band[0], band[-1]
    m = _merge_range_covering_cell(ws, row_1b, p)
    if (
        m is not None
        and int(m.min_row) == int(m.max_row) == row_1b
        and int(m.max_col) > int(m.min_col)
        and int(m.min_col) <= p <= int(m.max_col)
    ):
        return int(m.min_col), int(m.max_col)
    return None


def _row_leg_kind(ws, colmap: GlColumnMap, row_1b: int) -> str:
    deb, _ = parse_money_cell(_cell_value_resolving_merge(ws, int(row_1b), _openpyxl_col(colmap.debit)))
    cre, _ = parse_money_cell(_cell_value_resolving_merge(ws, int(row_1b), _openpyxl_col(colmap.credit)))
    if (deb or 0) > 0.015:
        return "debit"
    if (cre or 0) > 0.015:
        return "credit"
    return "other"


def detect_transaction_merge_format(
    ws,
    colmap: GlColumnMap,
    block_lo: int,
    block_hi: int,
) -> dict[str, Any]:
    """
    Capture particulars horizontal merge spans for debit/credit legs in one journal block.

    Returns a target format dict for :func:`standardize_transaction_merge_format`.
    """
    acct_c = _openpyxl_col(colmap.particulars)
    date_c = _openpyxl_col(colmap.date)
    desc_c = _openpyxl_col(colmap.details)
    block_lo = int(block_lo)
    block_hi = int(block_hi)

    debit_span: tuple[int, int] | None = None
    credit_span: tuple[int, int] | None = None
    for r in range(block_lo, block_hi + 1):
        kind = _row_leg_kind(ws, colmap, r)
        span = _particulars_horizontal_merge_span(ws, r, acct_c)
        if span is None:
            continue
        if kind == "debit" and debit_span is None:
            debit_span = span
        elif kind == "credit" and credit_span is None:
            credit_span = span

    target = debit_span or credit_span
    if debit_span and credit_span and debit_span != credit_span:
        target = debit_span

    date_vert: tuple[int, int, int, int] | None = None
    desc_vert: tuple[int, int, int, int] | None = None
    for col, slot in ((date_c, "date"), (desc_c, "desc")):
        m = _merge_range_covering_cell(ws, block_lo, col)
        if m is not None and int(m.max_row) > int(m.min_row):
            quad = (int(m.min_row), int(m.min_col), int(m.max_row), int(m.max_col))
            if slot == "date":
                date_vert = quad
            else:
                desc_vert = quad

    return {
        "debit_particulars_merge": debit_span,
        "credit_particulars_merge": credit_span,
        "target_particulars_merge": target,
        "date_vertical_merge": date_vert,
        "desc_vertical_merge": desc_vert,
    }


def standardize_transaction_merge_format(
    ws,
    colmap: GlColumnMap,
    block_lo: int,
    block_hi: int,
    target_format: dict[str, Any],
) -> None:
    """Align particulars horizontal merges on each leg row without clearing cell values."""
    target = target_format.get("target_particulars_merge")
    if not target:
        return
    mc1, mc2 = int(target[0]), int(target[1])
    acct_c = _openpyxl_col(colmap.particulars)
    deb_c = _openpyxl_col(colmap.debit)
    cred_c = _openpyxl_col(colmap.credit)

    for r in range(int(block_lo), int(block_hi) + 1):
        cur = _particulars_horizontal_merge_span(ws, r, acct_c)
        if cur == (mc1, mc2):
            continue

        acct_val = _cell_value_resolving_merge(ws, r, acct_c)
        deb_val = _cell_value_resolving_merge(ws, r, deb_c)
        cred_val = _cell_value_resolving_merge(ws, r, cred_c)

        for m in list(ws.merged_cells.ranges):
            if (
                int(m.min_row) == r == int(m.max_row)
                and int(m.max_col) > int(m.min_col)
                and int(m.min_col) <= acct_c <= int(m.max_col)
            ):
                _safe_unmerge_cells(ws, m)

        ws.merge_cells(start_row=r, start_column=mc1, end_row=r, end_column=mc2)

        if acct_val not in (None, ""):
            _set_sheet_cell_value_merge_aware(ws, r, mc1, acct_val)
        if deb_val not in (None, ""):
            _set_sheet_cell_value_merge_aware(ws, r, deb_c, deb_val)
        if cred_val not in (None, ""):
            _set_sheet_cell_value_merge_aware(ws, r, cred_c, cred_val)


def _coerce_insert_anchor_to_block_bottom(
    ws,
    colmap: GlColumnMap,
    insert_after_1b: int,
    *,
    line_count: int = 1,
) -> int:
    """
    Never insert between debit/credit legs when adding compound entries.

    For multi-line inserts (``line_count >= 2``), always coerce ``insert_after`` to the
    journal block bottom. Single-line inserts keep ``insert_below_exact`` placement unless
    the anchor row is a blank spacer inside the block.
    """
    row_1b = int(insert_after_1b)
    floor = int(colmap.data_start_row)
    if row_1b < floor:
        return max(floor, row_1b)

    date_c = _openpyxl_col(colmap.date)
    desc_c = _openpyxl_col(colmap.details)
    block_lo, block_hi = _journal_block_bounds_workbook(ws, colmap, row_1b, date_c, desc_c)

    if int(line_count) >= 2 and block_hi > row_1b:
        return block_hi
    if block_lo <= row_1b <= block_hi and not _worksheet_row_has_posting_content(ws, colmap, row_1b):
        return block_hi
    return row_1b


def _resolve_gl_insert_anchor(
    ws,
    colmap: GlColumnMap,
    insert_after_1b: int,
    *,
    insert_below_exact: bool,
    line_count: int = 1,
) -> int:
    """Resolve the Excel row new GL lines should be inserted below."""
    row_1b = int(insert_after_1b)
    floor = int(colmap.data_start_row)
    if row_1b < floor:
        return max(floor, row_1b)

    if not insert_below_exact:
        row_1b = resolve_gl_insert_after_row(ws, colmap, row_1b)
    return _coerce_insert_anchor_to_block_bottom(ws, colmap, row_1b, line_count=int(line_count))


def _insert_gl_rows_after(
    ws,
    colmap: GlColumnMap,
    insert_after_1b: int,
    rows: list[dict[str, Any]],
    *,
    max_col: int,
    insert_below_exact: bool = False,
) -> list[int]:
    """Insert GL lines below ``insert_after_1b``; return 1-based Excel row numbers written."""
    if not rows:
        return []
    date_c = _openpyxl_col(colmap.date)
    desc_c = _openpyxl_col(colmap.details)
    anchor = _resolve_gl_insert_anchor(
        ws,
        colmap,
        int(insert_after_1b),
        insert_below_exact=bool(insert_below_exact),
        line_count=len(rows),
    )
    deb_c = _openpyxl_col(colmap.debit)
    cred_c = _openpyxl_col(colmap.credit)
    tpl_top, tpl_bot = infer_last_entry_template_rows(ws, anchor, date_c, desc_c)
    block_merges, horizontal_by_tpl_row = _collect_journal_template_merges(ws, tpl_top, tpl_bot)
    block_h = max(1, tpl_bot - tpl_top + 1)
    credit_style = _detect_credit_leg_amount_style(ws, colmap, tpl_bot)

    ordered = _sort_insert_specs_debit_first(
        [
            {k: v for k, v in spec.items() if k not in ("insert_after", "_insert_below_exact")}
            for spec in rows
        ]
    )
    at_row = anchor + 1
    k = len(ordered)
    anchor_lo, anchor_hi = _journal_block_bounds_workbook(ws, colmap, anchor, date_c, desc_c)
    merge_fmt = detect_transaction_merge_format(ws, colmap, anchor_lo, anchor_hi)
    if k >= 2:
        insert_sheet_rows(ws, at_row, amount=k)
        _prepare_gl_structure_from_last_template(
            ws,
            row_start=at_row,
            row_end=at_row + k - 1,
            tpl_top=tpl_top,
            block_h=block_h,
            block_merges=block_merges,
            horizontal_by_tpl_row=horizontal_by_tpl_row,
            max_col=max_col,
        )
        _write_journal_block_lines(
            ws,
            colmap,
            row_start=at_row,
            line_specs=ordered,
            tpl_top=tpl_top,
            tpl_bot=tpl_bot,
            block_merges=block_merges,
            horizontal_by_tpl_row=horizontal_by_tpl_row,
            credit_style=credit_style,
            max_col=max_col,
        )
        standardize_transaction_merge_format(
            ws, colmap, at_row, at_row + k - 1, merge_fmt
        )
        return list(range(at_row, at_row + k))

    inserted: list[int] = []
    for spec in ordered:
        fields = dict(spec)
        insert_sheet_rows(ws, at_row, amount=1)
        style_src = max(int(colmap.data_start_row), at_row - 1)
        _copy_row_style(ws, style_src, at_row, max_col=max_col)
        _set_gl_row_fields(ws, colmap, at_row, fields, credit_style=credit_style, tpl_top=tpl_top)
        tpl_deb_nf = ws.cell(row=style_src, column=deb_c).number_format or ""
        tpl_cred_nf = ws.cell(row=style_src, column=cred_c).number_format or ""
        _set_cell_number_format_merge_aware(ws, at_row, deb_c, tpl_deb_nf or "General")
        _set_cell_number_format_merge_aware(ws, at_row, cred_c, tpl_cred_nf or "General")
        standardize_transaction_merge_format(ws, colmap, at_row, at_row, merge_fmt)
        inserted.append(at_row)
        at_row += 1
    return inserted


def _append_gl_rows_at_sheet_tail(
    ws,
    colmap: GlColumnMap,
    rows: list[dict[str, Any]],
    *,
    max_col: int,
) -> list[int]:
    if not rows:
        return []
    date_c = _openpyxl_col(colmap.date)
    desc_c = _openpyxl_col(colmap.details)
    deb_c = _openpyxl_col(colmap.debit)
    cred_c = _openpyxl_col(colmap.credit)

    inserted: list[int] = []
    for spec in rows:
        mr = ws.max_row or 1
        tpl_top, tpl_bot = infer_last_entry_template_rows(ws, mr, date_c, desc_c)
        block_merges, horizontal_by_tpl_row = _collect_journal_template_merges(ws, tpl_top, tpl_bot)
        block_h = max(1, tpl_bot - tpl_top + 1)
        credit_style = _detect_credit_leg_amount_style(ws, colmap, tpl_bot)
        row_start = mr + 1
        row_end = row_start
        _prepare_gl_structure_from_last_template(
            ws,
            row_start=row_start,
            row_end=row_end,
            tpl_top=tpl_top,
            block_h=block_h,
            block_merges=block_merges,
            horizontal_by_tpl_row=horizontal_by_tpl_row,
            max_col=max_col,
        )
        _set_gl_row_fields(ws, colmap, row_start, spec, credit_style=credit_style, tpl_top=tpl_top)
        tpl_deb_nf = ws.cell(row=tpl_top, column=deb_c).number_format or ""
        tpl_cred_nf = ws.cell(row=tpl_top, column=cred_c).number_format or ""
        _set_cell_number_format_merge_aware(ws, row_start, deb_c, tpl_deb_nf or "General")
        _set_cell_number_format_merge_aware(ws, row_start, cred_c, tpl_cred_nf or "General")
        inserted.append(row_start)
    return inserted


def _snapshot_sheet_row_values(ws, row_1b: int, *, max_col: int) -> list[Any]:
    return [_cell_value_resolving_merge(ws, row_1b, c) for c in range(1, max_col + 1)]


def _snapshot_row_raw(ws, row_1b: int, *, max_col: int) -> list[Any]:
    """Physical cell values on one sheet row (no merged top-left spillover)."""
    return [ws.cell(row=int(row_1b), column=c).value for c in range(1, int(max_col) + 1)]


def _write_row_raw(ws, row_1b: int, values: list[Any], *, max_col: int) -> None:
    """Write physical values onto one sheet row after merges were removed."""
    row_1b = int(row_1b)
    for c in range(1, int(max_col) + 1):
        v = values[c - 1] if c - 1 < len(values) else None
        _set_sheet_cell_value_merge_aware(ws, row_1b, c, "" if v is None else v)


def _restore_sheet_row_values(ws, row_1b: int, values: list[Any]) -> None:
    for c, v in enumerate(values, start=1):
        if v is None:
            continue
        anchor_r, anchor_c = _merged_top_left(ws, row_1b, c)
        if anchor_r != row_1b:
            continue
        _set_sheet_cell_value_merge_aware(ws, anchor_r, anchor_c, v)


def _propagate_journal_stripe_labels(
    ws,
    block_lo: int,
    block_hi: int,
    *,
    skip_row: int,
    date_col_1b: int,
    desc_col_1b: int,
) -> None:
    """Copy date/memo from the journal anchor row onto surviving legs after unmerge."""
    if block_hi <= block_lo:
        return
    anchor_d = ws.cell(int(block_lo), int(date_col_1b)).value
    anchor_desc = ws.cell(int(block_lo), int(desc_col_1b)).value
    for r in range(int(block_lo), int(block_hi) + 1):
        if r == int(skip_row):
            continue
        if anchor_d not in (None, ""):
            date_ar, date_ac = _merged_top_left(ws, r, int(date_col_1b))
            _set_sheet_cell_value_merge_aware(ws, date_ar, date_ac, anchor_d)
        if anchor_desc not in (None, ""):
            desc_ar, desc_ac = _merged_top_left(ws, r, int(desc_col_1b))
            _set_sheet_cell_value_merge_aware(ws, desc_ar, desc_ac, anchor_desc)


def _unmerge_ranges_intersecting_rows(ws, row_lo: int, row_hi: int) -> None:
    to_remove: list[Any] = []
    for m in ws.merged_cells.ranges:
        if int(m.max_row) >= row_lo and int(m.min_row) <= row_hi:
            to_remove.append(m)
    for m in to_remove:
        _safe_unmerge_cells(ws, m)


def _journal_block_bounds_for_row(ws, row_1b: int, date_col_1b: int, desc_col_1b: int) -> tuple[int, int, bool]:
    """Return ``(lo, hi, multi_row_journal)`` for the journal band touching ``row_1b``."""
    row_1b = int(row_1b)
    lo = hi = row_1b
    found_vert = False
    for col in (int(date_col_1b), int(desc_col_1b)):
        m = _merge_range_covering_cell(ws, row_1b, col)
        if m is not None and int(m.max_row) > int(m.min_row):
            lo = min(lo, int(m.min_row))
            hi = max(hi, int(m.max_row))
            found_vert = True
    # When only one stripe column is vertically merged, still honor that span even if the
    # other column (date vs description) is single-row on the anchor leg.
    for col in (int(date_col_1b), int(desc_col_1b)):
        m = _merge_range_covering_cell(ws, lo, col)
        if m is not None and int(m.max_row) > int(m.min_row):
            lo = min(lo, int(m.min_row))
            hi = max(hi, int(m.max_row))
            found_vert = True
    return lo, hi, found_vert


def _worksheet_row_has_tr(ws, row_1b: int, tr_col_1b: int) -> bool:
    v = _cell_value_resolving_merge(ws, int(row_1b), int(tr_col_1b))
    return v is not None and str(v).strip() != ""


def _worksheet_row_tr_label(ws, row_1b: int, tr_col_1b: int) -> str:
    v = _cell_value_resolving_merge(ws, int(row_1b), int(tr_col_1b))
    return str(v).strip() if v is not None else ""


def _worksheet_row_has_posting_content(ws, colmap: GlColumnMap, row_1b: int) -> bool:
    """True when a sheet row still looks like a GL leg (account or amount)."""
    snap = {
        "account": _effective_particulars_account_label(
            _worksheet_row_tuple_for_colmap(ws, int(row_1b), colmap),
            colmap,
        ),
        "debit": _cell_value_resolving_merge(ws, int(row_1b), _openpyxl_col(colmap.debit)),
        "credit": _cell_value_resolving_merge(ws, int(row_1b), _openpyxl_col(colmap.credit)),
    }
    deb, _ = parse_money_cell(snap.get("debit"))
    cre, _ = parse_money_cell(snap.get("credit"))
    if (deb or 0) > 0.015 or (cre or 0) > 0.015:
        return True
    return bool(_norm_account_label(snap.get("account")))


def _journal_block_tr_breaks_entry(
    tr_label: str,
    anchor_tr: str,
) -> bool:
    """True when ``tr_label`` starts a different journal entry than ``anchor_tr``."""
    if not tr_label:
        return False
    if anchor_tr and tr_label != anchor_tr:
        return True
    return not bool(anchor_tr)


def _row_starts_new_journal_entry(
    ws,
    row_1b: int,
    block_lo: int,
    date_col_1b: int,
) -> bool:
    """True when ``row_1b`` carries its own date anchor and begins a new journal entry."""
    row_1b = int(row_1b)
    block_lo = int(block_lo)
    if row_1b <= block_lo:
        return False
    val = _cell_value_resolving_merge(ws, row_1b, int(date_col_1b))
    if val in (None, ""):
        return False
    m = _merge_range_covering_cell(ws, row_1b, int(date_col_1b))
    if m is not None and int(m.min_row) < row_1b:
        return False
    return True


def _extend_journal_block_hi_through_spacers(
    ws,
    colmap: GlColumnMap,
    hi: int,
    max_row: int,
    *,
    block_lo: int,
    date_col_1b: int,
    tr_col_1b: int | None,
    anchor_tr: str,
) -> int:
    """Include blank spacer rows between debit/credit legs of the same entry."""
    block_lo = int(block_lo)
    while hi < max_row:
        below = hi + 1
        if _row_starts_new_journal_entry(ws, below, block_lo, date_col_1b):
            break
        if tr_col_1b is not None:
            tr_below = _worksheet_row_tr_label(ws, below, tr_col_1b)
            if _journal_block_tr_breaks_entry(tr_below, anchor_tr):
                break
        if _worksheet_row_has_posting_content(ws, colmap, below):
            hi = below
            continue
        below2 = below + 1
        if below2 > max_row:
            break
        if _row_starts_new_journal_entry(ws, below2, block_lo, date_col_1b):
            break
        if tr_col_1b is not None:
            tr_below2 = _worksheet_row_tr_label(ws, below2, tr_col_1b)
            if _journal_block_tr_breaks_entry(tr_below2, anchor_tr):
                break
        if _worksheet_row_has_posting_content(ws, colmap, below2):
            hi = below
            continue
        break
    return int(hi)


def _journal_block_bounds_workbook(
    ws,
    colmap: GlColumnMap,
    row_1b: int,
    date_col_1b: int,
    desc_col_1b: int,
) -> tuple[int, int]:
    """
    Workbook rows that belong to the same visual journal entry as ``row_1b``.

    Uses merged date/memo stripes plus continuation rows (same rule as
    :func:`partition_journal_blocks`), including blank spacer rows between legs.
    """
    lo, hi, _ = _journal_block_bounds_for_row(ws, int(row_1b), int(date_col_1b), int(desc_col_1b))
    data_floor = int(colmap.data_start_row)
    max_row = int(ws.max_row or hi)
    tr_col_1b = _openpyxl_col(colmap.tr_number) if colmap.tr_number is not None else None
    anchor_tr = _worksheet_row_tr_label(ws, lo, tr_col_1b) if tr_col_1b is not None else ""

    while lo > data_floor:
        above = lo - 1
        if tr_col_1b is not None:
            tr_above = _worksheet_row_tr_label(ws, above, tr_col_1b)
            if _journal_block_tr_breaks_entry(tr_above, anchor_tr):
                break
            if not anchor_tr and tr_above:
                anchor_tr = tr_above
        if not _worksheet_row_has_posting_content(ws, colmap, above):
            break
        lo = above
    if tr_col_1b is not None:
        anchor_tr = anchor_tr or _worksheet_row_tr_label(ws, lo, tr_col_1b)

    hi = _extend_journal_block_hi_through_spacers(
        ws,
        colmap,
        hi,
        max_row,
        block_lo=lo,
        date_col_1b=int(date_col_1b),
        tr_col_1b=tr_col_1b,
        anchor_tr=anchor_tr,
    )
    return int(lo), int(hi)


def resolve_gl_insert_after_row(
    ws,
    colmap: GlColumnMap,
    insert_after_1b: int,
) -> int:
    """
    Excel row to insert below when adding journal lines.

    If ``insert_after_1b`` is a debit leg of a multi-row entry, returns the bottom row of
    that entry so new lines are not placed between the entry's debit and credit legs.
    """
    row_1b = int(insert_after_1b)
    if row_1b < int(colmap.data_start_row):
        return max(int(colmap.data_start_row), row_1b)
    date_c = _openpyxl_col(colmap.date)
    desc_c = _openpyxl_col(colmap.details)
    _lo, hi = _journal_block_bounds_workbook(ws, colmap, row_1b, date_c, desc_c)
    return int(hi)


def _adjust_block_vertical_merges(
    block_merges: list[tuple[int, int, int, int]],
    *,
    block_lo: int,
    block_hi: int,
    new_hi: int,
) -> list[tuple[int, int, int, int]]:
    """Shrink full-block vertical merges after one leg row is removed."""
    out: list[tuple[int, int, int, int]] = []
    for mr1, mc1, mr2, mc2 in block_merges:
        if mr1 == block_lo and mr2 == block_hi and new_hi > block_lo:
            out.append((block_lo, mc1, new_hi, mc2))
        elif mr2 < block_lo or mr1 > block_hi:
            out.append((mr1, mc1, mr2, mc2))
        elif mr1 >= block_lo and mr2 <= block_hi:
            continue
        else:
            out.append((mr1, mc1, mr2, mc2))
    return out


def _norm_account_label(v: Any) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip())


def verify_gl_delete_targets(
    ws,
    colmap: GlColumnMap,
    checks: list[dict[str, Any]],
    *,
    credit_style: str,
    ws_values: Any | None = None,
) -> str | None:
    """
    Ensure each workbook row still matches what the user selected in the GL editor.

    Returns an error message when a row number shifted or no longer refers to that posting.
    """
    for spec in checks:
        try:
            er = int(spec.get("excel_row") or 0)
        except (TypeError, ValueError):
            continue
        if er < 1:
            continue
        exp_acct = _norm_account_label(spec.get("account"))
        if not exp_acct:
            return f"Row **{er}** has no account label in the editor — refresh and reselect."

        snap = _read_gl_row_snapshot(
            ws, colmap, er, credit_style=credit_style, ws_values=ws_values
        )
        got_acct = _norm_account_label(snap.get("account"))
        if not got_acct:
            try:
                exp_deb = float(spec.get("debit") or 0)
                exp_cred = float(spec.get("credit") or 0)
            except (TypeError, ValueError):
                exp_deb = exp_cred = 0.0
            got_deb = float(snap.get("debit") or 0)
            got_cred = float(snap.get("credit") or 0)
            amt_tol = 0.02
            amounts_match = (
                abs(got_deb - exp_deb) <= amt_tol
                and abs(got_cred - exp_cred) <= amt_tol
                and (got_deb > amt_tol or got_cred > amt_tol)
            )
            if not amounts_match:
                return (
                    f"Row **{er}** is empty or not a posting line in the workbook. "
                    "Refresh Financials and reselect rows to delete."
                )
            continue
        if got_acct != exp_acct:
            return (
                f"Row **{er}** no longer matches **{exp_acct}** (workbook shows **{got_acct}**). "
                "Refresh Financials and reselect the correct **Row #**."
            )
    return None


def _posting_column_indices_1b(colmap: GlColumnMap, *, max_col_1b: int) -> list[int]:
    """Openpyxl 1-based columns cleared for faux delete (account band + amount cols)."""
    cols: set[int] = set()
    p0 = int(colmap.particulars)
    scan_hi = _account_column_scan_upper_bound(colmap, p0, max(1, int(max_col_1b)))
    for c0 in range(p0, min(scan_hi, int(max_col_1b))):
        cols.add(_openpyxl_col(c0))
    for attr in ("particulars", "debit", "credit", "tr_number", "currency"):
        idx = getattr(colmap, attr, None)
        if idx is not None and int(idx) >= 0:
            cols.add(_openpyxl_col(int(idx)))
    return sorted(cols)


def _is_merged_slave_cell(ws, row_1b: int, col_1b: int) -> bool:
    from openpyxl.cell.cell import MergedCell

    return isinstance(ws.cell(row=int(row_1b), column=int(col_1b)), MergedCell)


def _materialize_writable_cell(ws, row_1b: int, col_1b: int):
    """Replace orphan read-only ``MergedCell`` instances with a writable ``Cell``."""
    from openpyxl.cell.cell import Cell, MergedCell

    row_1b = int(row_1b)
    col_1b = int(col_1b)
    cell = ws._cells.get((row_1b, col_1b))
    if cell is None:
        return ws.cell(row=row_1b, column=col_1b)
    if isinstance(cell, MergedCell):
        nc = Cell(ws, row=row_1b, column=col_1b)
        ws._cells[(row_1b, col_1b)] = nc
        return nc
    return cell


def _safe_unmerge_cells(ws, m) -> None:
    """Drop a merge range even when ``insert_rows`` left slave cells out of ``ws._cells``."""
    try:
        ws.unmerge_cells(str(m))
        return
    except KeyError:
        pass
    try:
        ws.merged_cells.ranges.discard(m)
    except Exception:
        try:
            ws.merged_cells.remove(m)
        except Exception:
            pass
    from openpyxl.cell.cell import Cell, MergedCell

    ar, ac = int(m.min_row), int(m.min_col)
    for r in range(int(m.min_row), int(m.max_row) + 1):
        for c in range(int(m.min_col), int(m.max_col) + 1):
            if (r, c) == (ar, ac):
                continue
            existing = ws._cells.get((r, c))
            if existing is None or isinstance(existing, MergedCell):
                ws._cells[(r, c)] = Cell(ws, row=r, column=c)


def _resolve_writable_coordinate(ws, row_1b: int, col_1b: int) -> tuple[int, int]:
    """
    Return a writable ``(row, col)`` for ``value`` assignment.

    Handles normal cells, merge anchors, vertical/horizontal splits (via unmerge), and orphan
    ``MergedCell`` slaves missing merge metadata.
    """
    row_1b = int(row_1b)
    col_1b = int(col_1b)
    m = _merge_range_covering_cell(ws, row_1b, col_1b)
    if m is not None:
        ar, ac = int(m.min_row), int(m.min_col)
        if row_1b == ar and col_1b == ac:
            return ar, ac
        if int(m.max_row) > ar or int(m.max_col) > ac:
            _safe_unmerge_cells(ws, m)
            return row_1b, col_1b
        return ar, ac
    if _is_merged_slave_cell(ws, row_1b, col_1b):
        for c in range(col_1b, 0, -1):
            if not _is_merged_slave_cell(ws, row_1b, c):
                return row_1b, c
        return row_1b, col_1b
    return row_1b, col_1b


def _set_sheet_cell_value_merge_aware(ws, row_1b: int, col_1b: int, value: Any) -> None:
    """
    Write one worksheet coordinate without touching read-only ``MergedCell`` instances.

    When ``(row_1b, col_1b)`` is not the merge anchor, unmerge that range first so only the
    targeted leg changes (vertical journal splits). When it is the anchor, write the anchor cell.
    """
    wr, wc = _resolve_writable_coordinate(ws, row_1b, col_1b)
    _materialize_writable_cell(ws, wr, wc).value = value


def _clear_posting_cell_on_row(ws, row_1b: int, col_1b: int, *, value: Any = "") -> None:
    """
    Clear one posting coordinate on ``row_1b`` without dissolving unrelated merges.

    Date/description vertical merges are left intact by only clearing posting columns.
    """
    _set_sheet_cell_value_merge_aware(ws, int(row_1b), int(col_1b), value)


def _record_has_posting(rec: dict[str, Any], *, amt_tol: float = 0.015) -> bool:
    acct = _norm_account_label(rec.get("account"))
    if not acct:
        return False
    deb = row_amount(rec.get("debit"))
    cre = row_amount(rec.get("credit"))
    return deb > amt_tol or cre > amt_tol


def _snapshot_has_amounts(snap: dict[str, Any], *, amt_tol: float = 0.015) -> bool:
    return row_amount(snap.get("debit")) > amt_tol or row_amount(snap.get("credit")) > amt_tol


def _posting_amount_fields_wiped(
    before: dict[str, Any],
    after: dict[str, Any],
    *,
    amt_tol: float = 0.015,
) -> tuple[bool, bool]:
    """Return ``(debit_wiped, credit_wiped)`` comparing logical debit/credit fields."""
    deb_o = row_amount(before.get("debit"))
    cre_o = row_amount(before.get("credit"))
    deb_n = row_amount(after.get("debit"))
    cre_n = row_amount(after.get("credit"))
    deb_wiped = deb_o > amt_tol and deb_n <= amt_tol
    cre_wiped = cre_o > amt_tol and cre_n <= amt_tol
    return deb_wiped, cre_wiped


def _parse_wiped_posting_amounts(
    new_rec: dict[str, Any],
    prior_rec: dict[str, Any],
    *,
    amt_tol: float = 0.015,
) -> bool:
    """True when re-parse zeroed amounts that existed on this line before faux delete."""
    if not _snapshot_has_amounts(prior_rec, amt_tol=amt_tol):
        return False
    deb_wiped, cre_wiped = _posting_amount_fields_wiped(prior_rec, new_rec, amt_tol=amt_tol)
    return deb_wiped or cre_wiped


def _accounts_allow_amount_preserve(
    prior_rec: dict[str, Any],
    new_rec: dict[str, Any],
) -> bool:
    """False only when both rows have non-empty, different account labels."""
    prior_acct = _norm_account_label(prior_rec.get("account"))
    new_acct = _norm_account_label(new_rec.get("account"))
    return not (prior_acct and new_acct and prior_acct != new_acct)


def _materialize_gl_row_formula_amounts(
    ws,
    colmap: GlColumnMap,
    row_1b: int,
    snap_before: dict[str, Any],
    *,
    credit_style: str,
    tpl_top: int,
) -> None:
    """Replace debit/credit formulas with their last computed numeric values."""
    if not snap_before.get("_debit_formula") and not snap_before.get("_credit_formula"):
        return
    patch: dict[str, Any] = {}
    if snap_before.get("_debit_formula"):
        patch["debit"] = snap_before.get("debit")
    if snap_before.get("_credit_formula"):
        patch["credit"] = snap_before.get("credit")
    if patch:
        _apply_gl_row_snapshot(
            ws, colmap, row_1b, patch, credit_style=credit_style, tpl_top=tpl_top
        )


def _materialize_all_posting_formulas_on_sheet(
    ws,
    colmap: GlColumnMap,
    *,
    credit_style: str,
    tpl_top: int,
    ws_values: Any | None,
) -> None:
    """
    Freeze every debit/credit formula on the GL sheet to its current numeric result.

    Run before faux-deletes so formulas that reference rows about to be cleared still evaluate
    correctly, and unrelated journal lines keep their amounts after delete.
    """
    data_floor = int(colmap.data_start_row)
    max_row = int(ws.max_row or data_floor)
    deb_c = _openpyxl_col(colmap.debit)
    cred_c = _openpyxl_col(colmap.credit)
    for row_1b in range(data_floor, max_row + 1):
        if not (
            _amount_cell_has_formula(ws, row_1b, deb_c)
            or _amount_cell_has_formula(ws, row_1b, cred_c)
        ):
            continue
        snap = _read_gl_row_posting_snapshot(
            ws,
            colmap,
            row_1b,
            credit_style=credit_style,
            ws_values=ws_values,
        )
        _materialize_gl_row_formula_amounts(
            ws,
            colmap,
            row_1b,
            snap,
            credit_style=credit_style,
            tpl_top=tpl_top,
        )


def _restore_gl_row_posting_if_wiped(
    ws,
    colmap: GlColumnMap,
    row_1b: int,
    snap_before: dict[str, Any],
    *,
    credit_style: str,
    tpl_top: int,
    ws_values: Any | None = None,
    amt_tol: float = 0.015,
) -> None:
    """Rewrite workbook amounts cleared as collateral when a sibling leg was faux-deleted."""
    if not _snapshot_has_amounts(snap_before, amt_tol=amt_tol):
        return
    cur = _read_gl_row_posting_snapshot(
        ws, colmap, row_1b, credit_style=credit_style, ws_values=ws_values
    )
    if not _accounts_allow_amount_preserve(snap_before, cur):
        return
    deb_wiped, cre_wiped = _posting_amount_fields_wiped(snap_before, cur, amt_tol=amt_tol)
    patch: dict[str, Any] = {}
    if deb_wiped:
        patch["debit"] = snap_before.get("debit")
    if cre_wiped:
        patch["credit"] = snap_before.get("credit")
    if _norm_account_label(snap_before.get("account")) and not _norm_account_label(cur.get("account")):
        patch["account"] = snap_before.get("account")
    if patch:
        patch["_ws_values"] = ws_values
        _apply_gl_row_snapshot(
            ws, colmap, row_1b, patch, credit_style=credit_style, tpl_top=tpl_top
        )


def _restore_journal_block_after_faux_clear(
    ws,
    colmap: GlColumnMap,
    cleared_row: int,
    block_snaps: dict[int, dict[str, Any]],
    *,
    credit_style: str,
    tpl_top: int,
    ws_values: Any | None = None,
) -> None:
    for row_1b, snap in block_snaps.items():
        if int(row_1b) == int(cleared_row):
            continue
        _materialize_gl_row_formula_amounts(
            ws,
            colmap,
            int(row_1b),
            snap,
            credit_style=credit_style,
            tpl_top=tpl_top,
        )
        _restore_gl_row_posting_if_wiped(
            ws,
            colmap,
            int(row_1b),
            snap,
            credit_style=credit_style,
            tpl_top=tpl_top,
            ws_values=ws_values,
        )


def stabilize_gl_records_after_faux_delete(
    prior: list[dict[str, Any]],
    parsed: list[dict[str, Any]],
    deleted_excel_rows: set[int] | list[int],
    *,
    amt_tol: float = 0.015,
) -> list[dict[str, Any]]:
    """
    After faux delete + workbook re-read, restore posting amounts for untouched lines.

    Rows in ``deleted_excel_rows`` are dropped. When the parser returns 0/empty debit or credit
    for a surviving ``_excel_row`` that still had a posting before delete, keep the prior amounts
    (and account if the new parse lost it).
    """
    deleted = {int(x) for x in deleted_excel_rows if int(x) > 0}
    prior_by_er: dict[int, dict[str, Any]] = {}
    for rec in prior:
        er = rec.get("_excel_row")
        if isinstance(er, int) and er > 0:
            prior_by_er[int(er)] = rec

    out: list[dict[str, Any]] = []
    for rec in parsed:
        er = rec.get("_excel_row")
        if not isinstance(er, int) or er <= 0:
            out.append(dict(rec))
            continue
        if er in deleted:
            continue
        prior_rec = prior_by_er.get(er)
        if prior_rec is None:
            out.append(dict(rec))
            continue
        merged = dict(rec)
        deb_wiped, cre_wiped = _posting_amount_fields_wiped(prior_rec, rec, amt_tol=amt_tol)
        if deb_wiped and _accounts_allow_amount_preserve(prior_rec, rec):
            merged["debit"] = prior_rec.get("debit")
        elif deb_wiped and row_amount(prior_rec.get("debit")) > amt_tol:
            merged["debit"] = prior_rec.get("debit")
        if cre_wiped and _accounts_allow_amount_preserve(prior_rec, rec):
            merged["credit"] = prior_rec.get("credit")
        elif cre_wiped and row_amount(prior_rec.get("credit")) > amt_tol:
            merged["credit"] = prior_rec.get("credit")
        if not _norm_account_label(merged.get("account")) and _norm_account_label(
            prior_rec.get("account")
        ):
            merged["account"] = prior_rec.get("account")
        out.append(merged)

    seen = {int(r.get("_excel_row") or 0) for r in out if int(r.get("_excel_row") or 0) > 0}
    for er, prior_rec in prior_by_er.items():
        if er in deleted or er in seen:
            continue
        if _record_has_posting(prior_rec, amt_tol=amt_tol):
            out.append(dict(prior_rec))
    out.sort(key=lambda r: int(r.get("_excel_row") or 0))
    return out


def faux_clear_gl_sheet_row(
    ws,
    row_1b: int,
    colmap: GlColumnMap,
    *,
    max_col: int,
) -> None:
    """
    Faux delete: clear posting content but keep the Excel row (no ``delete_rows``).

    Clears account / debit / credit / Tr (not date or memo merges) so sibling journal legs keep
    their amounts and shared date stripes stay merged.

    The row is omitted from GL reads/counts because :func:`_row_tuple_to_record_dynamic` skips
    lines with no account and no debit/credit amounts.
    """
    row_1b = int(row_1b)
    if row_1b < 1:
        return
    for col_1b in _posting_column_indices_1b(colmap, max_col_1b=int(max_col)):
        _clear_posting_cell_on_row(ws, row_1b, col_1b, value="")


def _delete_gl_row_merge_aware(
    ws,
    row_1b: int,
    colmap: GlColumnMap,
    *,
    max_col: int,
) -> None:
    """Faux-delete one GL line (clear cells; row stays on sheet)."""
    faux_clear_gl_sheet_row(ws, int(row_1b), colmap, max_col=max_col)


def apply_gl_edit_plan(
    workbook_path: str,
    plan: GlEditPlan,
    *,
    sheet_name: str = GL_SHEET_NAME_DEFAULT,
    layout: Optional[dict[str, Any]] = None,
) -> None:
    """
    Apply row updates, faux-deletes (clear rows), inserts (between rows or at tail), and swaps.

    ``delete_rows`` clears workbook lines in place; empty rows are excluded when the GL is read.
    Each insert dict may include ``insert_after`` (1-based Excel row); omit or ``0`` to append at the
    sheet tail. Pairwise swaps exchange cell values only (sheet row indices unchanged).
    """
    if not (
        plan.updates
        or plan.delete_rows
        or plan.insert_rows
        or plan.swap_rows
    ):
        return

    delete_rows = sorted({int(x) for x in plan.delete_rows if int(x) >= 1})
    _max_gl_delete_batch = 100
    if len(delete_rows) > _max_gl_delete_batch:
        raise ValueError(
            f"Refusing to clear {len(delete_rows)} rows at once (limit {_max_gl_delete_batch}). "
            "Select fewer workbook rows in the GL editor."
        )

    wb, ws, colmap, max_c, credit_style, wb_v, ws_v = _open_gl_sheet_for_write(
        workbook_path, sheet_name=sheet_name, layout=layout
    )
    date_c = _openpyxl_col(colmap.date)
    desc_c = _openpyxl_col(colmap.details)
    tpl_top, tpl_bot = infer_last_entry_template_rows(ws, ws.max_row or 1, date_c, desc_c)

    try:
        if plan.delete_row_checks and delete_rows:
            verify_err = verify_gl_delete_targets(
                ws,
                colmap,
                plan.delete_row_checks,
                credit_style=credit_style,
                ws_values=ws_v,
            )
            if verify_err:
                raise ValueError(verify_err)

        for row_a, row_b in plan.swap_rows:
            ra, rb = int(row_a), int(row_b)
            if ra < 1 or rb < 1 or ra == rb:
                continue
            snap_a = _read_gl_row_snapshot(
                ws, colmap, ra, credit_style=credit_style, ws_values=ws_v
            )
            snap_b = _read_gl_row_snapshot(
                ws, colmap, rb, credit_style=credit_style, ws_values=ws_v
            )
            _apply_gl_row_snapshot(ws, colmap, ra, snap_b, credit_style=credit_style, tpl_top=tpl_top)
            _apply_gl_row_snapshot(ws, colmap, rb, snap_a, credit_style=credit_style, tpl_top=tpl_top)

        for spec in plan.updates:
            er = spec.get("excel_row")
            if not isinstance(er, int) or er < colmap.data_start_row:
                continue
            if er in plan.delete_rows:
                continue
            fields = {k: v for k, v in spec.items() if k != "excel_row"}
            if fields:
                _set_gl_row_fields(ws, colmap, er, fields, credit_style=credit_style, tpl_top=tpl_top)

        data_floor = int(colmap.data_start_row)
        if delete_rows:
            _materialize_all_posting_formulas_on_sheet(
                ws,
                colmap,
                credit_style=credit_style,
                tpl_top=tpl_top,
                ws_values=ws_v,
            )
        for er in sorted({int(x) for x in delete_rows if int(x) >= data_floor}):
            block_lo, block_hi = _journal_block_bounds_workbook(
                ws, colmap, er, date_c, desc_c
            )
            block_snaps = {
                r: _read_gl_row_posting_snapshot(
                    ws, colmap, r, credit_style=credit_style, ws_values=ws_v
                )
                for r in range(int(block_lo), int(block_hi) + 1)
            }
            faux_clear_gl_sheet_row(ws, er, colmap, max_col=max_c)
            _restore_journal_block_after_faux_clear(
                ws,
                colmap,
                er,
                block_snaps,
                credit_style=credit_style,
                tpl_top=tpl_top,
                ws_values=ws_v,
            )

        if plan.insert_rows:
            spec_err = validate_gl_insert_specs_have_accounts(plan.insert_rows)
            if spec_err:
                raise ValueError(spec_err)
            # Freeze formula amounts before row insert; openpyxl row shifts break references.
            _materialize_all_posting_formulas_on_sheet(
                ws,
                colmap,
                credit_style=credit_style,
                tpl_top=tpl_top,
                ws_values=ws_v,
            )
            tail_specs: list[dict[str, Any]] = []
            by_after: dict[int, list[dict[str, Any]]] = {}
            for spec in plan.insert_rows:
                row = dict(spec)
                after_raw = row.pop("insert_after", 0)
                try:
                    after = int(after_raw or 0)
                except (TypeError, ValueError):
                    after = 0
                if after > 0:
                    exact = bool(row.pop("insert_below_exact", False))
                    if exact:
                        resolved = after
                    else:
                        resolved = resolve_gl_insert_after_row(ws, colmap, after)
                    row["_insert_below_exact"] = exact
                    by_after.setdefault(resolved, []).append(row)
                else:
                    tail_specs.append(row)
            anchor_block_snaps: dict[int, dict[int, dict[str, Any]]] = {}
            for after in by_after:
                block_lo, block_hi = _journal_block_bounds_workbook(
                    ws, colmap, int(after), date_c, desc_c
                )
                anchor_block_snaps[int(after)] = {
                    r: _read_gl_row_posting_snapshot(
                        ws,
                        colmap,
                        r,
                        credit_style=credit_style,
                        ws_values=ws_v,
                    )
                    for r in range(int(block_lo), int(block_hi) + 1)
                }
            inserted_rows: list[int] = []
            for after in sorted(by_after.keys(), reverse=True):
                batch = _sort_insert_specs_debit_first(by_after[after])
                exact = bool(batch and batch[0].get("_insert_below_exact"))
                inserted_rows.extend(
                    _insert_gl_rows_after(
                        ws,
                        colmap,
                        after,
                        batch,
                        max_col=max_c,
                        insert_below_exact=exact,
                    )
                )
            for block_snaps in anchor_block_snaps.values():
                for row_1b, snap in block_snaps.items():
                    _restore_gl_row_posting_if_wiped(
                        ws,
                        colmap,
                        int(row_1b),
                        snap,
                        credit_style=credit_style,
                        tpl_top=tpl_top,
                        ws_values=ws_v,
                    )
            if tail_specs:
                tail_err = validate_gl_insert_specs_have_accounts(tail_specs)
                if tail_err:
                    raise ValueError(tail_err)
                inserted_rows.extend(
                    _append_gl_rows_at_sheet_tail(ws, colmap, tail_specs, max_col=max_c)
                )
            if inserted_rows:
                part_err = verify_gl_posting_rows_have_particulars(
                    ws,
                    colmap,
                    inserted_rows,
                    credit_style=credit_style,
                )
                if part_err:
                    raise ValueError(part_err)

        wb.save(workbook_path)
    finally:
        _close_gl_workbooks(wb, wb_v)
