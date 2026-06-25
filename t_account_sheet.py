"""
Parse Finjan-style spatial T-account worksheets (grid of mini T-accounts on one sheet).

Primary use: supplement journal GL with opening / brought-forward rows that appear only on the
T-account sheet (e.g. Cash «QAR 6,167 Beg Balance»). Journal activity stays on the GL sheet.
"""

from __future__ import annotations

from dataclasses import dataclass, field
import re
from datetime import date
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

import gl_analytics as gla

DEFAULT_CURRENCY_ISO = "QAR"
_AMT_TOL = 0.015
_MAX_SCAN_ROWS = 200_000
_MAX_SCAN_COLS = 60

_DATE_HEADER = frozenset({"date"})
_SKIP_HEADER_TEXT = frozenset(
    {
        "date",
        "debit",
        "credit",
        "particulars",
        "accounts",
        "account",
        "journal entries template",
        "journal entries",
        "total",
    }
)
_BEG_BALANCE_RE = re.compile(
    r"(?:\bbeg(?:inning)?\s*bal(?:ance)?\b|\bbrought\s+(?:forward|fwd)\b|\bcarried\s+forward\b|\bb/?\s*f(?:orward)?\b)",
    re.I,
)
_MONEY_NUM_RE = re.compile(r"-?[\d,]+(?:\.\d+)?")
_ISO_TOKEN_RE = re.compile(r"\b([A-Z]{3})\b")


def infer_t_accounts_sheet_name(sheet_names: list[str], gl_sheet: str) -> Optional[str]:
    """Pick a T-account sheet when Settings names a matching worksheet (never a random non-GL tab)."""
    gl_norm = (gl_sheet or "").strip()
    for name in sheet_names:
        if name == gl_norm:
            continue
        ln = name.lower().replace("_", "-")
        if "t-account" in ln or "t account" in ln or ln in ("t-accounts", "taccounts"):
            return name
    return None


def _norm_cell(val: Any) -> str:
    return re.sub(r"\s+", " ", str(val or "").strip())


def _norm_key(val: Any) -> str:
    return _norm_cell(val).casefold()


def _sniff_iso(text: str) -> str:
    m = _ISO_TOKEN_RE.search(text.upper())
    return m.group(1) if m else DEFAULT_CURRENCY_ISO


def _parse_ta_amount(val: Any) -> tuple[Optional[float], str, bool]:
    """
    Parse amount cells like ``QAR 6,167 Beg Balance``.

    Returns ``(amount, currency_iso, is_opening_label)``.
    """
    if val is None:
        return None, DEFAULT_CURRENCY_ISO, False
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val), DEFAULT_CURRENCY_ISO, False
    s = _norm_cell(val)
    if not s:
        return None, DEFAULT_CURRENCY_ISO, False
    is_opening = bool(_BEG_BALANCE_RE.search(s)) or gla.is_brought_forward_text(s)
    iso = _sniff_iso(s)
    nums = _MONEY_NUM_RE.findall(s.replace(",", ""))
    if not nums:
        return None, iso, is_opening
    try:
        amt = float(nums[0].replace(",", ""))
    except ValueError:
        return None, iso, is_opening
    return amt, iso, is_opening


def _looks_like_date(val: Any) -> bool:
    if val is None:
        return False
    if hasattr(val, "year"):
        return True
    s = str(val).strip()
    if not s:
        return False
    if re.match(r"^\d{1,2}[-/]\w{3,}", s, re.I):
        return True
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return True
    return False


def _coerce_tac_date(val: Any, *, default_year: int) -> Optional[date]:
    if val is None:
        return None
    if hasattr(val, "date") and callable(val.date):
        d = val.date()
    elif hasattr(val, "year"):
        d = val
    else:
        from excel_engine import _coerce_journal_sheet_date

        d = _coerce_journal_sheet_date(val)
        if d is None:
            return None
    if d.year < 1900:
        try:
            return d.replace(year=int(default_year))
        except ValueError:
            return d
    return d


def _fill_merged_grid(grid: list[list[Any]], ws) -> None:
    nrow = len(grid)
    ncol = len(grid[0]) if nrow else 0
    for mrange in ws.merged_cells.ranges:
        r1, r2 = mrange.min_row - 1, mrange.max_row - 1
        c1, c2 = mrange.min_col - 1, mrange.max_col - 1
        if nrow == 0 or ncol == 0 or r1 >= nrow or c1 >= ncol:
            continue
        r2 = min(r2, nrow - 1)
        c2 = min(c2, ncol - 1)
        anchor = grid[r1][c1]
        if anchor is None:
            continue
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                if cc < len(grid[rr]):
                    grid[rr][cc] = anchor


def _sheet_grid(ws, max_r: int, max_c: int) -> list[list[Any]]:
    grid: list[list[Any]] = []
    for r in range(max_r):
        row_vals: list[Any] = []
        for c in range(max_c):
            row_vals.append(ws.cell(row=r + 1, column=c + 1).value)
        grid.append(row_vals)
    return grid


@dataclass
class TAccountBlock:
    account: str
    subheader_row_0: int
    end_row_0: int
    debit_date_col: int
    credit_date_col: Optional[int]
    col_lo: int
    col_hi: int
    footer_total: Optional[float] = None
    footer_currency_iso: str = DEFAULT_CURRENCY_ISO
    lines: list[dict[str, Any]] = field(default_factory=list)


def _plausible_account_title(text: str) -> bool:
    t = _norm_cell(text)
    if len(t) < 2:
        return False
    key = t.casefold()
    if key in _SKIP_HEADER_TEXT:
        return False
    if key.startswith("total "):
        return False
    if _looks_like_date(t):
        return False
    if _MONEY_NUM_RE.search(t.replace(",", "")) and _sniff_iso(t) in t.upper():
        return False
    return True


def _account_title_above(grid: list[list[Any]], sub_row_0: int, col_lo: int, col_hi: int) -> str:
  best = ""
  for look in (sub_row_0 - 1, sub_row_0 - 2):
    if look < 0:
      continue
    for c in range(col_lo, min(col_hi + 1, len(grid[look]))):
      t = _norm_cell(grid[look][c])
      if _plausible_account_title(t) and len(t) > len(best):
        best = t
  return best


def _row_mostly_empty(grid: list[list[Any]], row_0: int, col_lo: int, col_hi: int) -> bool:
    if row_0 < 0 or row_0 >= len(grid):
        return True
    row = grid[row_0]
    for c in range(col_lo, min(col_hi + 1, len(row))):
      v = row[c]
      if v is not None and str(v).strip():
        return False
    return True


def _cluster_date_column_groups(date_cols: list[int]) -> list[list[int]]:
    """
    Pair each «Date» header with its credit-side «Date» (within a few columns).

    Side-by-side T-accounts on one row yield multiple disjoint pairs instead of one wide band.
    """
    sorted_cols = sorted(set(date_cols))
    groups: list[list[int]] = []
    used: set[int] = set()
    for c in sorted_cols:
        if c in used:
            continue
        group = [c]
        used.add(c)
        for c2 in sorted_cols:
            if c2 in used:
                continue
            if c < c2 <= c + 4:
                group.append(c2)
                used.add(c2)
                break
        groups.append(group)
    return groups


def _find_tac_blocks(grid: list[list[Any]]) -> list[TAccountBlock]:
    nrow = len(grid)
    ncol = len(grid[0]) if nrow else 0
    candidates: list[tuple[int, list[int]]] = []
    for r in range(nrow):
        date_cols = [
            c
            for c in range(ncol)
            if _norm_key(grid[r][c]) in _DATE_HEADER
        ]
        if not date_cols:
            continue
        for group in _cluster_date_column_groups(date_cols):
            candidates.append((r, group))

    blocks: list[TAccountBlock] = []
    for idx, (sub_r, date_cols) in enumerate(candidates):
        debit_date_col = date_cols[0]
        credit_date_col = (
            date_cols[-1]
            if len(date_cols) > 1 and date_cols[-1] > debit_date_col
            else None
        )
        if credit_date_col is not None and credit_date_col - debit_date_col < 2:
            credit_date_col = None
        title_lo = debit_date_col
        title_hi = credit_date_col if credit_date_col is not None else debit_date_col + 2
        title_hi = min(title_hi, ncol - 1)

        account = _account_title_above(grid, sub_r, title_lo, title_hi)
        if not account:
            continue

        col_lo = max(0, debit_date_col - 1)
        col_hi = credit_date_col + 3 if credit_date_col is not None else debit_date_col + 4
        col_hi = min(col_hi, ncol - 1)

        end_r = sub_r
        next_sub = nrow
        for j in range(idx + 1, len(candidates)):
            nr, ng = candidates[j]
            if nr > sub_r and ng[0] >= col_lo and ng[0] <= col_hi:
                next_sub = nr
                break
            if nr > sub_r and ng[0] > col_hi:
                next_sub = nr
                break
        empty_run = 0
        for r in range(sub_r + 1, min(next_sub, nrow)):
            if _row_mostly_empty(grid, r, col_lo, col_hi):
                empty_run += 1
                if empty_run >= 2:
                    break
                continue
            empty_run = 0
            end_r = r

        blocks.append(
            TAccountBlock(
                account=account,
                subheader_row_0=sub_r,
                end_row_0=end_r,
                debit_date_col=debit_date_col,
                credit_date_col=credit_date_col,
                col_lo=col_lo,
                col_hi=col_hi,
            )
        )
    return blocks


def _side_amount_cols(date_col: int, col_hi: int) -> list[int]:
    cols = [date_col + 1, date_col + 2, date_col + 3]
    out: list[int] = []
    for c in cols:
        if c <= col_hi:
            out.append(c)
    return out


def _best_amount_in_row(
    grid: list[list[Any]],
    row_0: int,
    cols: list[int],
) -> tuple[Optional[float], str, bool, int]:
    best_amt: Optional[float] = None
    best_iso = DEFAULT_CURRENCY_ISO
    best_open = False
    best_col = -1
    for c in cols:
        if c >= len(grid[row_0]):
            continue
        amt, iso, is_open = _parse_ta_amount(grid[row_0][c])
        if amt is not None and abs(amt) > _AMT_TOL:
            if best_amt is None or abs(amt) > abs(best_amt):
                best_amt = amt
                best_iso = iso
                best_open = is_open
                best_col = c
        # opening label may sit in the cell next to the amount
        for adj in (c - 1, c + 1):
            if 0 <= adj < len(grid[row_0]):
                adj_t = _norm_cell(grid[row_0][adj])
                if _BEG_BALANCE_RE.search(adj_t) or gla.is_brought_forward_text(adj_t):
                    best_open = True
    return best_amt, best_iso, best_open, best_col


def _parse_block_lines(
    grid: list[list[Any]],
    block: TAccountBlock,
    *,
    default_year: int,
) -> None:
    sub_r = block.subheader_row_0
    debit_amt_cols = _side_amount_cols(block.debit_date_col, block.col_hi)
    credit_amt_cols: list[int] = []
    if block.credit_date_col is not None:
        credit_amt_cols = _side_amount_cols(block.credit_date_col, block.col_hi)

    lines: list[dict[str, Any]] = []
    footer_amt: Optional[float] = None
    footer_iso = DEFAULT_CURRENCY_ISO

    for r in range(sub_r + 1, block.end_row_0 + 1):
        row = grid[r]
        if len(row) <= block.col_lo:
            continue
        d_date = row[block.debit_date_col] if block.debit_date_col < len(row) else None
        deb_amt, deb_iso, deb_open, deb_col = _best_amount_in_row(grid, r, debit_amt_cols)
        cre_amt: Optional[float] = None
        cre_iso = DEFAULT_CURRENCY_ISO
        cre_open = False
        c_date = None
        if block.credit_date_col is not None and credit_amt_cols:
            c_date = row[block.credit_date_col] if block.credit_date_col < len(row) else None
            cre_amt, cre_iso, cre_open, _ = _best_amount_in_row(grid, r, credit_amt_cols)

        has_date = _looks_like_date(d_date) or _looks_like_date(c_date)
        if deb_amt is not None and abs(deb_amt) > _AMT_TOL:
            if deb_open:
                lines.append(
                    {
                        "side": "debit",
                        "amount": deb_amt,
                        "currency_iso": deb_iso,
                        "gl_date": _coerce_tac_date(d_date or c_date, default_year=default_year),
                        "is_opening": True,
                        "excel_row_1b": r + 1,
                    }
                )
            elif not has_date and deb_col >= 0:
                footer_amt = deb_amt
                footer_iso = deb_iso

        if cre_amt is not None and abs(cre_amt) > _AMT_TOL:
            if cre_open:
                lines.append(
                    {
                        "side": "credit",
                        "amount": cre_amt,
                        "currency_iso": cre_iso,
                        "gl_date": _coerce_tac_date(c_date or d_date, default_year=default_year),
                        "is_opening": True,
                        "excel_row_1b": r + 1,
                    }
                )
            elif not has_date:
                footer_amt = cre_amt
                footer_iso = cre_iso

    block.lines = lines
    block.footer_total = footer_amt
    block.footer_currency_iso = footer_iso


def parse_t_account_sheet_blocks(
    workbook_path: str,
    sheet_name: str,
    *,
    default_year: Optional[int] = None,
) -> list[TAccountBlock]:
    """Detect T-account regions and parse line-level debits/credits per block."""
    yr = int(default_year or date.today().year)
    wb = load_workbook(workbook_path, read_only=False, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        max_r = min(ws.max_row or 1, _MAX_SCAN_ROWS)
        max_c = min(ws.max_column or 1, _MAX_SCAN_COLS)
        grid = _sheet_grid(ws, max_r, max_c)
        _fill_merged_grid(grid, ws)
        blocks = _find_tac_blocks(grid)
        for block in blocks:
            _parse_block_lines(grid, block, default_year=yr)
        return blocks
    finally:
        wb.close()


def opening_gl_records_from_blocks(
    blocks: list[TAccountBlock],
    *,
    source_sheet: str,
) -> list[dict[str, Any]]:
    """Convert opening / beg-balance T-account lines into GL-shaped records."""
    out: list[dict[str, Any]] = []
    for block in blocks:
        for line in block.lines:
            if not line.get("is_opening"):
                continue
            amt = float(line.get("amount") or 0)
            if abs(amt) <= _AMT_TOL:
                continue
            side = str(line.get("side") or "debit")
            deb = amt if side == "debit" else 0.0
            cre = amt if side == "credit" else 0.0
            gd = line.get("gl_date")
            gd_s = gd.isoformat() if hasattr(gd, "isoformat") else str(gd or "")
            rec: dict[str, Any] = {
                "gl_date": gd_s,
                "description": gla.BROUGHT_FORWARD_LABEL,
                "account": block.account,
                "debit": deb,
                "credit": cre,
                "currency_iso": str(line.get("currency_iso") or DEFAULT_CURRENCY_ISO),
                "opening_balance": True,
                "brought_forward": True,
                "source_sheet": source_sheet,
                "source": "t_account_sheet",
            }
            er = line.get("excel_row_1b")
            if isinstance(er, int) and er > 0:
                rec["_excel_row"] = er
            out.append(rec)
    return out


def extract_t_account_opening_supplements(
    workbook_path: str,
    sheet_name: str,
    *,
    default_year: Optional[int] = None,
) -> list[dict[str, Any]]:
    """Read one worksheet and return opening/BF GL records for journal supplementation."""
    blocks = parse_t_account_sheet_blocks(
        workbook_path, sheet_name, default_year=default_year
    )
    return opening_gl_records_from_blocks(blocks, source_sheet=sheet_name)


def reconcile_t_account_footers(
    blocks: list[TAccountBlock],
    account_net: dict[str, float],
    *,
    tol: float = 0.02,
) -> list[dict[str, Any]]:
    """
    Compare parsed footer totals to journal-derived net balances (debit − credit).

    Returns mismatch rows: ``account``, ``footer_total``, ``journal_net``, ``delta``.
    """
    import account_buckets as ab

    mismatches: list[dict[str, Any]] = []
    for block in blocks:
        if block.footer_total is None:
            continue
        key = ab.fold_account_key(block.account)
        jnet = account_net.get(key)
        if jnet is None:
            continue
        delta = float(block.footer_total) - float(jnet)
        if abs(delta) > tol:
            mismatches.append(
                {
                    "account": block.account,
                    "footer_total": float(block.footer_total),
                    "journal_net": float(jnet),
                    "delta": delta,
                }
            )
    return mismatches


def merge_t_account_opening_supplements(
    journal_records: list[dict[str, Any]],
    supplement_records: list[dict[str, Any]],
    *,
    amt_tol: float = _AMT_TOL,
) -> list[dict[str, Any]]:
    """
    Append T-account opening rows when the journal lacks an equivalent opening/BF for that account.
    """
    if not supplement_records:
        return list(journal_records)

    import account_buckets as ab

    def _opening_net(rec: dict[str, Any]) -> float:
        deb = float(rec.get("debit") or 0)
        cre = float(rec.get("credit") or 0)
        return deb - cre

    existing_keys: set[tuple[str, float]] = set()
    for rec in journal_records:
        if not gla.is_opening_or_brought_forward_rec(rec):
            continue
        key = ab.fold_account_key(str(rec.get("account") or ""))
        net = _opening_net(rec)
        if abs(net) > amt_tol:
            existing_keys.add((key, round(net, 2)))

    out = list(journal_records)
    for sup in supplement_records:
        key = ab.fold_account_key(str(sup.get("account") or ""))
        net = _opening_net(sup)
        if abs(net) <= amt_tol:
            continue
        sig = (key, round(net, 2))
        if sig in existing_keys:
            continue
        # skip if any opening already exists for this account (avoid duplicate/conflict)
        if any(
            ab.fold_account_key(str(r.get("account") or "")) == key
            and gla.is_opening_or_brought_forward_rec(r)
            for r in journal_records
        ):
            continue
        out.append(dict(sup))
        existing_keys.add(sig)
    return out


def read_t_account_opening_from_workbook(
    workbook_path: str,
    *,
    gl_sheet_name: str,
    t_accounts_sheet_name: Optional[str] = None,
    default_year: Optional[int] = None,
) -> list[dict[str, Any]]:
    """
    Return opening/BF supplement rows from the configured T-account worksheet only.

    Requires an explicit sheet name in Settings (or a worksheet whose name matches
    «t-account» / «t-accounts»). Never scans arbitrary non-GL tabs.
    """
    ext = Path(workbook_path).suffix.lower()
    if ext not in (".xlsx", ".xlsm"):
        return []
    ta_sheet = (t_accounts_sheet_name or "").strip()
    if not ta_sheet:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            ta_sheet = infer_t_accounts_sheet_name(list(wb.sheetnames), gl_sheet_name) or ""
        finally:
            wb.close()
    if not ta_sheet or ta_sheet == gl_sheet_name:
        return []
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if ta_sheet not in wb.sheetnames:
            return []
    finally:
        wb.close()
    return extract_t_account_opening_supplements(
        workbook_path, ta_sheet, default_year=default_year
    )
