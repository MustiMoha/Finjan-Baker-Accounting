"""
openpyxl row insert/delete with merged-range bookkeeping.

Streamlit / pandas are not used here — the worksheet is the source of truth.
"""

from __future__ import annotations

from typing import TYPE_CHECKING, Iterable

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

MergeQuad = tuple[int, int, int, int]  # min_row, min_col, max_row, max_col (1-based)


def merged_ranges_on_row(ws: Worksheet, row_1b: int) -> list[MergeQuad]:
    """Return every merged rectangle that includes ``row_1b``."""
    row_1b = int(row_1b)
    out: list[MergeQuad] = []
    for m in ws.merged_cells.ranges:
        if int(m.min_row) <= row_1b <= int(m.max_row):
            out.append((int(m.min_row), int(m.min_col), int(m.max_row), int(m.max_col)))
    return out


def _unmerge_quads(ws: Worksheet, quads: Iterable[MergeQuad]) -> None:
    seen: set[str] = set()
    for mr1, mc1, mr2, mc2 in quads:
        ref = f"{ws.cell(mr1, mc1).coordinate}:{ws.cell(mr2, mc2).coordinate}"
        if ref not in seen:
            seen.add(ref)
            try:
                ws.unmerge_cells(ref)
            except KeyError:
                for m in list(ws.merged_cells.ranges):
                    if str(m) == ref:
                        try:
                            ws.merged_cells.ranges.discard(m)
                        except Exception:
                            ws.merged_cells.remove(m)
                        break
                from openpyxl.cell.cell import Cell, MergedCell

                for r in range(int(mr1), int(mr2) + 1):
                    for c in range(int(mc1), int(mc2) + 1):
                        if (r, c) == (int(mr1), int(mc1)):
                            continue
                        existing = ws._cells.get((r, c))
                        if existing is None or isinstance(existing, MergedCell):
                            ws._cells[(r, c)] = Cell(ws, row=r, column=c)


def _adjust_merge_after_row_delete(quad: MergeQuad, deleted_row: int) -> MergeQuad | None:
    """
    Compute merged range coordinates after ``deleted_row`` is removed.

    openpyxl will call ``delete_rows(deleted_row, 1)`` so every row below shifts up by one.
    """
    mr1, mc1, mr2, mc2 = quad
    deleted_row = int(deleted_row)

    if mr2 < deleted_row:
        return quad
    if mr1 > deleted_row:
        return (mr1 - 1, mc1, mr2 - 1, mc2)
    if mr1 == mr2 == deleted_row:
        return None
    # Deleted row lies inside the vertical span — shrink the bottom edge by one.
    return (mr1, mc1, mr2 - 1, mc2)


def delete_sheet_row(ws: Worksheet, row_1b: int) -> None:
    """
    Delete exactly one worksheet row and fix merged regions that intersect it.

    Steps:
    1. Snapshot merge quads touching this row.
    2. Unmerge them (openpyxl cannot delete rows through active merges cleanly).
    3. ``ws.delete_rows(row_1b, 1)`` — openpyxl shifts row indices at ``row_1b+1..max`` up by 1.
    4. Re-merge each adjusted quad (single-row merges that covered only the deleted row are dropped).
    """
    row_1b = int(row_1b)
    if row_1b < 1:
        return

    touched = merged_ranges_on_row(ws, row_1b)
    if touched:
        _unmerge_quads(ws, touched)

    # Structural delete — formulas, styles, and non-merged values on lower rows move up one row.
    ws.delete_rows(row_1b, 1)

    for quad in touched:
        adjusted = _adjust_merge_after_row_delete(quad, row_1b)
        if adjusted is None:
            continue
        nr1, nc1, nr2, nc2 = adjusted
        if nr2 < nr1:
            continue
        ws.merge_cells(start_row=nr1, start_column=nc1, end_row=nr2, end_column=nc2)


def insert_sheet_rows(ws: Worksheet, row_1b: int, *, amount: int = 1) -> None:
    """
    Insert blank rows at ``row_1b`` and expand merged ranges that include that index.

    Rows at or below ``row_1b`` shift down by ``amount``. Merges strictly above are unchanged;
    merges at/after the insertion point have their bounds shifted down.
    """
    row_1b = int(row_1b)
    amount = int(amount)
    if row_1b < 1 or amount < 1:
        return

    # Merges that start at or below the insertion row need expansion/shift.
    expand: list[MergeQuad] = []
    for m in ws.merged_cells.ranges:
        if int(m.min_row) >= row_1b or int(m.max_row) >= row_1b:
            if int(m.min_row) <= row_1b <= int(m.max_row):
                expand.append((int(m.min_row), int(m.min_col), int(m.max_row), int(m.max_col)))
            elif int(m.min_row) > row_1b:
                expand.append((int(m.min_row), int(m.min_col), int(m.max_row), int(m.max_col)))

    if expand:
        _unmerge_quads(ws, expand)

    ws.insert_rows(row_1b, amount)

    for mr1, mc1, mr2, mc2 in expand:
        if mr1 <= row_1b <= mr2:
            ws.merge_cells(
                start_row=mr1,
                start_column=mc1,
                end_row=mr2 + amount,
                end_column=mc2,
            )
        elif mr1 > row_1b:
            ws.merge_cells(
                start_row=mr1 + amount,
                start_column=mc1,
                end_row=mr2 + amount,
                end_column=mc2,
            )
