"""Fill admin-uploaded Excel templates with statement data (openpyxl only; no pandas writes)."""

from __future__ import annotations

from copy import copy
from html import escape
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

import gl_analytics as gla
from ui_locale import html_rtl_css, tr


def _anchor_sheet_cell(wb, defined_name: str):
    """Return (worksheet, row, col_index) for a single-cell defined name; fallback A6 on active sheet."""
    ws_fb = wb.active
    if defined_name not in wb.defined_names:
        return ws_fb, 6, 1
    dn = wb.defined_names[defined_name]
    dests = list(dn.destinations)
    if not dests:
        return ws_fb, 6, 1
    sheet_title, coord = dests[0]
    coord_clean = str(coord).replace("$", "")
    try:
        col_ltr, row = coordinate_from_string(coord_clean)
        col_ix = column_index_from_string(col_ltr)
        return wb[sheet_title], row, col_ix
    except Exception:
        return ws_fb, 6, 1


def _write_scalar_named(wb, name: str, value: Any) -> None:
    if name not in wb.defined_names:
        return
    ws, row, col_ix = _anchor_sheet_cell(wb, name)
    ws.cell(row=row, column=col_ix, value=value)


def trial_balance_to_xlsx_bytes(template_bytes: bytes, tb: pd.DataFrame, *, param_as_of: str = "") -> bytes:
    bio = BytesIO(template_bytes)
    wb = load_workbook(bio)
    ws, row0, c0 = _anchor_sheet_cell(wb, "DATA_BODY")
    if param_as_of:
        _write_scalar_named(wb, "PARAM_AS_OF", param_as_of)
    r = row0
    for _, rec in tb.iterrows():
        ws.cell(row=r, column=c0, value=str(rec.get("account", "") or ""))
        d_cell = pd.to_numeric(rec.get("debits", 0), errors="coerce")
        c_cell = pd.to_numeric(rec.get("credits", 0), errors="coerce")
        nb_cell = pd.to_numeric(rec.get("net_balance", 0), errors="coerce")
        ws.cell(row=r, column=c0 + 1, value=0.0 if pd.isna(d_cell) else float(d_cell))
        ws.cell(row=r, column=c0 + 2, value=0.0 if pd.isna(c_cell) else float(c_cell))
        ws.cell(row=r, column=c0 + 3, value=0.0 if pd.isna(nb_cell) else float(nb_cell))
        r += 1
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def income_summary_to_xlsx_bytes(template_bytes: bytes, totals: dict[str, float], *, param_period: str = "") -> bytes:
    bio = BytesIO(template_bytes)
    wb = load_workbook(bio)
    ws, row0, c0 = _anchor_sheet_cell(wb, "DATA_BODY")
    if param_period:
        _write_scalar_named(wb, "PARAM_PERIOD_LABEL", param_period)
    labels = [
        ("total_revenue", "Total income"),
        ("total_expenses", "Total spending"),
        ("capital_net", "Capital (net)"),
        ("equity_net", "Equity (net)"),
        ("assets_net", "Assets (net)"),
        ("liabilities_net", "Liabilities (net)"),
    ]
    r = row0
    for key, lbl in labels:
        if key not in totals:
            continue
        ws.cell(row=r, column=c0, value=lbl)
        ws.cell(row=r, column=c0 + 1, value=float(totals.get(key, 0.0)))
        r += 1
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _safe_workbook_sheet_title(title: str) -> str:
    t = "".join(ch if ch not in r'\/:*?[]' else "_" for ch in (title or "GL").strip()) or "GL"
    return t[:31]


def general_ledger_sheet_exact_copy_to_xlsx_bytes(workbook_path: str, *, sheet_name: str) -> bytes:
    """
    Copy one worksheet from the master workbook into a new .xlsx byte-for-byte layout-wise:
    cell values, styles, merges, and basic row/column sizing — same grid as stored in Storage.
    """
    ext = Path(workbook_path).suffix.lower()
    if ext not in (".xlsx", ".xlsm"):
        raise ValueError(f"Expected .xlsx or .xlsm master workbook; got {ext!r}")

    wb_src = load_workbook(workbook_path, data_only=False, keep_vba=(ext == ".xlsm"))
    try:
        if sheet_name not in wb_src.sheetnames:
            raise KeyError(f"Sheet {sheet_name!r} not found; available: {wb_src.sheetnames}")
        src = wb_src[sheet_name]

        wb_dst = Workbook()
        d0 = wb_dst.active
        wb_dst.remove(d0)
        dst_title = _safe_workbook_sheet_title(sheet_name)
        dst = wb_dst.create_sheet(title=dst_title)

        mr = int(src.max_row or 1)
        mc = int(src.max_column or 1)

        for row in src.iter_rows(min_row=1, max_row=mr, min_col=1, max_col=mc):
            for cell in row:
                nv = dst.cell(row=cell.row, column=cell.column, value=cell.value)
                if getattr(cell, "has_style", False):
                    nv.font = copy(cell.font)
                    nv.border = copy(cell.border)
                    nv.fill = copy(cell.fill)
                    nv.number_format = cell.number_format
                    nv.protection = copy(cell.protection)
                    nv.alignment = copy(cell.alignment)

        for rng in list(src.merged_cells.ranges):
            dst.merge_cells(str(rng))

        for col_letter, dim in src.column_dimensions.items():
            if dim.width is not None:
                dst.column_dimensions[col_letter].width = dim.width
            if getattr(dim, "hidden", None):
                dst.column_dimensions[col_letter].hidden = dim.hidden

        for row_idx, dim in src.row_dimensions.items():
            if dim.height is not None:
                dst.row_dimensions[row_idx].height = dim.height
            if getattr(dim, "hidden", None):
                dst.row_dimensions[row_idx].hidden = dim.hidden

        out = BytesIO()
        wb_dst.save(out)
        return out.getvalue()
    finally:
        wb_src.close()


# --- Styled TB / BS (no admin Excel template required) --------------------------------

_TBL_HEADER_FILL = "B4C6E7"
_TBL_HEADER_FG = "1F497D"
_TBL_BORDER_BLUE = "1F497D"
_TBL_NUM_FMT = "#,##0.00"


def _fmt_amt_tb_cell(v: float, *, tol: float = 1e-12) -> str:
    """Trial balance amount: blank near zero; negatives in parentheses in credit column."""
    x = pd.to_numeric(v, errors="coerce")
    fv = 0.0 if pd.isna(x) else float(x)
    if abs(fv) <= tol:
        return "-"
    if fv < 0:
        return f"({abs(fv):,.2f})"
    return f"{fv:,.2f}"


def _tb_display_line_items(tb: pd.DataFrame | None) -> list[tuple[str, str, float, float]]:
    """Sorted Finjan-style rows: (section, account, debits, credits). Uses net balance columns."""
    if tb is None or tb.empty:
        return []
    work = tb
    if "section" not in work.columns or "category" not in work.columns:
        work = gla.trial_balance_for_display(work)
    items: list[tuple[str, str, str, float, float, int, str]] = []
    cat_rank = {c: i for i, c in enumerate(gla.TB_CATEGORY_ORDER)}
    for _, rec in work.iterrows():
        acct = str(rec.get("account") or "").strip()
        section = str(rec.get("section") or rec.get("category") or "").strip()
        net = pd.to_numeric(rec.get("net_balance", 0), errors="coerce")
        net_f = 0.0 if pd.isna(net) else float(net)
        deb, cre = gla.net_balance_to_dr_cr(net_f)
        if "debits" in rec and "credits" in rec and "net_balance" not in rec:
            deb = float(pd.to_numeric(rec.get("debits", 0), errors="coerce") or 0)
            cre = float(pd.to_numeric(rec.get("credits", 0), errors="coerce") or 0)
        cat = str(rec.get("category") or "")
        items.append((section, acct, deb, cre, cat_rank.get(cat, 99), acct.casefold()))
    items.sort(key=lambda it: (it[4], it[5]))
    return [(s, a, d, c) for s, a, d, c, _r, _k in items]


def _tb_sorted_line_items(tb: pd.DataFrame | None) -> list[tuple[str, str, float, float]]:
    """Sorted rows: (account_code, title, debits, credits) — legacy numeric-code layout."""
    lines = _tb_display_line_items(tb)
    out: list[tuple[str, str, float, float]] = []
    for _sec, acct, deb, cre in lines:
        code, title = gla.split_account_code_and_title(acct)
        out.append((code, title if title else acct, deb, cre))
    return out


def _xlsx_apply_tb_heading_style(cell) -> None:
    ln = Side(style="thin", color=_TBL_BORDER_BLUE)
    cell.font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
    cell.fill = PatternFill(fill_type="solid", start_color=_TBL_HEADER_FILL, end_color=_TBL_HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.border = Border(left=ln, right=ln, top=ln, bottom=ln)


def _xlsx_apply_tb_body_style(cell, *, h_align: str = "left") -> None:
    ln = Side(style="thin", color=_TBL_BORDER_BLUE)
    cell.font = Font(color=_TBL_HEADER_FG, size=11)
    cell.alignment = Alignment(horizontal=h_align, vertical="center")
    cell.border = Border(left=ln, right=ln, top=ln, bottom=Side(style="thin", color="C5D9F1"))


def _xlsx_write_amount(ws, row: int, col: int, value: float, *, tol: float = 1e-12, bold: bool = False) -> None:
    c = ws.cell(row=row, column=col)
    thin_top = Side(style="thin", color="C5D9F1")
    ln_lr = Side(style="thin", color=_TBL_BORDER_BLUE)
    if abs(float(value)) <= tol:
        c.value = None
        c.border = Border(left=ln_lr, right=ln_lr, top=thin_top, bottom=thin_top)
        c.font = Font(bold=bold, color=_TBL_HEADER_FG, size=11)
        c.alignment = Alignment(horizontal="right", vertical="center")
        return
    c.value = float(value)
    c.number_format = _TBL_NUM_FMT
    c.font = Font(bold=bold, color=_TBL_HEADER_FG, size=11)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = Border(left=ln_lr, right=ln_lr, top=thin_top, bottom=thin_top)


def trial_balance_formatted_xlsx_bytes(
    tb: pd.DataFrame,
    *,
    title: str = "Trial Balance",
    as_of: str | None = None,
    sheet_title: str = "Trial Balance",
) -> bytes:
    """Standalone workbook: Finjan-style category + account + net DR/CR columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = str(sheet_title)[:31] or "TB"

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    tcell = ws.cell(row=r, column=1, value=str(title))
    tcell.font = Font(bold=True, color=_TBL_HEADER_FG, size=14)
    tcell.alignment = Alignment(horizontal="center")
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    subt = ws.cell(row=r, column=1, value="Trial balance")
    subt.font = Font(color=_TBL_HEADER_FG, size=11)
    subt.alignment = Alignment(horizontal="center")
    r += 1
    if as_of:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ac = ws.cell(row=r, column=1, value=str(as_of))
        ac.font = Font(color=_TBL_HEADER_FG, size=10)
        ac.alignment = Alignment(horizontal="center")
        r += 1
    r += 1

    for cix, hdr in enumerate(("", "Account", "DR", "CR"), start=1):
        _xlsx_apply_tb_heading_style(ws.cell(row=r, column=cix, value=hdr))
    r += 1

    lines = _tb_display_line_items(tb)
    tot_d = sum(x[2] for x in lines)
    tot_c = sum(x[3] for x in lines)

    tol = 1e-9
    prev_sec = ""
    for sec, acct, deb, cre in lines:
        sec_out = sec if sec != prev_sec else ""
        if sec:
            prev_sec = sec
        ws.cell(row=r, column=1, value=sec_out or None)
        _xlsx_apply_tb_body_style(ws.cell(row=r, column=1), h_align="left")
        ws.cell(row=r, column=2, value=str(acct))
        _xlsx_apply_tb_body_style(ws.cell(row=r, column=2), h_align="left")
        _xlsx_write_amount(ws, r, 3, deb if abs(deb) > tol else 0.0)
        _xlsx_write_amount(ws, r, 4, cre if abs(cre) > tol else 0.0)
        r += 1

    thick = Side(style="medium", color=_TBL_BORDER_BLUE)
    thin_lr = Side(style="thin", color=_TBL_BORDER_BLUE)
    tr = r
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)
    tmerge = ws.cell(row=tr, column=1, value="Total")
    tmerge.font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
    tmerge.alignment = Alignment(horizontal="center", vertical="center")
    tmerge.border = Border(top=thick, bottom=thick, left=thin_lr, right=thin_lr)
    ws.cell(row=tr, column=2).border = Border(top=thick, bottom=thick, left=thin_lr, right=thin_lr)
    wc3 = ws.cell(row=tr, column=3)
    wc3.value = float(tot_d)
    wc3.number_format = _TBL_NUM_FMT
    wc3.font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
    wc3.alignment = Alignment(horizontal="right", vertical="center")
    wc3.border = Border(top=thick, bottom=thick, left=thin_lr, right=thin_lr)
    wc4 = ws.cell(row=tr, column=4)
    wc4.value = float(tot_c)
    wc4.number_format = _TBL_NUM_FMT
    wc4.font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
    wc4.alignment = Alignment(horizontal="right", vertical="center")
    wc4.border = Border(top=thick, bottom=thick, left=thin_lr, right=thin_lr)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _line_label_plain(code: str, tit: str) -> str:
    t = (tit or "").strip()
    c = str(code or "").strip()
    if c.isdigit() and t:
        return f"{c} — {t}"
    if t:
        return t
    return c


def _bs_classic_detail_amount(grp_key: str, deb: float, cre: float, tol: float = 1e-12) -> float:
    if grp_key == "Asset":
        return float(deb) - float(cre)
    return float(cre) - float(deb)


_BS_PERIOD_NET_EPS = 1e-9
_BS_RETAINED_BOOK_LABEL = "Retained earnings — workbook"


def _balance_sheet_should_show_period_net_line(net: float, *, tol: float = _BS_PERIOD_NET_EPS) -> bool:
    return abs(net) > tol


def _balance_sheet_period_net_equity_line_label(net: float) -> str:
    """Net income ⇒ retained earnings; net loss ⇒ accumulated deficit (same signed amount convention)."""
    return "Accumulated deficit" if net < 0 else "Retained earnings"


def _equity_extra_line_count(retained_book: float | None, period_net: float | None) -> int:
    n = 0
    if retained_book is not None:
        n += 1
    if period_net is not None and _balance_sheet_should_show_period_net_line(period_net):
        n += 1
    return n


def _equity_section_has_supplemental_lines(retained_book: float | None, period_net: float | None, *, tol: float = _BS_PERIOD_NET_EPS) -> bool:
    if retained_book is not None:
        return True
    return period_net is not None and abs(period_net) > tol


def _balance_sheet_has_display_rows(
    bs_groups: dict[str, pd.DataFrame],
    *,
    retained_earnings_excel: float | None,
    period_net_income_for_equity: float | None = None,
) -> bool:
    for key in ("Asset", "Liability", "Equity"):
        df = bs_groups.get(key)
        if isinstance(df, pd.DataFrame) and not df.empty:
            return True
    if retained_earnings_excel is not None:
        return True
    return (
        period_net_income_for_equity is not None
        and abs(period_net_income_for_equity) > _BS_PERIOD_NET_EPS
    )


def _stmt_nan() -> float:
    return float("nan")


def statement_df_display_for_streamlit(df: pd.DataFrame) -> pd.DataFrame:
    """
    Format ``amount`` / ``section_total`` as strings so Streamlit shows truly empty cells.

    Nullable numeric columns and stylers often surface as the literal text «None» in the grid.
    """
    if df.empty:
        return df
    out = df.copy()

    def _fmt_cell(v: Any) -> str:
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except TypeError:
            return ""
        try:
            fv = float(v)
        except (TypeError, ValueError):
            return ""
        if fv != fv:
            return ""
        return f"{fv:.2f}"

    for col in ("amount", "section_total"):
        if col not in out.columns:
            continue
        out[col] = out[col].map(_fmt_cell)
    return out


def balance_sheet_streamlit_df(
    bs_groups: dict[str, pd.DataFrame],
    *,
    retained_earnings_excel_amount: float | None = None,
    period_net_income_for_equity: float | None = None,
) -> pd.DataFrame:
    """Single table for ``st.dataframe`` (same interaction model as trial balance)."""
    empty = pd.DataFrame(columns=["account", "amount", "section_total"])
    if not _balance_sheet_has_display_rows(
        bs_groups,
        retained_earnings_excel=retained_earnings_excel_amount,
        period_net_income_for_equity=period_net_income_for_equity,
    ):
        return empty

    rows: list[dict[str, Any]] = []
    na = _stmt_nan()

    total_liab = 0.0
    total_eq = 0.0
    total_ast = 0.0
    tol = _BS_PERIOD_NET_EPS

    for sec_upper, grp_key, total_lbl in (
        ("Assets", "Asset", "Total assets"),
        ("Liabilities", "Liability", "Total liabilities"),
        ("Equity", "Equity", "Total equity"),
    ):
        sub = bs_groups.get(grp_key)
        if grp_key != "Equity" and (sub is None or sub.empty):
            continue
        if grp_key == "Equity" and (sub is None or sub.empty) and not _equity_section_has_supplemental_lines(
            retained_earnings_excel_amount,
            period_net_income_for_equity,
            tol=tol,
        ):
            continue
        rows.append({"account": str(sec_upper).upper(), "amount": na, "section_total": na})
        lines = _tb_sorted_line_items(sub)
        sec_sum = 0.0
        for code, tit, deb, cre in lines:
            amt = _bs_classic_detail_amount(grp_key, deb, cre)
            sec_sum += amt
            rows.append({"account": _line_label_plain(code, tit), "amount": amt, "section_total": na})
        if grp_key == "Equity" and retained_earnings_excel_amount is not None:
            re_amt = float(retained_earnings_excel_amount)
            sec_sum += re_amt
            rows.append({"account": _BS_RETAINED_BOOK_LABEL, "amount": re_amt, "section_total": na})
        if grp_key == "Equity" and period_net_income_for_equity is not None and _balance_sheet_should_show_period_net_line(
            float(period_net_income_for_equity)
        ):
            pn = float(period_net_income_for_equity)
            lab = _balance_sheet_period_net_equity_line_label(pn)
            sec_sum += pn
            rows.append({"account": lab, "amount": pn, "section_total": na})
        rows.append({"account": str(total_lbl).title(), "amount": na, "section_total": sec_sum})
        if grp_key == "Asset":
            total_ast = sec_sum
        elif grp_key == "Liability":
            total_liab = sec_sum
        elif grp_key == "Equity":
            total_eq = sec_sum

    liq = total_liab + total_eq
    rows.append({"account": "Total liabilities and equity", "amount": na, "section_total": liq})

    return pd.DataFrame(rows)


def _is_line_amount_revenue(deb: float, cre: float) -> float:
    return float(cre) - float(deb)


def _is_line_amount_expense(deb: float, cre: float) -> float:
    return float(deb) - float(cre)


def income_statement_streamlit_df(
    is_groups: dict[str, pd.DataFrame],
    *,
    total_revenue: float | None = None,
    total_expenses: float | None = None,
) -> pd.DataFrame:
    """Single P&L-style table for ``st.dataframe`` (aligned with TB column config)."""
    empty = pd.DataFrame(columns=["account", "amount", "section_total"])
    rev = is_groups.get("Revenue")
    exp = is_groups.get("Expense")
    has_lines = (isinstance(rev, pd.DataFrame) and not rev.empty) or (
        isinstance(exp, pd.DataFrame) and not exp.empty
    )
    if not has_lines:
        return empty

    rows: list[dict[str, Any]] = []
    na = _stmt_nan()
    tr_eff: float | None = float(total_revenue) if total_revenue is not None else None
    te_eff: float | None = float(total_expenses) if total_expenses is not None else None

    if isinstance(rev, pd.DataFrame) and not rev.empty:
        rows.append({"account": "REVENUE", "amount": na, "section_total": na})
        lines = _tb_sorted_line_items(rev)
        rsum = 0.0
        for code, tit, deb, cre in lines:
            amt = _is_line_amount_revenue(deb, cre)
            rsum += amt
            rows.append({"account": _line_label_plain(code, tit), "amount": amt, "section_total": na})
        if tr_eff is None:
            tr_eff = rsum
        rows.append({"account": "Total revenue", "amount": na, "section_total": float(tr_eff)})

    if isinstance(exp, pd.DataFrame) and not exp.empty:
        rows.append({"account": "EXPENSES", "amount": na, "section_total": na})
        lines = _tb_sorted_line_items(exp)
        esum = 0.0
        for code, tit, deb, cre in lines:
            amt = _is_line_amount_expense(deb, cre)
            esum += amt
            rows.append({"account": _line_label_plain(code, tit), "amount": amt, "section_total": na})
        if te_eff is None:
            te_eff = esum
        rows.append({"account": "Total expenses", "amount": na, "section_total": float(te_eff)})

    tr_fin = float(tr_eff) if tr_eff is not None else 0.0
    te_fin = float(te_eff) if te_eff is not None else 0.0
    rows.append({"account": "Net income", "amount": na, "section_total": tr_fin - te_fin})
    return pd.DataFrame(rows)


def _statement_account_row_is_emphasis(account_val: Any) -> bool:
    """Bold Streamlit rows for section headings and totals (balance sheet / income statement)."""
    s = str(account_val).strip()
    if not s:
        return False
    upper = s.upper()
    if upper in frozenset({"ASSETS", "LIABILITIES", "EQUITY", "REVENUE", "EXPENSES"}):
        return True
    low = s.lower()
    if low == "net income":
        return True
    if low == "net loss":
        return True
    if low == "accumulated deficit":
        return True
    if low == "total liabilities and equity":
        return True
    return low.startswith("total ")


def style_statement_streamlit(df: pd.DataFrame):
    """Return a pandas ``Styler`` so category headings and totals render bold (matches TB widget family)."""
    if df.empty or "account" not in df.columns:
        return df
    df_disp = statement_df_display_for_streamlit(df)
    bold_mask = df_disp["account"].map(_statement_account_row_is_emphasis)

    def _bold_row(row: pd.Series) -> list[str]:
        if bool(bold_mask.loc[row.name]):
            return ["font-weight: bold"] * len(row)
        return [""] * len(row)

    return df_disp.style.apply(_bold_row, axis=1)


def balance_sheet_formatted_xlsx_bytes(
    bs_groups: dict[str, pd.DataFrame],
    *,
    title: str = "Balance Sheet",
    as_of: str | None = None,
    retained_earnings_excel_amount: float | None = None,
    period_net_income_for_equity: float | None = None,
) -> bytes:
    """Classic statement layout: description | detail amount | totals column (openpyxl)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    thin_lr = Side(style="thin", color=_TBL_BORDER_BLUE)
    thin_bt = Side(style="thin", color="C5D9F1")
    sec_fill = PatternFill(fill_type="solid", start_color="DCE6F2", end_color="DCE6F2")
    tot_fill = PatternFill(fill_type="solid", start_color="E9F0FA", end_color="E9F0FA")
    grand_fill = PatternFill(fill_type="solid", start_color=_TBL_HEADER_FILL, end_color=_TBL_HEADER_FILL)
    tol = 1e-9

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    tcell = ws.cell(row=r, column=1, value=str(title).upper())
    tcell.font = Font(bold=True, color=_TBL_HEADER_FG, size=14)
    tcell.alignment = Alignment(horizontal="left", vertical="center")
    r += 1
    if as_of:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        aa = ws.cell(row=r, column=1, value=str(as_of))
        aa.font = Font(color=_TBL_HEADER_FG, size=10)
        aa.alignment = Alignment(horizontal="left", vertical="center")
        r += 1
    r += 1

    if not _balance_sheet_has_display_rows(
        bs_groups,
        retained_earnings_excel=retained_earnings_excel_amount,
        period_net_income_for_equity=period_net_income_for_equity,
    ):
        ws.cell(row=r, column=1, value="No accounts classified as assets, liabilities, or equity in this range.")
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    total_liab = 0.0
    total_eq = 0.0
    total_ast = 0.0

    for sec_upper, grp_key, total_lbl in (
        ("ASSETS", "Asset", "Total assets"),
        ("LIABILITIES", "Liability", "Total liabilities"),
        ("EQUITY", "Equity", "Total equity"),
    ):
        sub = bs_groups.get(grp_key)
        if grp_key != "Equity" and (sub is None or sub.empty):
            continue
        if grp_key == "Equity" and (sub is None or sub.empty) and not _equity_section_has_supplemental_lines(
            retained_earnings_excel_amount,
            period_net_income_for_equity,
            tol=tol,
        ):
            continue
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        sh = ws.cell(row=r, column=1, value=str(sec_upper))
        sh.font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
        sh.fill = sec_fill
        sh.border = Border(left=thin_lr, right=thin_lr, top=thin_lr, bottom=thin_lr)
        r += 1

        lines = _tb_sorted_line_items(sub)
        sec_sum = 0.0
        all_lines_len = len(lines) + _equity_extra_line_count(retained_earnings_excel_amount, period_net_income_for_equity)
        for li, (code, tit, deb, cre) in enumerate(lines):
            amt = _bs_classic_detail_amount(grp_key, deb, cre)
            sec_sum += amt
            lab = _line_label_plain(code, tit)
            if not str(lab).strip():
                lab = f"{code} — {tit}" if str(code).isdigit() and (tit or "").strip() else ((tit or code) or "")
            c1 = ws.cell(row=r, column=1, value=str(lab))
            c1.font = Font(color=_TBL_HEADER_FG, size=11)
            c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c1.border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            c2 = ws.cell(row=r, column=2)
            c2.value = float(amt)
            c2.number_format = _TBL_NUM_FMT
            c2.font = Font(color=_TBL_HEADER_FG, size=11)
            c2.alignment = Alignment(horizontal="right", vertical="center")
            last_detail = li == all_lines_len - 1
            if last_detail:
                c2.border = Border(
                    left=thin_lr,
                    right=thin_lr,
                    top=thin_bt,
                    bottom=Side(style="medium", color=_TBL_BORDER_BLUE),
                )
            else:
                c2.border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            c3 = ws.cell(row=r, column=3)
            c3.border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            r += 1

        if grp_key == "Equity" and retained_earnings_excel_amount is not None:
            re_amt = float(retained_earnings_excel_amount)
            sec_sum += re_amt
            lc1 = ws.cell(row=r, column=1, value=_BS_RETAINED_BOOK_LABEL)
            lc1.font = Font(color=_TBL_HEADER_FG, size=11)
            lc1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            lc1.border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            c2re = ws.cell(row=r, column=2)
            if abs(re_amt) > tol:
                c2re.value = float(re_amt)
                c2re.number_format = _TBL_NUM_FMT
            c2re.font = Font(color=_TBL_HEADER_FG, size=11)
            c2re.alignment = Alignment(horizontal="right", vertical="center")
            last_supp = not (
                period_net_income_for_equity is not None
                and _balance_sheet_should_show_period_net_line(float(period_net_income_for_equity))
            )
            c2re.border = (
                Border(
                    left=thin_lr,
                    right=thin_lr,
                    top=thin_bt,
                    bottom=Side(style="medium", color=_TBL_BORDER_BLUE),
                )
                if last_supp
                else Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            )
            ws.cell(row=r, column=3).border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            r += 1

        if (
            grp_key == "Equity"
            and period_net_income_for_equity is not None
            and _balance_sheet_should_show_period_net_line(float(period_net_income_for_equity))
        ):
            pn = float(period_net_income_for_equity)
            sec_sum += pn
            plab = _balance_sheet_period_net_equity_line_label(pn)
            lc1p = ws.cell(row=r, column=1, value=plab)
            lc1p.font = Font(color=_TBL_HEADER_FG, size=11)
            lc1p.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            lc1p.border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            c2pn = ws.cell(row=r, column=2)
            if abs(pn) > tol:
                c2pn.value = float(pn)
                c2pn.number_format = _TBL_NUM_FMT
            c2pn.font = Font(color=_TBL_HEADER_FG, size=11)
            c2pn.alignment = Alignment(horizontal="right", vertical="center")
            c2pn.border = Border(
                left=thin_lr,
                right=thin_lr,
                top=thin_bt,
                bottom=Side(style="medium", color=_TBL_BORDER_BLUE),
            )
            ws.cell(row=r, column=3).border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            r += 1

        if grp_key == "Liability":
            total_liab = sec_sum
        elif grp_key == "Equity":
            total_eq = sec_sum
        elif grp_key == "Asset":
            total_ast = sec_sum

        ws.cell(row=r, column=1, value=str(total_lbl).title())
        ws.cell(row=r, column=1).font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(row=r, column=1).border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
        ws.cell(row=r, column=1).fill = tot_fill
        ws.cell(row=r, column=2).border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
        ws.cell(row=r, column=2).fill = tot_fill
        ctot = ws.cell(row=r, column=3, value=float(sec_sum))
        ctot.number_format = _TBL_NUM_FMT
        ctot.font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
        ctot.alignment = Alignment(horizontal="right", vertical="center")
        ctot.border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
        ctot.fill = tot_fill
        r += 1

    thick = Side(style="medium", color=_TBL_BORDER_BLUE)
    liq_tot = total_liab + total_eq
    ws.cell(row=r, column=1, value="Total liabilities and equity")
    ws.cell(row=r, column=1).font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=r, column=1).border = Border(left=thin_lr, right=thin_lr, top=thick, bottom=thick)
    ws.cell(row=r, column=1).fill = grand_fill
    ws.cell(row=r, column=2).border = Border(left=thin_lr, right=thin_lr, top=thick, bottom=thick)
    ws.cell(row=r, column=2).fill = grand_fill
    gcell = ws.cell(row=r, column=3, value=float(liq_tot))
    gcell.number_format = _TBL_NUM_FMT
    gcell.font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
    gcell.alignment = Alignment(horizontal="right", vertical="center")
    gcell.border = Border(left=thin_lr, right=thin_lr, top=thick, bottom=thick)
    gcell.fill = grand_fill
    r += 1

    has_assets = isinstance(bs_groups.get("Asset"), pd.DataFrame) and not bs_groups["Asset"].empty
    if has_assets and abs(total_ast - liq_tot) > 1e-6:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        note = ws.cell(
            row=r,
            column=1,
            value=f"Accounting equation check: assets − (liabilities + equity) = {total_ast - liq_tot:.2f} (expect 0).",
        )
        note.font = Font(italic=True, color=_TBL_HEADER_FG, size=9)

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def income_statement_formatted_xlsx_bytes(
    is_groups: dict[str, pd.DataFrame],
    *,
    title: str = "Income Statement",
    period_label: str | None = None,
    total_revenue: float | None = None,
    total_expenses: float | None = None,
) -> bytes:
    """Classic income statement layout (revenue, expenses, net income)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Stmt"

    thin_lr = Side(style="thin", color=_TBL_BORDER_BLUE)
    thin_bt = Side(style="thin", color="C5D9F1")
    sec_fill = PatternFill(fill_type="solid", start_color="DCE6F2", end_color="DCE6F2")
    tot_fill = PatternFill(fill_type="solid", start_color="E9F0FA", end_color="E9F0FA")
    grand_fill = PatternFill(fill_type="solid", start_color=_TBL_HEADER_FILL, end_color=_TBL_HEADER_FILL)
    tol = 1e-9

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(row=r, column=1, value=str(title).upper()).font = Font(bold=True, color=_TBL_HEADER_FG, size=14)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    r += 1
    if period_label:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.cell(row=r, column=1, value=str(period_label)).font = Font(color=_TBL_HEADER_FG, size=10)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
        r += 1
    r += 1

    rev = is_groups.get("Revenue")
    exp = is_groups.get("Expense")
    has_any = (isinstance(rev, pd.DataFrame) and not rev.empty) or (
        isinstance(exp, pd.DataFrame) and not exp.empty
    )
    if not has_any:
        ws.cell(row=r, column=1, value="No revenue or expense accounts in this range.")
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    tr = 0.0
    te = 0.0

    if isinstance(rev, pd.DataFrame) and not rev.empty:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.cell(row=r, column=1, value="REVENUE").font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
        ws.cell(row=r, column=1).fill = sec_fill
        ws.cell(row=r, column=1).border = Border(left=thin_lr, right=thin_lr, top=thin_lr, bottom=thin_lr)
        r += 1
        lines = _tb_sorted_line_items(rev)
        rsum = 0.0
        for li, (code, tit, deb, cre) in enumerate(lines):
            amt = _is_line_amount_revenue(deb, cre)
            rsum += amt
            lab = _line_label_plain(code, tit)
            if not str(lab).strip():
                lab = f"{code} — {tit}" if str(code).isdigit() and (tit or "").strip() else ((tit or code) or "")
            ws.cell(row=r, column=1, value=str(lab))
            ws.cell(row=r, column=1).font = Font(color=_TBL_HEADER_FG, size=11)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=r, column=1).border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            c2 = ws.cell(row=r, column=2)
            if abs(amt) > tol:
                c2.value = float(amt)
                c2.number_format = _TBL_NUM_FMT
            c2.font = Font(color=_TBL_HEADER_FG, size=11)
            c2.alignment = Alignment(horizontal="right", vertical="center")
            if li == len(lines) - 1:
                c2.border = Border(
                    left=thin_lr,
                    right=thin_lr,
                    top=thin_bt,
                    bottom=Side(style="medium", color=_TBL_BORDER_BLUE),
                )
            else:
                c2.border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            ws.cell(row=r, column=3).border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            r += 1
        tr = float(total_revenue) if total_revenue is not None else rsum
        ws.cell(row=r, column=1, value="Total revenue").font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(row=r, column=1).border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
        ws.cell(row=r, column=1).fill = tot_fill
        ws.cell(row=r, column=2).border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
        ws.cell(row=r, column=2).fill = tot_fill
        ctot = ws.cell(row=r, column=3, value=float(tr))
        ctot.number_format = _TBL_NUM_FMT
        ctot.font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
        ctot.alignment = Alignment(horizontal="right", vertical="center")
        ctot.border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
        ctot.fill = tot_fill
        r += 1

    if isinstance(exp, pd.DataFrame) and not exp.empty:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.cell(row=r, column=1, value="EXPENSES").font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
        ws.cell(row=r, column=1).fill = sec_fill
        ws.cell(row=r, column=1).border = Border(left=thin_lr, right=thin_lr, top=thin_lr, bottom=thin_lr)
        r += 1
        lines = _tb_sorted_line_items(exp)
        esum = 0.0
        for li, (code, tit, deb, cre) in enumerate(lines):
            amt = _is_line_amount_expense(deb, cre)
            esum += amt
            lab = _line_label_plain(code, tit)
            if not str(lab).strip():
                lab = f"{code} — {tit}" if str(code).isdigit() and (tit or "").strip() else ((tit or code) or "")
            ws.cell(row=r, column=1, value=str(lab))
            ws.cell(row=r, column=1).font = Font(color=_TBL_HEADER_FG, size=11)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=r, column=1).border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            c2 = ws.cell(row=r, column=2)
            if abs(amt) > tol:
                c2.value = float(amt)
                c2.number_format = _TBL_NUM_FMT
            c2.font = Font(color=_TBL_HEADER_FG, size=11)
            c2.alignment = Alignment(horizontal="right", vertical="center")
            if li == len(lines) - 1:
                c2.border = Border(
                    left=thin_lr,
                    right=thin_lr,
                    top=thin_bt,
                    bottom=Side(style="medium", color=_TBL_BORDER_BLUE),
                )
            else:
                c2.border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            ws.cell(row=r, column=3).border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
            r += 1
        te = float(total_expenses) if total_expenses is not None else esum
        ws.cell(row=r, column=1, value="Total expenses").font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(row=r, column=1).border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
        ws.cell(row=r, column=1).fill = tot_fill
        ws.cell(row=r, column=2).border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
        ws.cell(row=r, column=2).fill = tot_fill
        ctot = ws.cell(row=r, column=3, value=float(te))
        ctot.number_format = _TBL_NUM_FMT
        ctot.font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
        ctot.alignment = Alignment(horizontal="right", vertical="center")
        ctot.border = Border(left=thin_lr, right=thin_lr, top=thin_bt, bottom=thin_bt)
        ctot.fill = tot_fill
        r += 1

    if total_revenue is not None:
        tr = float(total_revenue)
    if total_expenses is not None:
        te = float(total_expenses)

    net = float(tr) - float(te)
    thick = Side(style="medium", color=_TBL_BORDER_BLUE)
    net_lbl_row = "Net loss" if net < -1e-9 else "Net income"
    ws.cell(row=r, column=1, value=net_lbl_row).font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=r, column=1).border = Border(left=thin_lr, right=thin_lr, top=thick, bottom=thick)
    ws.cell(row=r, column=1).fill = grand_fill
    ws.cell(row=r, column=2).border = Border(left=thin_lr, right=thin_lr, top=thick, bottom=thick)
    ws.cell(row=r, column=2).fill = grand_fill
    gcell = ws.cell(row=r, column=3, value=float(net))
    gcell.number_format = _TBL_NUM_FMT
    gcell.font = Font(bold=True, color=_TBL_HEADER_FG, size=11)
    gcell.alignment = Alignment(horizontal="right", vertical="center")
    gcell.border = Border(left=thin_lr, right=thin_lr, top=thick, bottom=thick)
    gcell.fill = grand_fill

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _fmt_amt_plain(v: float, *, tol: float = 1e-12, blank_near_zero: bool = True) -> str:
    """Format a numeric cell. TB-style views omit near-zero; BS detail can show explicit 0.00."""
    x = pd.to_numeric(v, errors="coerce")
    fv = 0.0 if pd.isna(x) else float(x)
    if blank_near_zero and abs(fv) <= tol:
        return ""
    return f"{fv:,.2f}"


# --- Streamlit previews matching formatted Excel (colors, typography, fills) ---------


_HTML_CSS_TRIAL_BALANCE_XLS = """
<style>
.fin-tb-doc { font-family: "Segoe UI", system-ui, sans-serif; color: #1f497d; width: 100%; max-width: 960px; margin: 0 auto; }
.fin-tb-doc .ttl { font-size: 14pt; font-weight: 700; text-transform: uppercase; text-align: center; margin: 8px 0 4px; color: #1f497d; }
.fin-tb-doc .sub { font-size: 10pt; text-align: center; color: #1f497d; margin: 0 0 14px; }
.fin-tb-tbl { border-collapse: collapse; font-size: 11pt; color: #1f497d; width: 100%; }
.fin-tb-tbl th {
  background: #b4c6e7; font-weight: 700; text-transform: uppercase;
  padding: 9px 8px; border: 1px solid #1f497d; letter-spacing: 0.02em; text-align: center;
}
.fin-tb-tbl td {
  padding: 7px 8px; border-left: 1px solid #1f497d; border-right: 1px solid #1f497d;
  border-bottom: 1px solid #c5d9f1; vertical-align: middle;
}
.fin-tb-tbl td.nr { text-align: right; white-space: nowrap; }
.fin-tb-tbl td.nc { text-align: center; white-space: nowrap; }
.fin-tb-tbl tr.tb-total td {
  font-weight: 700; border-top: 2px solid #1f497d !important; border-bottom: 2px solid #1f497d !important;
}
</style>
"""


_HTML_CSS_SHEET3COL_XLS = """
<style>
.fin-xl3 { font-family: "Segoe UI", system-ui, sans-serif; color: #1f497d; width: 100%; max-width: 920px; margin: 0 auto; }
.fin-xl3 .ttl { font-size: 14pt; font-weight: 700; text-transform: uppercase; text-align: left; margin: 8px 0 4px; color: #1f497d; }
.fin-xl3 .sub { font-size: 10pt; text-align: left; color: #1f497d; margin: 0 0 14px; }
.fin-xl3-tbl { border-collapse: collapse; font-size: 11pt; width: 100%; color: #1f497d; }
.fin-xl3-tbl td {
  padding: 7px 8px; border-left: 1px solid #1f497d; border-right: 1px solid #1f497d;
  border-bottom: 1px solid #c5d9f1; vertical-align: middle;
}
.fin-xl3-tbl td.nr { text-align: right; white-space: nowrap; }
.fin-xl3-tbl td.sec {
  font-weight: 700; background: #dce6f2; border: 1px solid #1f497d;
}
.fin-xl3-tbl td.amt-strong-b { border-bottom: 2px solid #1f497d !important; }
.fin-xl3-tbl td.totl { font-weight: 700; background: #e9f0fa; padding-left: 18px;
  box-shadow: inset 1px 0 0 #1f497d, inset -1px 0 0 #1f497d;
}
.fin-xl3-tbl td.totmid { border-bottom: 1px solid #c5d9f1; background: #e9f0fa; border-left: 1px solid #1f497d;
  border-right: 1px solid #1f497d;
}
.fin-xl3-tbl td.totn { font-weight: 700; background: #e9f0fa; text-align: right;
  border-left: 1px solid #1f497d; border-right: 1px solid #1f497d; border-bottom: 1px solid #c5d9f1;
}
.fin-xl3-tbl td.glbl {
  font-weight: 700; background: #b4c6e7; border-top: 2px solid #1f497d !important;
  border-bottom: 2px solid #1f497d !important; border-left: 1px solid #1f497d !important;
  border-right: 1px solid #1f497d !important;
}
.fin-xl3-tbl td.gmid {
  background: #b4c6e7; border-top: 2px solid #1f497d !important;
  border-bottom: 2px solid #1f497d !important; border-left: 1px solid #1f497d !important;
  border-right: 1px solid #1f497d !important;
}
.fin-xl3-tbl td.gn {
  font-weight: 700; background: #b4c6e7; text-align: right;
  border-top: 2px solid #1f497d !important; border-bottom: 2px solid #1f497d !important;
  border-left: 1px solid #1f497d !important; border-right: 1px solid #1f497d !important;
}
</style>
"""


def trial_balance_streamlit_html(
    tb: pd.DataFrame,
    *,
    title: str = "Trial Balance",
    as_of: str | None = None,
) -> tuple[str, int]:
    """Finjan-style trial balance: category labels, account, net DR / CR columns."""
    org_title = tr(title)
    as_of_s = escape(tr(str(as_of).strip())) if (as_of and str(as_of).strip()) else ""
    lines = _tb_display_line_items(tb)
    tot_d = sum(x[2] for x in lines)
    tot_c = sum(x[3] for x in lines)
    thead = (
        "<tr><th></th><th>"
        + escape(tr("Account"))
        + "</th><th>"
        + escape(tr("DR"))
        + "</th><th>"
        + escape(tr("CR"))
        + "</th></tr>"
    )
    tol = 1e-12
    trs: list[str] = []
    prev_sec = ""
    for sec, acct, deb, cre in lines:
        sec_out = tr(sec) if sec and sec != prev_sec else ""
        if sec:
            prev_sec = sec
        dcell = escape(_fmt_amt_tb_cell(deb, tol=tol))
        ccell = escape(_fmt_amt_tb_cell(cre, tol=tol))
        trs.append(
            f"<tr><td class='sec'>{escape(sec_out)}</td><td>{escape(acct)}</td>"
            f"<td class='nr'>{dcell}</td><td class='nr'>{ccell}</td></tr>"
        )
    totals_row = (
        "<tr class='tb-total'><td colspan='2' style='text-align:center'>"
        + escape(tr("Total"))
        + "</td>"
        f"<td class='nr'>{escape(_fmt_amt_plain(tot_d, tol=tol))}</td>"
        f"<td class='nr'>{escape(_fmt_amt_plain(tot_c, tol=tol))}</td></tr>"
    )
    tb_gap = tot_d - tot_c
    tb_rel = max(1e-6, 0.01 * max(abs(tot_d), abs(tot_c), 1.0))
    tb_footer = ""
    if abs(tb_gap) > tb_rel:
        gap_msg = (
            f"Trial balance check: total debits − total credits = {tb_gap:,.2f} (expect 0). "
            "Review compound journal entries in the workbook or unmapped GL lines."
        )
        tb_footer = (
            "<div class='sub' style='color:#92400e'>"
            + escape(tr(gap_msg))
            + "</div>"
        )
    sub_html = (
        "<div class='sub'>"
        + escape(tr("Trial balance"))
        + (f"<br>{as_of_s}" if as_of_s else "")
        + "</div>"
    )
    body = "".join(trs)
    html = (
        html_rtl_css()
        + _HTML_CSS_TRIAL_BALANCE_XLS
        + "<div class='fin-tb-doc'>"
        + f"<div class='ttl'>{escape(org_title)}</div>"
        + sub_html
        + "<table class='fin-tb-tbl'><thead>"
        + thead
        + "</thead><tbody>"
        + body
        + totals_row
        + "</tbody></table>"
        + tb_footer
        + "</div>"
    )
    h_px = min(1320, 140 + len(lines) * 30)
    return html, h_px


def balance_sheet_streamlit_html(
    bs_groups: dict[str, pd.DataFrame],
    *,
    title: str = "Balance Sheet",
    as_of: str | None = None,
    retained_earnings_excel_amount: float | None = None,
    period_net_income_for_equity: float | None = None,
) -> tuple[str, int]:
    """HTML matching ``balance_sheet_formatted_xlsx_bytes`` layout and styling."""
    title = tr(title)
    if not _balance_sheet_has_display_rows(
        bs_groups,
        retained_earnings_excel=retained_earnings_excel_amount,
        period_net_income_for_equity=period_net_income_for_equity,
    ):
        frag = (
            html_rtl_css()
            + _HTML_CSS_SHEET3COL_XLS
            + "<div class='fin-xl3'><p>"
            + escape(tr("No accounts classified as assets, liabilities, or equity in this range."))
            + "</p></div>"
        )
        return frag, 120

    tol = 1e-9
    total_liab = 0.0
    total_eq = 0.0
    total_ast = 0.0
    rows_out: list[str] = []

    ttl = escape(str(title).upper())
    sub = escape(tr(str(as_of).strip())) if (as_of and str(as_of).strip()) else ""
    sub_blk = f"<div class='sub'>{sub}</div>" if sub else ""

    for sec_upper, grp_key, total_lbl in (
        ("ASSETS", "Asset", "Total assets"),
        ("LIABILITIES", "Liability", "Total liabilities"),
        ("EQUITY", "Equity", "Total equity"),
    ):
        sub_df = bs_groups.get(grp_key)
        if grp_key != "Equity" and (sub_df is None or sub_df.empty):
            continue
        if grp_key == "Equity" and (sub_df is None or sub_df.empty) and not _equity_section_has_supplemental_lines(
            retained_earnings_excel_amount,
            period_net_income_for_equity,
            tol=_BS_PERIOD_NET_EPS,
        ):
            continue
        rows_out.append(
            "<tr><td colspan='3' class='sec'>" + escape(tr(sec_upper)) + "</td></tr>"
        )
        lines = _tb_sorted_line_items(sub_df)
        sec_sum = 0.0
        n_lines = len(lines) + _equity_extra_line_count(
            retained_earnings_excel_amount, period_net_income_for_equity
        )
        for li, (code, tit, deb, cre) in enumerate(lines):
            amt = _bs_classic_detail_amount(grp_key, deb, cre)
            sec_sum += amt
            lab = _line_label_plain(code, tit)
            if not str(lab).strip():
                lab = (
                    f"{code} — {tit}"
                    if str(code).isdigit() and (tit or "").strip()
                    else ((tit or code) or "")
                )
            last_detail = li == n_lines - 1
            amt_cell = escape(_fmt_amt_plain(amt, tol=tol, blank_near_zero=False))
            amt_cls = "nr" + (" amt-strong-b" if last_detail else "")
            rows_out.append(
                f"<tr><td>{escape(str(lab))}</td>"
                f"<td class=\"{amt_cls}\">{amt_cell}</td>"
                f"<td></td></tr>"
            )

        if grp_key == "Equity" and retained_earnings_excel_amount is not None:
            re_amt = float(retained_earnings_excel_amount)
            sec_sum += re_amt
            amt_cell = escape(_fmt_amt_plain(re_amt, tol=tol))
            last_book = not (
                period_net_income_for_equity is not None
                and _balance_sheet_should_show_period_net_line(float(period_net_income_for_equity))
            )
            book_cls = "nr" + (" amt-strong-b" if last_book else "")
            rows_out.append(
                f"<tr><td>{escape(tr(_BS_RETAINED_BOOK_LABEL))}</td>"
                f"<td class=\"{book_cls}\">{amt_cell}</td>"
                f"<td></td></tr>"
            )

        if (
            grp_key == "Equity"
            and period_net_income_for_equity is not None
            and _balance_sheet_should_show_period_net_line(float(period_net_income_for_equity))
        ):
            pn = float(period_net_income_for_equity)
            sec_sum += pn
            plab = tr(_balance_sheet_period_net_equity_line_label(pn))
            amt_cell = escape(_fmt_amt_plain(pn, tol=tol))
            rows_out.append(
                f"<tr><td>{escape(plab)}</td>"
                f"<td class=\"nr amt-strong-b\">{amt_cell}</td>"
                f"<td></td></tr>"
            )

        if grp_key == "Liability":
            total_liab = sec_sum
        elif grp_key == "Equity":
            total_eq = sec_sum
        elif grp_key == "Asset":
            total_ast = sec_sum

        tlab = escape(tr(total_lbl))
        totamt = escape(_fmt_amt_plain(sec_sum, tol=tol))
        rows_out.append(
            f"<tr><td class='totl'>{tlab}</td><td class='totmid'></td>"
            f"<td class='totn'>{totamt}</td></tr>"
        )

    liq_tot = total_liab + total_eq
    liq_esc = escape(_fmt_amt_plain(liq_tot, tol=tol))
    rows_out.append(
        "<tr><td class='glbl'>" + escape(tr("Total liabilities and equity")) + "</td>"
        "<td class='gmid'></td>"
        f"<td class='gn'>{liq_esc}</td></tr>"
    )

    footer = ""
    has_assets = isinstance(bs_groups.get("Asset"), pd.DataFrame) and not bs_groups["Asset"].empty
    if has_assets and abs(total_ast - liq_tot) > 1e-6:
        gap = total_ast - liq_tot
        footer = (
            "<div class='sub'>"
            + escape(
                tr(f"Accounting equation check: assets − (liabilities + equity) = {gap:,.2f} (expect 0).")
            )
            + "</div>"
        )

    n_body = len(rows_out)
    html = (
        html_rtl_css()
        + _HTML_CSS_SHEET3COL_XLS
        + "<div class='fin-xl3'>"
        + f"<div class='ttl'>{ttl}</div>"
        + sub_blk
        + "<table class='fin-xl3-tbl'><tbody>"
        + "".join(rows_out)
        + "</tbody></table>"
        + footer
        + "</div>"
    )
    h_px = min(1600, 120 + n_body * 30 + (40 if footer else 0))
    return html, h_px


def income_statement_streamlit_html(
    is_groups: dict[str, pd.DataFrame],
    *,
    title: str = "Income Statement",
    period_label: str | None = None,
    total_revenue: float | None = None,
    total_expenses: float | None = None,
) -> tuple[str, int]:
    """HTML matching ``income_statement_formatted_xlsx_bytes`` layout and styling."""
    title = tr(title)
    rev = is_groups.get("Revenue")
    exp = is_groups.get("Expense")
    has_any = (isinstance(rev, pd.DataFrame) and not rev.empty) or (
        isinstance(exp, pd.DataFrame) and not exp.empty
    )
    if not has_any:
        frag = (
            html_rtl_css()
            + _HTML_CSS_SHEET3COL_XLS
            + "<div class='fin-xl3'><p>"
            + escape(tr("No revenue or expense accounts in this range."))
            + "</p></div>"
        )
        return frag, 120

    tol = 1e-9
    tr_rev = 0.0
    te = 0.0
    rows_out: list[str] = []

    ttl = escape(str(title).upper())
    pl = escape(tr(str(period_label).strip())) if (period_label and str(period_label).strip()) else ""
    sub_blk = f"<div class='sub'>{pl}</div>" if pl else ""

    if isinstance(rev, pd.DataFrame) and not rev.empty:
        rows_out.append("<tr><td colspan='3' class='sec'>" + escape(tr("REVENUE")) + "</td></tr>")
        lines = _tb_sorted_line_items(rev)
        rsum = 0.0
        for li, (code, tit, deb, cre) in enumerate(lines):
            amt = _is_line_amount_revenue(deb, cre)
            rsum += amt
            lab = _line_label_plain(code, tit)
            if not str(lab).strip():
                lab = (
                    f"{code} — {tit}"
                    if str(code).isdigit() and (tit or "").strip()
                    else ((tit or code) or "")
                )
            last_detail = li == len(lines) - 1
            amt_cell = escape(_fmt_amt_plain(amt, tol=tol))
            amt_cls = "nr" + (" amt-strong-b" if last_detail else "")
            rows_out.append(
                f"<tr><td>{escape(str(lab))}</td><td class=\"{amt_cls}\">{amt_cell}</td><td></td></tr>"
            )
        tr_rev = float(total_revenue) if total_revenue is not None else rsum
        rows_out.append(
            "<tr><td class='totl'>" + escape(tr("Total revenue")) + "</td><td class='totmid'></td>"
            f"<td class='totn'>{escape(_fmt_amt_plain(tr_rev, tol=tol))}</td></tr>"
        )

    if isinstance(exp, pd.DataFrame) and not exp.empty:
        rows_out.append("<tr><td colspan='3' class='sec'>" + escape(tr("EXPENSES")) + "</td></tr>")
        lines = _tb_sorted_line_items(exp)
        esum = 0.0
        for li, (code, tit, deb, cre) in enumerate(lines):
            amt = _is_line_amount_expense(deb, cre)
            esum += amt
            lab = _line_label_plain(code, tit)
            if not str(lab).strip():
                lab = (
                    f"{code} — {tit}"
                    if str(code).isdigit() and (tit or "").strip()
                    else ((tit or code) or "")
                )
            last_detail = li == len(lines) - 1
            amt_cell = escape(_fmt_amt_plain(amt, tol=tol))
            amt_cls = "nr" + (" amt-strong-b" if last_detail else "")
            rows_out.append(
                f"<tr><td>{escape(str(lab))}</td><td class=\"{amt_cls}\">{amt_cell}</td><td></td></tr>"
            )
        te = float(total_expenses) if total_expenses is not None else esum
        rows_out.append(
            "<tr><td class='totl'>" + escape(tr("Total expenses")) + "</td><td class='totmid'></td>"
            f"<td class='totn'>{escape(_fmt_amt_plain(te, tol=tol))}</td></tr>"
        )

    if total_revenue is not None:
        tr_rev = float(total_revenue)
    if total_expenses is not None:
        te = float(total_expenses)
    net = float(tr_rev) - float(te)
    net_row_lbl = tr("Net loss") if net < -1e-9 else tr("Net income")

    rows_out.append(
        f"<tr><td class='glbl'>{escape(net_row_lbl)}</td><td class='gmid'></td>"
        f"<td class='gn'>{escape(_fmt_amt_plain(net, tol=tol))}</td></tr>"
    )

    n_body = len(rows_out)
    html = (
        html_rtl_css()
        + _HTML_CSS_SHEET3COL_XLS
        + "<div class='fin-xl3'>"
        + f"<div class='ttl'>{ttl}</div>"
        + sub_blk
        + "<table class='fin-xl3-tbl'><tbody>"
        + "".join(rows_out)
        + "</tbody></table></div>"
    )
    h_px = min(1600, 120 + n_body * 30)
    return html, h_px


def trial_balance_preview_html(
    tb: pd.DataFrame, *, title: str = "Trial Balance", as_of: str | None = None
) -> tuple[str, int]:
    """Backward-compatible alias for :func:`trial_balance_streamlit_html`."""
    return trial_balance_streamlit_html(tb, title=title, as_of=as_of)


def load_template_bytes(client: Any, secrets: dict, kind: str) -> tuple[Any | None, str | None]:
    """Fetch template from Storage using app_settings.statement_templates_json[kind].object_path."""
    import database as db
    import supabase_storage_documents as sbd

    doc = db.fetch_statement_templates_json(client)
    meta = doc.get(kind)
    if not isinstance(meta, dict):
        return None, f"No template uploaded for «{kind}». Add one in Settings."
    path = str(meta.get("object_path") or "").strip()
    if not path:
        return None, f"No template path for «{kind}»."
    bucket = sbd.documents_bucket(secrets)
    try:
        raw = sbd.download_document_bytes(client, bucket, path)
        return raw, None
    except Exception as e:
        return None, str(e)
